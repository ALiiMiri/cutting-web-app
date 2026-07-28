"""Excel export for a persisted grouped cutting order."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def create_cutting_order_workbook(order):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "نتایج برش گروهی"
    sheet.sheet_view.rightToLeft = True

    blue = PatternFill("solid", fgColor="4472C4")
    light_blue = PatternFill("solid", fgColor="D9EAF7")
    green = PatternFill("solid", fgColor="D9EAD3")
    grey = PatternFill("solid", fgColor="E7E6E6")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    sheet["A1"] = f"سفارش برش گروهی {order['order_number']}"
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].fill = blue
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A1:N1")
    sheet["A3"] = "وضعیت"
    sheet["B3"] = order["status_label"]
    sheet["D3"] = "نسخه"
    sheet["E3"] = order["version"]
    sheet["G3"] = "تعداد پروژه"
    sheet["H3"] = len(order["projects"])
    sheet["J3"] = "تعداد شاخه"
    sheet["K3"] = len(order["bars"])

    sheet["A5"] = "پروژه‌های این سفارش برش"
    sheet["A5"].font = Font(bold=True)
    project_headers = ["شناسه", "شماره سفارش", "کد پروژه", "نام مشتری"]
    for column, value in enumerate(project_headers, start=1):
        cell = sheet.cell(6, column, value)
        cell.fill = light_blue
        cell.font = Font(bold=True)
        cell.border = border
    for offset, project in enumerate(order["projects"], start=7):
        values = [
            project["project_id"],
            project["project_order_ref_snapshot"],
            project["project_code_snapshot"],
            project["project_name_snapshot"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(offset, column, value)
            cell.border = border

    row = 8 + len(order["projects"])
    headers = [
        "شاخه",
        "وضعیت",
        "پروفیل",
        "رنگ",
        "منبع",
        "طول اولیه (cm)",
        "قطعات و محل مصرف",
        "تعداد برش",
        "افت تیغ (cm)",
        "باقی‌مانده برنامه (cm)",
        "باقی‌مانده واقعی (cm)",
        "نتیجه باقی‌مانده",
        "زمان برش",
        "اپراتور",
    ]
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row, column, value)
        cell.fill = light_blue
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for bar in order["bars"]:
        row += 1
        pieces_text = "\n".join(
            (
                f"{piece['length']:g} cm | سفارش "
                f"{piece.get('project_order_ref_snapshot') or piece.get('project_id') or '—'}"
                f" | ردیف {piece.get('door_row_number') or '—'}"
                f" | {piece.get('door_location_snapshot') or 'بدون محل'}"
                f" | {piece['member_label']} | {piece['cut_instruction']}"
            )
            for piece in bar["pieces"]
        )
        if bar["source_type"] == "inventory_piece":
            source = f"قطعه انبار #{bar['source_inventory_piece_id']}"
        else:
            source = "شاخه کامل جدید"
        if bar["returned_piece_id"]:
            remainder_result = f"بازگشت به انبار؛ قطعه #{bar['returned_piece_id']}"
        elif bar["waste_item_id"]:
            remainder_result = f"ثبت ضایعات #{bar['waste_item_id']}"
        elif bar["status"] == "cut":
            remainder_result = "بدون باقی‌مانده"
        else:
            remainder_result = "هنوز برش نخورده"
        values = [
            bar["sequence_no"],
            bar["status_label"],
            bar["profile_name_snapshot"],
            bar["color_name_snapshot"],
            source,
            bar["initial_length"],
            pieces_text,
            len(bar["pieces"]),
            bar["kerf_loss"],
            bar["planned_remaining"],
            bar["actual_remaining"],
            remainder_result,
            bar["cut_at"],
            bar["cut_by_username"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if bar["status"] == "cut":
                cell.fill = green
            elif bar["status"] == "cancelled":
                cell.fill = grey

    widths = {
        "A": 10,
        "B": 18,
        "C": 24,
        "D": 16,
        "E": 23,
        "F": 18,
        "G": 75,
        "H": 13,
        "I": 16,
        "J": 23,
        "K": 23,
        "L": 30,
        "M": 21,
        "N": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = f"A{8 + len(order['projects']) + 1}"
    sheet.auto_filter.ref = f"A{8 + len(order['projects'])}:N{row}"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
