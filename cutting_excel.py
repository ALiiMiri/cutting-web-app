"""Excel presentation for a calculated cutting plan."""

import json

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from measurements import centimeters_to_measurement_unit, measurement_unit_labels


STRATEGY_LABELS = {
    "minimize_waste": "حداقل‌سازی ضایعات",
    "minimize_pieces": "حداقل‌سازی تعداد منابع",
    "minimize_new_profiles": "حداقل‌سازی شاخه‌های جدید",
}

STATUS_LABELS = {
    "discarded": "ضایعات واقعی",
    "reusable": "قابل بازگشت به انبار",
    "none": "بدون باقی‌مانده",
}


def create_cutting_plan_snapshot(plan):
    """Return only the applied-plan fields needed to reproduce its report."""
    return {
        "schema_version": 1,
        "processed_bins": plan["processed_bins"],
        "profile_summaries": plan["profile_summaries"],
        "stats": plan["stats"],
        "total_bins": plan["total_bins"],
        "blade_width": plan["blade_width"],
        "optimization_strategy": plan["optimization_strategy"],
    }


def resolve_applied_cutting_plan(application_status):
    """Return the saved plan and any warning required before a base-only export."""
    if application_status.get("status") != "completed":
        return None, (
            "هنوز کسر انبار از طریق محاسبه برش برای این سفارش انجام نشده است. "
            "اگر ادامه دهید، فایل Excel بدون شیت «نتایج برش» ساخته می‌شود."
        )

    raw_snapshot = (application_status.get("application") or {}).get(
        "plan_snapshot_json"
    )
    if not raw_snapshot:
        return None, (
            "کسر انبار این سفارش قبل از قابلیت ذخیره طرح برش انجام شده است؛ "
            "بنابراین شیت نتایج برش در فایل Excel نمایش داده نمی‌شود."
        )

    try:
        candidate = json.loads(raw_snapshot)
    except (TypeError, ValueError, json.JSONDecodeError):
        candidate = None
    if (
        isinstance(candidate, dict)
        and candidate.get("schema_version") == 1
        and isinstance(candidate.get("processed_bins"), list)
        and isinstance(candidate.get("profile_summaries"), list)
        and isinstance(candidate.get("stats"), dict)
    ):
        return candidate, None

    return None, (
        "کسر انبار برای این سفارش ثبت شده، اما جزئیات طرح آن برای ساخت "
        "شیت نتایج برش معتبر نیست."
    )


def add_cutting_results_sheet(workbook, plan, measurement_unit="cm"):
    """Add a complete, auditable cutting-results sheet and return it."""
    unit_labels = measurement_unit_labels(measurement_unit)

    def export_length(value):
        return centimeters_to_measurement_unit(value, measurement_unit)

    sheet = workbook.create_sheet("نتایج برش")
    sheet.sheet_view.rightToLeft = True

    title_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_fill = PatternFill(
        start_color="D9EAF7", end_color="D9EAF7", fill_type="solid"
    )
    discarded_fill = PatternFill(
        start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"
    )
    reusable_fill = PatternFill(
        start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    sheet["A1"] = "نتایج محاسبه برش و وزن باقی‌مانده"
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells("A1:L1")

    stats = plan["stats"]
    summary_rows = [
        ("تعداد کل شاخه‌ها/قطعات مبنا", plan["total_bins"], "عدد"),
        (
            "استراتژی انتخاب منابع",
            STRATEGY_LABELS[plan["optimization_strategy"]],
            "",
        ),
        (
            "ضخامت تیغ برش",
            round(export_length(plan["blade_width"]), 1),
            unit_labels["fa"],
        ),
        (
            "مجموع افت ناشی از تیغ",
            round(export_length(stats["total_kerf_length"]), 1),
            unit_labels["fa"],
        ),
        (
            "ضایعات واقعی",
            round(export_length(stats["discarded_length"]), 1),
            unit_labels["fa"],
        ),
        ("وزن ضایعات واقعی", round(stats["discarded_weight"], 2), "کیلوگرم"),
        (
            "باقی‌مانده قابل‌بازیافت",
            round(export_length(stats["reusable_length"]), 1),
            unit_labels["fa"],
        ),
        (
            "وزن باقی‌مانده قابل‌بازیافت",
            round(stats["reusable_weight"], 2),
            "کیلوگرم",
        ),
        (
            "کل باقی‌مانده",
            round(export_length(stats["total_remaining_length"]), 1),
            unit_labels["fa"],
        ),
        ("وزن کل باقی‌مانده", round(stats["total_remaining_weight"], 2), "کیلوگرم"),
        ("درصد طول باقی‌مانده", round(stats["total_remaining_percentage"], 1), "درصد"),
    ]
    for row_index, (label, value, unit) in enumerate(summary_rows, start=3):
        sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row_index, column=2, value=value)
        sheet.cell(row=row_index, column=3, value=unit)

    table_row = len(summary_rows) + 5
    headers = [
        "شاخه",
        "نوع پروفیل",
        "رنگ",
        "وزن هر متر (kg)",
        "منبع",
        f"طول اولیه ({unit_labels['short']})",
        f"قطعات برش ({unit_labels['short']})",
        "تعداد برش",
        f"افت تیغ ({unit_labels['short']})",
        f"باقی‌مانده ({unit_labels['short']})",
        "وزن باقی‌مانده (kg)",
        "وضعیت باقی‌مانده",
    ]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=table_row, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    for row_offset, bin_data in enumerate(plan["processed_bins"], start=1):
        row_index = table_row + row_offset
        values = [
            bin_data["index"],
            bin_data["profile_type"],
            bin_data["color_name"],
            bin_data["weight_per_meter"],
            (
                f"قطعه انبار، شناسه {bin_data['inventory_piece_id']}"
                if bin_data["from_inventory_piece"]
                else "شاخه جدید"
            ),
            export_length(bin_data["initial_length"]),
            " | ".join(
                f"{export_length(piece['length']):g}؛ {piece['member_label']}؛ "
                f"{piece['cut_instruction']}"
                for piece in bin_data["piece_details"]
            ),
            bin_data["cut_count"],
            export_length(bin_data["kerf_loss"]),
            export_length(bin_data["remaining"]),
            bin_data["remaining_weight"],
            STATUS_LABELS[bin_data["remaining_type"]],
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = thin_border
            if bin_data["remaining_type"] == "discarded":
                cell.fill = discarded_fill
            elif bin_data["remaining_type"] == "reusable":
                cell.fill = reusable_fill

    profile_row = table_row + len(plan["processed_bins"]) + 3
    sheet.cell(row=profile_row, column=1, value="خلاصه به تفکیک پروفیل و رنگ").font = Font(
        bold=True, size=13
    )
    profile_headers = [
        "پروفیل / رنگ",
        f"طول استاندارد شاخه ({unit_labels['short']})",
        "وزن هر متر",
        "تعداد شاخه/قطعه",
        f"طول ضایعات واقعی ({unit_labels['short']})",
        "وزن ضایعات واقعی",
        f"طول قابل‌بازیافت ({unit_labels['short']})",
        "وزن قابل‌بازیافت",
        f"کل طول باقی‌مانده ({unit_labels['short']})",
        "وزن کل باقی‌مانده",
    ]
    for column_index, header in enumerate(profile_headers, start=1):
        cell = sheet.cell(row=profile_row + 1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for row_offset, summary in enumerate(plan["profile_summaries"], start=2):
        values = [
            f"{summary['profile_type']} — {summary['color_name']}",
            round(export_length(summary["default_length"]), 1),
            round(summary["weight_per_meter"], 3),
            summary["bin_count"],
            round(export_length(summary["discarded_length"]), 1),
            round(summary["discarded_weight"], 2),
            round(export_length(summary["reusable_length"]), 1),
            round(summary["reusable_weight"], 2),
            round(export_length(summary["total_remaining_length"]), 1),
            round(summary["total_remaining_weight"], 2),
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=profile_row + row_offset, column=column_index, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 12,
        "B": 24,
        "C": 18,
        "D": 18,
        "E": 22,
        "F": 18,
        "G": 58,
        "H": 14,
        "I": 18,
        "J": 20,
        "K": 24,
        "L": 24,
    }
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    sheet.freeze_panes = f"A{table_row + 1}"
    return sheet
