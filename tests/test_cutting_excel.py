import io
import unittest

from openpyxl import Workbook, load_workbook

from cutting_calculator import calculate_cutting_plan
from cutting_excel import (
    add_cutting_results_sheet,
    create_cutting_plan_snapshot,
    resolve_applied_cutting_plan,
)


def _profile(default_length=500):
    return {
        "id": 1,
        "name": "پروفیل تست",
        "weight_per_meter": 1.5,
        "min_waste": 20,
        "default_length": default_length,
    }


def _door(door_id, color, frame_type):
    return {
        "id": door_id,
        "width": 100,
        "height": 200,
        "quantity": 1,
        "noe_profile": "پروفیل تست",
        "rang": color,
        "kolaft": frame_type,
    }


def _find_row(sheet, first_cell_value):
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == first_cell_value:
            return row
    raise AssertionError(f"Row not found: {first_cell_value}")


class CuttingExcelTests(unittest.TestCase):
    def test_unapplied_project_requires_warning_and_has_no_cutting_plan(self):
        snapshot, warning = resolve_applied_cutting_plan(
            {"status": "not_applied", "application": None}
        )

        self.assertIsNone(snapshot)
        self.assertIn("هنوز کسر انبار", warning)
        self.assertIn("بدون شیت «نتایج برش»", warning)

    def test_completed_project_uses_the_exact_saved_plan(self):
        plan = calculate_cutting_plan(
            [_door(1, "مشکی", "سه طرفه")],
            [_profile()],
        )
        expected = create_cutting_plan_snapshot(plan)
        import json

        snapshot, warning = resolve_applied_cutting_plan(
            {
                "status": "completed",
                "application": {
                    "plan_snapshot_json": json.dumps(expected, ensure_ascii=False)
                },
            }
        )

        self.assertIsNone(warning)
        self.assertEqual(snapshot, expected)

    def test_cutting_sheet_keeps_color_length_cut_type_and_kerf(self):
        plan = calculate_cutting_plan(
            [
                _door(1, "مشکی", "سه طرفه"),
                _door(2, "سفید", "دو طرفه"),
            ],
            [_profile()],
        )
        workbook = Workbook()
        sheet = add_cutting_results_sheet(workbook, plan, "cm")

        header_row = _find_row(sheet, "شاخه")
        headers = [sheet.cell(header_row, column).value for column in range(1, 13)]
        self.assertEqual(
            headers,
            [
                "شاخه",
                "نوع پروفیل",
                "رنگ",
                "وزن هر متر (kg)",
                "منبع",
                "طول اولیه (cm)",
                "قطعات برش (cm)",
                "تعداد برش",
                "افت تیغ (cm)",
                "باقی‌مانده (cm)",
                "وزن باقی‌مانده (kg)",
                "وضعیت باقی‌مانده",
            ],
        )

        data_rows = range(header_row + 1, header_row + 1 + plan["total_bins"])
        self.assertEqual(
            {sheet.cell(row, 3).value for row in data_rows}, {"مشکی", "سفید"}
        )
        self.assertTrue(all(sheet.cell(row, 6).value == 500 for row in data_rows))
        self.assertAlmostEqual(
            sum(sheet.cell(row, 9).value for row in data_rows),
            plan["stats"]["total_kerf_length"],
        )

        black_instructions = " ".join(
            sheet.cell(row, 7).value
            for row in data_rows
            if sheet.cell(row, 3).value == "مشکی"
        )
        white_instructions = " ".join(
            sheet.cell(row, 7).value
            for row in data_rows
            if sheet.cell(row, 3).value == "سفید"
        )
        self.assertIn("فارسی‌بُر", black_instructions)
        self.assertIn("بالا: صاف ۹۰ درجه؛ پایین: صاف ۹۰ درجه", white_instructions)
        self.assertNotIn("بالای چهارچوب", white_instructions)

        profile_header_row = _find_row(sheet, "پروفیل / رنگ")
        profile_labels = {
            sheet.cell(profile_header_row + offset, 1).value
            for offset in range(1, len(plan["profile_summaries"]) + 1)
        }
        self.assertEqual(
            profile_labels,
            {"پروفیل تست — مشکی", "پروفیل تست — سفید"},
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        loaded = load_workbook(output, data_only=True)
        self.assertIn("نتایج برش", loaded.sheetnames)

    def test_millimeter_project_converts_every_cutting_length(self):
        plan = calculate_cutting_plan(
            [_door(1, "مشکی", "سه طرفه")],
            [_profile()],
        )
        workbook = Workbook()
        sheet = add_cutting_results_sheet(workbook, plan, "mm")

        header_row = _find_row(sheet, "شاخه")
        data_rows = range(header_row + 1, header_row + 1 + plan["total_bins"])
        self.assertTrue(all(sheet.cell(row, 6).value == 5000 for row in data_rows))
        self.assertAlmostEqual(
            sum(sheet.cell(row, 9).value for row in data_rows),
            plan["stats"]["total_kerf_length"] * 10,
        )
        kerf_summary_row = _find_row(sheet, "مجموع افت ناشی از تیغ")
        self.assertEqual(sheet.cell(kerf_summary_row, 3).value, "میلی‌متر")
        self.assertAlmostEqual(
            sheet.cell(kerf_summary_row, 2).value,
            plan["stats"]["total_kerf_length"] * 10,
        )


if __name__ == "__main__":
    unittest.main()
