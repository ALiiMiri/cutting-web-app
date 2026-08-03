
from flask import Flask, g, render_template, request, redirect, url_for, flash, session, render_template_string, get_flashed_messages
from flask_login import LoginManager, login_required, current_user, logout_user
import os
import sqlite3
import traceback  # برای نمایش خطای کامل
from flask import send_file, jsonify
import time
import arabic_reshaper
from bidi.algorithm import get_display
from weasyprint import HTML, CSS
from datetime import datetime, date
import jdatetime
import random
import secrets
import hmac

# Import date utilities
from date_utils import (
    get_shamsi_timestamp, 
    get_shamsi_datetime_str, 
    get_shamsi_datetime_iso,
    gregorian_to_shamsi,
    gregorian_to_shamsi_date
)

from math import ceil
import json
from collections import defaultdict
from config import Config
from database import (
    get_db_connection,
    check_table_exists,
    get_all_projects,
    get_projects_paginated,
    get_recent_customers,
    add_project_db,
    get_project_details_db,
    generate_unique_project_code,
    get_doors_for_project_db,
    add_door_with_hardware_db,
    update_door_with_hardware_db,
    add_door_code_with_hardware_db,
    update_door_code_with_hardware_db,
    get_next_door_code_db,
    get_all_custom_columns,
    get_active_custom_columns,
    get_active_custom_columns_values,
    add_custom_column,
    update_custom_column_status,
    get_project_custom_columns,
    get_project_visible_custom_columns,
    set_project_column_visibility,
    remove_project_column,
    get_column_id_by_key,
    get_custom_column_options,
    add_option_to_column,
    delete_column_option,
    update_door_custom_value,
    get_door_custom_values,
    update_project_db,
    delete_project_db,
    check_column_can_hide_internal,
    update_custom_column_option,
    get_non_empty_custom_columns_for_project,
    get_price_settings_db,
    save_quote_db,
    get_all_saved_quotes_db,
    delete_quote_db,
    delete_multiple_quotes_db,
    save_doors_batch_db,
    batch_update_doors_db,
    get_column_type_db,
    get_column_id_from_option_db,
    initialize_inventory_tables,
    get_all_profile_types,
    add_profile_type,
    get_profile_details,
    get_inventory_settings,
    update_inventory_settings,
    get_inventory_stats,
    delete_profile_type,
    update_profile_type,
    get_profile_stock_details,
    add_inventory_stock,
    remove_inventory_stock,
    add_inventory_piece,
    remove_inventory_piece,
    get_inventory_logs,
    get_inventory_cutting_application_status,
    apply_cutting_plan_inventory_transaction,
    get_latest_reversible_inventory_operation,
    init_db,
    get_available_inventory_pieces,
    user_can_edit_project,
    user_can_edit_project_assignment,
    get_assignable_project_users,
    assign_project_user,
    get_project_assignment_logs,
    get_project_dashboard_counts,
    get_hardware_catalog_options,
    add_hardware_catalog_option,
    archive_hardware_catalog_option,
    move_hardware_catalog_option,
    get_profile_bracket_settings,
    update_profile_bracket_setting,
)
from door_hardware import (
    HARDWARE_CATALOG_CATEGORIES,
    HardwareValidationError,
    hardware_summary,
    normalize_door_hardware,
)

# Import blueprints
from routes import register_blueprints

# Import backup manager
import backup_manager
from maintenance import disable_maintenance, enable_maintenance, maintenance_status

# Import auth utilities
from auth_utils import (
    get_orders_view_preference,
    get_user_by_id,
    set_orders_view_preference,
)

# Import decorators
from decorators import (
    admin_required,
    manager_or_admin_required,
    roles_required,
    staff_or_admin_required,
    prevent_read_only,
)
from security_utils import PROJECT_EDIT_ENDPOINTS, access_denial_message, csrf_protected, get_csrf_token
from cutting_calculator import (
    CuttingPlanError,
    calculate_cutting_plan as build_cutting_plan,
    make_inventory_variant_key,
    normalize_color_name,
)
from cutting_excel import (
    add_cutting_results_sheet,
    create_cutting_plan_snapshot,
    resolve_applied_cutting_plan,
)
from cutting_order_excel import create_cutting_order_workbook
from cutting_orders import (
    CuttingOrderError,
    archive_project_safely,
    can_view_cutting_order,
    cancel_cutting_order,
    confirm_bar_cut,
    confirm_bars_cut,
    create_cutting_order,
    get_cutting_order,
    get_project_cutting_blockers,
    list_cutting_orders,
    project_cutting_history,
    reserve_cutting_order,
    revise_cutting_order,
    send_cutting_order,
)
from hardware_calculator import calculate_project_hardware
from factory_requirements import (
    FactoryRequirementError,
    calculate_factory_requirements,
    normalize_bracket_mode,
)
from measurements import (
    centimeters_to_measurement_unit,
    dimension_to_centimeters,
    format_measurement_value,
    measurement_unit_labels,
    normalize_measurement_unit,
)

# --- تنظیمات اولیه ---
DB_NAME = Config.DB_NAME


# --- تابع کمکی برای بررسی وجود جدول ---





# --- Flask App Setup ---
app = Flask(__name__, template_folder='templates')
app.secret_key = Config.SECRET_KEY
app.logger.setLevel("INFO")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
)

# Configure Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'auth.login'
login_manager.login_message = 'لطفاً برای دسترسی به این صفحه وارد شوید.'
login_manager.login_message_category = 'warning'


@app.context_processor
def inject_security_values():
    def can_edit_project(project_id):
        if not current_user.is_authenticated:
            return False
        return user_can_edit_project(current_user.id, current_user.role, project_id)

    return {
        'csrf_token': get_csrf_token(),
        'current_user_can_edit_project': can_edit_project,
    }


# Configure Flask to use UTF-8 encoding
@app.before_request
def record_request_start():
    """Log enough context to identify requests that never complete."""
    g.request_started_at = time.monotonic()
    g.request_id = secrets.token_hex(4)
    if request.endpoint != "healthz":
        app.logger.info(
            "REQUEST_START id=%s method=%s path=%s",
            g.request_id,
            request.method,
            request.path,
        )


@app.after_request
def set_charset(response):
    """Ensure all responses use UTF-8 encoding"""
    if 'Content-Type' in response.headers:
        content_type = response.headers['Content-Type']
        if 'charset=' not in content_type:
            response.headers['Content-Type'] = content_type + '; charset=utf-8'
    else:
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
    if request.endpoint and (
        request.endpoint.startswith('admin.') or request.endpoint.startswith('auth.')
    ):
        response.headers['Cache-Control'] = 'no-store'
    if request.endpoint != "healthz":
        started_at = getattr(g, "request_started_at", None)
        duration_ms = (
            round((time.monotonic() - started_at) * 1000)
            if started_at is not None
            else -1
        )
        app.logger.info(
            "REQUEST_END id=%s status=%s duration_ms=%s",
            getattr(g, "request_id", "-"),
            response.status_code,
            duration_ms,
        )
    return response

# Configure Jinja2 to use UTF-8
app.jinja_env.autoescape = True
app.jinja_env.auto_reload = True

# اضافه کردن فیلتر شمسی به Jinja2
@app.template_filter('shamsi')
def shamsi_filter(dt):
    """تبدیل تاریخ میلادی به شمسی برای استفاده در template ها"""
    return gregorian_to_shamsi(dt)

@app.template_filter('shamsi_date')
def shamsi_date_filter(dt):
    """تبدیل تاریخ میلادی به شمسی (فقط تاریخ) برای استفاده در template ها"""
    return gregorian_to_shamsi_date(dt)

# Flask-Login user loader
@login_manager.user_loader
def load_user(user_id):
    return get_user_by_id(user_id)

# Global before_request to protect all routes
@app.before_request
def require_login():
    """نیاز به لاگین برای همه routeها به جز login و static"""
    maintenance = maintenance_status()
    if maintenance and request.endpoint not in ("static",):
        return render_template_string(
            "<html dir='rtl'><meta charset='utf-8'><title>در حال به‌روزرسانی</title>"
            "<body style='font-family:Tahoma;text-align:center;padding:80px'>"
            "<h1>سامانه موقتاً در حال نگهداری است</h1>"
            "<p>اطلاعات شما محفوظ است؛ لطفاً چند دقیقه دیگر دوباره تلاش کنید.</p></body></html>"
        ), 503
    # مسیرهای استثنا (که نیاز به لاگین ندارند)
    allowed_endpoints = ['auth.login', 'healthz', 'static']
    
    # اگر کاربر لاگین نیست و مسیر جاری در لیست استثنا نیست
    if not current_user.is_authenticated:
        if request.endpoint not in allowed_endpoints:
            flash('لطفاً برای دسترسی به سیستم وارد شوید.', 'warning')
            return redirect(url_for('auth.login'))

    if current_user.is_authenticated:
        # غیرفعال‌کردن حساب یا تغییر رمز باید همان لحظه همه ورودهای قبلی را ببندد.
        session_version = session.get('auth_version')
        if not current_user.is_active or session_version != current_user.session_version:
            logout_user()
            session.pop('auth_version', None)
            flash('برای حفظ امنیت، لطفاً دوباره وارد شوید.', 'warning')
            return redirect(url_for('auth.login'))

        denial_message = access_denial_message(
            current_user.role, request.method, request.endpoint, request.path
        )
        if denial_message:
            flash(denial_message, 'error')
            return redirect(url_for('index'))

        project_id = (request.view_args or {}).get('project_id')
        if (
            project_id is not None
            and request.endpoint in PROJECT_EDIT_ENDPOINTS
            and not user_can_edit_project(current_user.id, current_user.role, project_id)
        ):
            flash('فقط مسئول این پروژه یا مدیر می‌تواند آن را تغییر دهد.', 'error')
            return redirect(url_for('view_project', project_id=project_id))
    
    # اگر کاربر لاگین است ولی باید رمز تغییر دهد
    if current_user.is_authenticated and hasattr(current_user, 'must_change_password') and current_user.must_change_password:
        # فقط به صفحات change_password و logout اجازه دسترسی
        if request.endpoint not in ['auth.change_password', 'auth.logout', 'static']:
            flash('لطفاً ابتدا رمز عبور خود را تغییر دهید.', 'warning')
            return redirect(url_for('auth.change_password'))

# --- مقداردهی اولیه دیتابیس ---
# فراخوانی تابع ایجاد جداول انبار در شروع برنامه
# فراخوانی تابع ایجاد جداول انبار در شروع برنامه
print("DEBUG: Initializing database tables...")
init_db()

# Register blueprints
register_blueprints(app)

# --- بررسی وجود جداول بعد از مقداردهی اولیه ---
print("\n--- Starting table checks ---")
check_table_exists("projects")
check_table_exists("doors")
check_table_exists("custom_columns")
check_table_exists("custom_column_options")
check_table_exists("door_custom_values")
print("--- Table checks completed ---\n")


# --- Routes (آدرس‌های وب) ---


@app.route("/healthz")
def healthz():
    """Check both worker responsiveness and read access to the live database."""
    connection = None
    try:
        connection = sqlite3.connect(
            f"file:{DB_NAME}?mode=ro",
            uri=True,
            timeout=1,
        )
        connection.execute("PRAGMA busy_timeout = 1000")
        connection.execute("SELECT 1 FROM schema_migrations LIMIT 1").fetchone()
        return jsonify(status="ok")
    except sqlite3.Error:
        app.logger.exception("HEALTHCHECK_DATABASE_FAILED")
        return jsonify(status="error"), 503
    finally:
        if connection is not None:
            connection.close()


@app.route("/")
def index():
    print("DEBUG: Route / (index) called.")
    try:
        digit_translation = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        search = request.args.get('search', '', type=str).strip().translate(digit_translation)
        sort_by = request.args.get('sort_by', 'id', type=str)
        sort_order = request.args.get('sort_order', 'DESC', type=str)
        date_from = request.args.get('date_from', '', type=str).strip().translate(digit_translation)
        date_to = request.args.get('date_to', '', type=str).strip().translate(digit_translation)
        customer_filter = request.args.get('customer_filter', '', type=str).strip()
        scope = request.args.get('scope', 'all', type=str)
        if scope not in ('all', 'mine', 'unassigned'):
            scope = 'all'
        per_page = request.args.get('per_page', 15, type=int)
        
        # Validate per_page
        if per_page not in [10, 15, 20, 30, 50]:
            per_page = 15
        
        # Get paginated projects
        result = get_projects_paginated(
            page=page,
            per_page=per_page,
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
            date_from=date_from,
            date_to=date_to,
            customer_filter=customer_filter,
            scope=scope,
            current_user_id=current_user.id,
        )
        for project in result['projects']:
            project['can_edit'] = user_can_edit_project_assignment(
                current_user.id,
                current_user.role,
                project['assigned_to_user_id'],
            )
        cutting_blockers = get_project_cutting_blockers(
            [project['id'] for project in result['projects']]
        )
        for project in result['projects']:
            project['cutting_blocker'] = cutting_blockers.get(project['id'])
        
        return render_template(
            "index.html",
            projects=result['projects'],
            pagination=result,
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
            date_from=date_from,
            date_to=date_to,
            customer_filter=customer_filter,
            scope=scope,
            per_page=per_page,
            recent_customers=get_recent_customers(),
            dashboard_counts=get_project_dashboard_counts(current_user.id),
            orders_view_preference=get_orders_view_preference(current_user.id),
            assignable_users=(
                get_assignable_project_users()
                if current_user.role in ('admin', 'manager') else []
            ),
        )
    except Exception as e:
        print(f"!!!!!! Unexpected error in index route: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش فهرست سفارش‌ها رخ داد.", "error")
        return render_template(
            "index.html",
            projects=[],
            pagination={"total": 0, "page": 1, "pages": 1, "per_page": 15},
            search="",
            sort_by="id",
            sort_order="DESC",
            date_from="",
            date_to="",
            customer_filter="",
            scope="all",
            per_page=15,
            recent_customers=[],
            assignable_users=[],
            dashboard_counts={"total": 0, "mine": 0, "unassigned": 0},
            orders_view_preference="table",
        )


@app.route("/preferences/orders-view", methods=["POST"])
@csrf_protected
def save_orders_view_preference():
    data = request.get_json(silent=True) or request.form
    preference = str(data.get("view", "")).strip()
    if preference not in ("table", "cards"):
        return jsonify({"success": False, "error": "نوع نمایش معتبر نیست."}), 400
    if not set_orders_view_preference(current_user.id, preference):
        return jsonify({"success": False, "error": "ذخیره نوع نمایش انجام نشد."}), 500
    return jsonify({"success": True, "view": preference})


@app.route("/cutting-orders")
def cutting_orders_list():
    return render_template(
        "cutting_orders.html",
        cutting_orders=list_cutting_orders(current_user.id, current_user.role),
    )


@app.route("/cutting-orders/calculate", methods=["POST"])
@staff_or_admin_required
@csrf_protected
def calculate_grouped_cutting_order():
    project_ids = request.form.getlist("project_ids")
    if not project_ids:
        flash("حداقل یک سفارش را برای محاسبه برش انتخاب کنید.", "warning")
        return redirect(url_for("index"))
    try:
        normalized_ids = [int(value) for value in project_ids]
    except (TypeError, ValueError):
        flash("انتخاب سفارش‌ها معتبر نیست.", "error")
        return redirect(url_for("index"))
    inaccessible = [
        project_id
        for project_id in normalized_ids
        if not user_can_edit_project(current_user.id, current_user.role, project_id)
    ]
    if inaccessible:
        flash(
            "کارمند فقط می‌تواند سفارش‌های واگذارشده به خودش را محاسبه کند.",
            "error",
        )
        return redirect(url_for("index"))
    try:
        order_id = create_cutting_order(normalized_ids, current_user.id)
        flash(
            "محاسبه ثبت شد؛ هنوز هیچ چیزی از انبار رزرو یا کم نشده است.",
            "success",
        )
        return redirect(url_for("cutting_order_details", order_id=order_id))
    except CuttingOrderError as exc:
        flash(str(exc), "error")
    except sqlite3.Error:
        app.logger.exception("GROUPED_CUTTING_ORDER_CREATE_FAILED")
        flash("ثبت سفارش برش انجام نشد؛ هیچ تغییری در انبار ایجاد نشد.", "error")
    return redirect(url_for("index"))


@app.route("/cutting-orders/<int:order_id>")
def cutting_order_details(order_id):
    if not can_view_cutting_order(
        order_id, current_user.id, current_user.role
    ):
        flash("شما اجازه مشاهده این سفارش برش را ندارید.", "error")
        return redirect(url_for("cutting_orders_list"))
    order = get_cutting_order(order_id)
    if not order:
        flash("سفارش برش پیدا نشد.", "error")
        return redirect(url_for("cutting_orders_list"))
    return render_template("cutting_order_details.html", order=order)


@app.route("/cutting-orders/<int:order_id>/reserve", methods=["POST"])
@manager_or_admin_required
@csrf_protected
def reserve_cutting_order_route(order_id):
    try:
        changed = reserve_cutting_order(order_id, current_user.id)
        flash(
            "تمام منابع این سفارش با موفقیت رزرو شد."
            if changed
            else "این سفارش قبلاً رزرو شده است.",
            "success" if changed else "warning",
        )
    except CuttingOrderError as exc:
        flash(str(exc), "error")
    except sqlite3.Error:
        app.logger.exception("CUTTING_ORDER_RESERVE_FAILED")
        flash("رزرو انجام نشد و هیچ رزرو ناقصی ثبت نشد.", "error")
    return redirect(url_for("cutting_order_details", order_id=order_id))


@app.route(
    "/cutting-orders/<int:order_id>/confirm-cuts",
    methods=["POST"],
)
@roles_required("admin", "manager", "factory")
@csrf_protected
def confirm_cutting_order_bars_route(order_id):
    single_bar_id = request.form.get("single_bar_id", type=int)
    selected_values = (
        [str(single_bar_id)]
        if single_bar_id
        else request.form.getlist("bar_ids")
    )
    try:
        bar_ids = []
        for value in selected_values:
            bar_id = int(value)
            if bar_id <= 0 or bar_id in bar_ids:
                raise ValueError
            bar_ids.append(bar_id)
    except (TypeError, ValueError):
        flash("انتخاب شاخه‌ها معتبر نیست.", "error")
        return redirect(url_for("cutting_order_details", order_id=order_id))

    updates = [
        {
            "bar_id": bar_id,
            "actual_remaining": request.form.get(f"actual_remaining_{bar_id}"),
        }
        for bar_id in bar_ids
    ]
    try:
        confirmed_count = confirm_bars_cut(
            order_id,
            updates,
            current_user.id,
        )
        flash(
            f"{confirmed_count} شاخه با موفقیت تأیید شد و موجودی همه آن‌ها ثبت شد.",
            "success",
        )
    except CuttingOrderError as exc:
        flash(str(exc), "error")
    except sqlite3.Error:
        app.logger.exception("CUTTING_ORDER_BULK_CUT_FAILED")
        flash(
            "تأیید گروهی انجام نشد؛ موجودی هیچ‌کدام از شاخه‌ها تغییر نکرد.",
            "error",
        )
    return redirect(url_for("cutting_order_details", order_id=order_id))


@app.route("/cutting-orders/<int:order_id>/send", methods=["POST"])
@manager_or_admin_required
@csrf_protected
def send_cutting_order_route(order_id):
    try:
        changed = send_cutting_order(order_id, current_user.id)
        flash(
            "سفارش قفل و برای کارخانه ارسال شد."
            if changed
            else "این سفارش قبلاً برای کارخانه ارسال شده است.",
            "success" if changed else "warning",
        )
    except CuttingOrderError as exc:
        flash(str(exc), "error")
    return redirect(url_for("cutting_order_details", order_id=order_id))


@app.route(
    "/cutting-orders/<int:order_id>/bars/<int:bar_id>/confirm-cut",
    methods=["POST"],
)
@roles_required("admin", "manager", "factory")
@csrf_protected
def confirm_cutting_order_bar_route(order_id, bar_id):
    try:
        changed = confirm_bar_cut(
            order_id,
            bar_id,
            current_user.id,
            request.form.get("actual_remaining"),
        )
        flash(
            "برش این شاخه تأیید و موجودی همان شاخه ثبت شد."
            if changed
            else "این شاخه قبلاً تأیید شده است.",
            "success" if changed else "warning",
        )
    except CuttingOrderError as exc:
        flash(str(exc), "error")
    except sqlite3.Error:
        app.logger.exception("CUTTING_ORDER_BAR_CUT_FAILED")
        flash("ثبت برش انجام نشد؛ موجودی این شاخه تغییر نکرد.", "error")
    return redirect(
        url_for(
            "cutting_order_details",
            order_id=order_id,
            _anchor=f"bar-{bar_id}",
        )
    )


@app.route("/cutting-orders/<int:order_id>/cancel", methods=["POST"])
@manager_or_admin_required
@csrf_protected
def cancel_cutting_order_route(order_id):
    try:
        changed = cancel_cutting_order(
            order_id, current_user.id, request.form.get("reason")
        )
        flash(
            "رزرو شاخه‌های برش‌نخورده آزاد شد؛ سابقه برش‌های انجام‌شده حفظ شد."
            if changed
            else "این سفارش قبلاً بسته شده است.",
            "success" if changed else "warning",
        )
    except CuttingOrderError as exc:
        flash(str(exc), "error")
    return redirect(url_for("cutting_order_details", order_id=order_id))


@app.route("/cutting-orders/<int:order_id>/revise", methods=["POST"])
@manager_or_admin_required
@csrf_protected
def revise_cutting_order_route(order_id):
    try:
        new_order_id = revise_cutting_order(
            order_id, current_user.id, request.form.get("reason")
        )
        flash(
            "رزرو شاخه‌های برش‌نخورده آزاد و نسخه جدید محاسبه شد؛ "
            "قطعاتی که قبلاً برش خورده‌اند دوباره در برنامه نیامده‌اند.",
            "success",
        )
        return redirect(
            url_for("cutting_order_details", order_id=new_order_id)
        )
    except CuttingOrderError as exc:
        flash(str(exc), "error")
    except sqlite3.Error:
        app.logger.exception("CUTTING_ORDER_REVISION_FAILED")
        flash("ساخت نسخه جدید انجام نشد.", "error")
    return redirect(url_for("cutting_order_details", order_id=order_id))


@app.route("/cutting-orders/<int:order_id>/excel")
def cutting_order_excel(order_id):
    if not can_view_cutting_order(
        order_id, current_user.id, current_user.role
    ):
        flash("شما اجازه دریافت این خروجی را ندارید.", "error")
        return redirect(url_for("cutting_orders_list"))
    order = get_cutting_order(order_id)
    if not order:
        flash("سفارش برش پیدا نشد.", "error")
        return redirect(url_for("cutting_orders_list"))
    output = create_cutting_order_workbook(order)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{order['order_number']}.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


@app.route("/home")
def home():
    """نام دیگر برای صفحه اصلی (برای سازگاری با تمپلیت‌ها)"""
    return index()


@app.route("/project/add", methods=["GET"])
def add_project_form():
    print("DEBUG: Route /project/add (GET - add_project_form) called.")
    order_ref = generate_unique_project_code()
    print(f"DEBUG: Generated order_ref (project code): {order_ref}")
    return render_template("add_project.html", order_ref=order_ref, measurement_unit="cm")


@app.route("/project/add", methods=["POST"])
@csrf_protected
@staff_or_admin_required
def add_project_route():
    print("DEBUG: Route /project/add (POST - add_project_route) called.")
    customer_name = request.form.get("customer_name", "").strip()
    order_ref = request.form.get("order_ref", "").strip()
    date_shamsi = request.form.get("date_shamsi", "").strip()
    try:
        measurement_unit = normalize_measurement_unit(
            request.form.get("measurement_unit", "cm")
        )
    except ValueError as exc:
        flash(str(exc), "error")
        return render_template(
            "add_project.html",
            order_ref=order_ref or generate_unique_project_code(),
            measurement_unit="cm",
            customer_name=customer_name,
            date_shamsi=date_shamsi,
        )
    
    # Validate date is required
    if not date_shamsi:
        flash("لطفاً تاریخ را انتخاب کنید. انتخاب تاریخ اجباری است.", "error")
        return render_template(
            "add_project.html",
            order_ref=order_ref or generate_unique_project_code(),
            measurement_unit=measurement_unit,
            customer_name=customer_name,
            date_shamsi=date_shamsi,
        )
    
    if not customer_name:
        flash("لطفاً نام مشتری را وارد کنید.", "error")
        return render_template(
            "add_project.html",
            order_ref=order_ref or generate_unique_project_code(),
            measurement_unit=measurement_unit,
            customer_name=customer_name,
            date_shamsi=date_shamsi,
        )
    
    # If order_ref is empty, generate a new one
    if not order_ref:
        order_ref = generate_unique_project_code()
    
    # Use order_ref as project_code (they are the same)
    project_code = order_ref
    
    new_id = add_project_db(
        customer_name,
        order_ref,
        date_shamsi,
        project_code,
        measurement_unit=measurement_unit,
        created_by_user_id=current_user.id,
    )
    if new_id:
        flash(
            f"سفارش جدید برای مشتری «{customer_name}» با شماره {order_ref} ایجاد شد.",
            "success",
        )
        print(f"DEBUG: Project ID {new_id} added with order_ref/project_code {order_ref}, name: '{customer_name}', date: {date_shamsi}, redirecting to project details.")
        return redirect(url_for("view_project", project_id=new_id))
    else:
        flash("خطایی در ذخیره پروژه رخ داد.", "error")
        return render_template(
            "add_project.html",
            order_ref=order_ref or generate_unique_project_code(),
            measurement_unit=measurement_unit,
            customer_name=customer_name,
            date_shamsi=date_shamsi,
        )


@app.route("/project/<int:project_id>/update", methods=["POST"])
@csrf_protected
@staff_or_admin_required
def update_project_route(project_id):
    """ویرایش پروژه از صفحه خانه (فرم مودال)"""
    try:
        customer_name = request.form.get("customer_name", "").strip()
        order_ref = request.form.get("order_ref", "").strip()
        date_shamsi = request.form.get("date_shamsi", "").strip()

        if not customer_name and not order_ref:
            flash("لطفاً حداقل نام مشتری یا شماره سفارش را وارد کنید.", "error")
            return redirect(url_for("index"))

        success = update_project_db(project_id, customer_name, order_ref, date_shamsi)
        if success:
            flash("سفارش با موفقیت ویرایش شد.", "success")
        else:
            flash("خطا در ویرایش سفارش.", "error")
        return redirect(url_for("index"))
    except Exception as e:
        print(f"!!!!!! Unexpected error in update_project_route: {e}")
        traceback.print_exc()
        flash("خطایی در ویرایش پروژه رخ داد.", "error")
        return redirect(url_for("index"))


@app.route("/project/<int:project_id>/assign", methods=["POST"])
@manager_or_admin_required
@csrf_protected
def assign_project_route(project_id):
    raw_user_id = request.form.get("assigned_to_user_id", "").strip()
    try:
        new_user_id = int(raw_user_id) if raw_user_id else None
    except ValueError:
        flash("کاربر انتخاب‌شده معتبر نیست.", "error")
        return redirect(url_for("index"))

    success, message = assign_project_user(project_id, new_user_id, current_user.id)
    flash(message, "success" if success else "error")
    return redirect(url_for("index"))


@app.route("/project/<int:project_id>/assignment_history")
@manager_or_admin_required
def project_assignment_history(project_id):
    project = get_project_details_db(project_id)
    if not project:
        flash("پروژه پیدا نشد.", "error")
        return redirect(url_for("index"))
    return render_template(
        "project_assignment_history.html",
        project=project,
        assignment_logs=get_project_assignment_logs(project_id),
    )


@app.route("/project/<int:project_id>/delete", methods=["POST", "GET"])
@manager_or_admin_required
@csrf_protected
def delete_project_route(project_id):
    """حذف پروژه (از صفحه خانه). GET فقط ریدایرکت می‌کند؛ حذف واقعی با POST انجام می‌شود."""
    try:
        if request.method == "GET":
            flash("برای حذف سفارش، از منوی سه‌نقطه همان سفارش استفاده کنید.", "warning")
            return redirect(url_for("index"))

        # 🔄 بکاپ خودکار قبل از حذف پروژه
        print(f"ایجاد بکاپ خودکار قبل از حذف پروژه {project_id}...")
        backup_success, backup_result = backup_manager.create_backup(
            reason=f"before_delete_project",
            user="system",
            metadata={"project_id": project_id, "action": "delete_project"}
        )
        if backup_success:
            print(f"✓ بکاپ قبل از حذف پروژه ایجاد شد: {backup_result}")
        else:
            flash(f"حذف متوقف شد؛ ایجاد بکاپ ایمنی ناموفق بود: {backup_result}", "error")
            return redirect(url_for("index"))

        cutting_history = project_cutting_history(project_id)
        if cutting_history["order_count"]:
            archived, affected_orders = archive_project_safely(
                project_id, current_user.id
            )
            if archived:
                flash(
                    "سفارش بایگانی شد؛ رزروهای برش‌نخورده آزاد شدند و "
                    "تاریخچه شاخه‌های واقعاً بریده‌شده، تکه‌ها و ضایعات حفظ شد."
                    + (
                        f" {len(affected_orders)} دستور برش باز نیز بسته شد."
                        if affected_orders
                        else ""
                    ),
                    "success",
                )
            else:
                flash("بایگانی سفارش انجام نشد.", "error")
            return redirect(url_for("index"))

        success = delete_project_db(project_id)
        if success:
            flash("سفارش با موفقیت حذف شد.", "success")
        else:
            flash("خطا در حذف سفارش.", "error")
        return redirect(url_for("index"))
    except Exception as e:
        print(f"!!!!!! Unexpected error in delete_project_route: {e}")
        traceback.print_exc()
        flash("خطایی در حذف سفارش رخ داد.", "error")
        return redirect(url_for("index"))


@app.route("/project/<int:project_id>")
def view_project(project_id):
    print(f"DEBUG: >>>>>>> Entering route /project/{project_id} (view_project) <<<<<<<")
    print(f"DEBUG: Request Headers (view_project):\n{request.headers}")
    
    # بررسی پارامتر force_refresh برای تازه‌سازی کامل صفحه
    force_refresh = request.args.get("force_refresh", "0") == "1"
    
    # اگر force_refresh فعال است، به صفحه treeview هدایت می‌کنیم
    if force_refresh:
        print(f"DEBUG: ریدایرکت به صفحه treeview با force_refresh=1")
        timestamp = int(time.time())
        return redirect(url_for("project_treeview", project_id=project_id, force_refresh=1, refresh_columns=1, t=timestamp))
    
    project_details = None
    door_list = []
    try:
        project_details = get_project_details_db(project_id)
        if not project_details:
            flash(f"پروژه با ID {project_id} یافت نشد.", "error")
            print(f"DEBUG: پروژه {project_id} یافت نشد، ریدایرکت به index.")
            return redirect(url_for("index"))
        door_list = get_doors_for_project_db(project_id)
        project_measurement_unit = normalize_measurement_unit(
            project_details.get("measurement_unit", "cm")
        )
        unit_labels = measurement_unit_labels(project_measurement_unit)
        for door in door_list:
            door["display_width"] = (
                format_measurement_value(
                    centimeters_to_measurement_unit(
                        door["width"], project_measurement_unit
                    )
                )
                if door.get("width") is not None
                else ""
            )
            door["display_height"] = (
                format_measurement_value(
                    centimeters_to_measurement_unit(
                        door["height"], project_measurement_unit
                    )
                )
                if door.get("height") is not None
                else ""
            )
            door["hardware_summary"] = hardware_summary(door)
        total_door_quantity = sum(
            int(door.get("quantity") or 0) for door in door_list
        )
        print(
            f"DEBUG: رندر کردن project_details.html برای پروژه {project_id} با {len(door_list)} درب."
        )
        return render_template(
            "project_details.html",
            project=project_details,
            doors=door_list,
            total_door_quantity=total_door_quantity,
            measurement_unit_label=unit_labels["fa"],
            hardware_options=get_hardware_catalog_options(),
            next_door_code=get_next_door_code_db(project_id),
            profile_options=[
                item["name"] for item in get_all_profile_types()
            ],
            profile_color_options=get_custom_column_options(
                get_column_id_by_key("rang")
            ),
        )
    except Exception as e:
        print(f"!!!!!! خطای جدی در روت view_project برای ID {project_id}: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش جزئیات پروژه رخ داد. لطفاً دوباره تلاش کنید.", "error")
        print(f"DEBUG: خطا در view_project، ریدایرکت به index.")
        return redirect(url_for("index"))


def _normalize_door_code_locations(data):
    door_code = " ".join(str(data.get("door_code", "")).split())
    if not door_code:
        raise ValueError("کد درب را وارد کنید.")
    if len(door_code) > 40:
        raise ValueError("کد درب نباید بیشتر از ۴۰ نویسه باشد.")
    raw_locations = data.get("locations")
    if not isinstance(raw_locations, list):
        raw_locations = [
            {"location": data.get("location", ""), "quantity": data.get("quantity", 1)}
        ]
    locations = []
    for raw in raw_locations:
        if not isinstance(raw, dict):
            raise ValueError("اطلاعات محل نصب معتبر نیست.")
        location = " ".join(str(raw.get("location", "")).split())
        if not location:
            raise ValueError("نام همه محل‌های نصب را وارد کنید.")
        if len(location) > 160:
            raise ValueError("نام محل نصب نباید بیشتر از ۱۶۰ نویسه باشد.")
        try:
            quantity = int(raw.get("quantity"))
        except (TypeError, ValueError) as exc:
            raise ValueError("تعداد هر محل نصب باید عدد صحیح باشد.") from exc
        if quantity <= 0:
            raise ValueError("تعداد هر محل نصب باید بزرگ‌تر از صفر باشد.")
        locations.append({"location": location, "quantity": quantity})
    if not locations:
        raise ValueError("حداقل یک محل نصب اضافه کنید.")
    if len(locations) > 100:
        raise ValueError("برای هر کد حداکثر ۱۰۰ محل نصب قابل ثبت است.")
    return door_code, locations


@app.route("/project/<int:project_id>/quick_add_door", methods=["POST"])
@csrf_protected
@staff_or_admin_required
def quick_add_door(project_id):
    """Add one door directly from the project overview modal."""
    project_info = get_project_details_db(project_id)
    if not project_info:
        return jsonify({"success": False, "error": "سفارش مورد نظر پیدا نشد."}), 404

    data = request.get_json(silent=True) or request.form
    try:
        door_code, locations = _normalize_door_code_locations(data)
        measurement_unit = normalize_measurement_unit(
            project_info.get("measurement_unit", "cm")
        )
        width = dimension_to_centimeters(data.get("width"), measurement_unit)
        height = dimension_to_centimeters(data.get("height"), measurement_unit)
        direction = str(data.get("direction", "راست"))
        frame_type = str(data.get("frame_type", "سه طرفه"))
        profile_name = " ".join(str(data.get("profile_name", "")).split())
        profile_color = " ".join(str(data.get("profile_color", "")).split())
        hardware_payload = data.get("hardware") or {}
        hardware = normalize_door_hardware(hardware_payload)
        bracket_mode = normalize_bracket_mode(hardware_payload.get("bracket_mode"))
    except (HardwareValidationError, FactoryRequirementError) as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except (TypeError, ValueError) as exc:
        return jsonify({"success": False, "error": str(exc) or "ابعاد معتبر نیست."}), 400

    if width <= 0 or height <= 0:
        return jsonify(
            {"success": False, "error": "عرض و ارتفاع باید بزرگ‌تر از صفر باشند."}
        ), 400
    if direction not in {"راست", "چپ"}:
        return jsonify({"success": False, "error": "جهت انتخاب‌شده معتبر نیست."}), 400
    if frame_type not in {"دو طرفه", "سه طرفه"}:
        return jsonify({"success": False, "error": "نوع چارچوب معتبر نیست."}), 400
    if not profile_name or not profile_color:
        return jsonify({"success": False, "error": "نوع و رنگ پروفیل را انتخاب کنید."}), 400

    door_id = add_door_code_with_hardware_db(
        project_id, door_code, width, height, direction, locations, hardware,
        bracket_mode=bracket_mode,
        frame_type=frame_type,
        profile_name=profile_name,
        profile_color=profile_color,
    )
    if door_id == "duplicate_code":
        return jsonify({"success": False, "error": "این کد درب قبلاً در سفارش استفاده شده است."}), 409
    if not door_id:
        return jsonify({"success": False, "error": "ذخیره درب انجام نشد."}), 500

    return jsonify(
        {
            "success": True,
            "message": "کد درب، یراق و محل‌های نصب به سفارش اضافه شد.",
        }
    )


@app.route(
    "/project/<int:project_id>/update_door/<int:door_id>", methods=["POST"]
)
@csrf_protected
@staff_or_admin_required
def update_door(project_id, door_id):
    """Update one door from the project overview modal."""
    project_info = get_project_details_db(project_id)
    if not project_info:
        return jsonify({"success": False, "error": "سفارش مورد نظر پیدا نشد."}), 404

    data = request.get_json(silent=True) or request.form
    try:
        door_code, locations = _normalize_door_code_locations(data)
        measurement_unit = normalize_measurement_unit(
            project_info.get("measurement_unit", "cm")
        )
        width = dimension_to_centimeters(data.get("width"), measurement_unit)
        height = dimension_to_centimeters(data.get("height"), measurement_unit)
        direction = str(data.get("direction", "راست"))
        frame_type = str(data.get("frame_type", "سه طرفه"))
        profile_name = " ".join(str(data.get("profile_name", "")).split())
        profile_color = " ".join(str(data.get("profile_color", "")).split())
        hardware_payload = data.get("hardware") or {}
        hardware = normalize_door_hardware(hardware_payload)
        bracket_mode = normalize_bracket_mode(hardware_payload.get("bracket_mode"))
    except (HardwareValidationError, FactoryRequirementError) as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except (TypeError, ValueError) as exc:
        return jsonify({"success": False, "error": str(exc) or "ابعاد معتبر نیست."}), 400

    if width <= 0 or height <= 0:
        return jsonify(
            {"success": False, "error": "عرض و ارتفاع باید بزرگ‌تر از صفر باشند."}
        ), 400
    if direction not in {"راست", "چپ"}:
        return jsonify({"success": False, "error": "جهت انتخاب‌شده معتبر نیست."}), 400
    if frame_type not in {"دو طرفه", "سه طرفه"}:
        return jsonify({"success": False, "error": "نوع چارچوب معتبر نیست."}), 400
    if not profile_name or not profile_color:
        return jsonify({"success": False, "error": "نوع و رنگ پروفیل را انتخاب کنید."}), 400

    updated = update_door_code_with_hardware_db(
        project_id, door_id, door_code, width, height, direction, locations, hardware,
        bracket_mode=bracket_mode,
        frame_type=frame_type,
        profile_name=profile_name,
        profile_color=profile_color,
    )
    if updated == "duplicate_code":
        return jsonify({"success": False, "error": "این کد درب قبلاً در سفارش استفاده شده است."}), 409
    if updated is False:
        return jsonify({"success": False, "error": "درب مورد نظر پیدا نشد."}), 404
    if updated is None:
        return jsonify({"success": False, "error": "ذخیره تغییرات انجام نشد."}), 500

    return jsonify({"success": True, "message": "کد درب، یراق و محل‌های نصب ویرایش شد."})


@app.route("/project/<int:project_id>/add_door", methods=["GET"])
def add_door_form(project_id):
    print(
        f"DEBUG: روت /project/{project_id}/add_door (GET - add_door_form) فراخوانی شد."
    )
    project_info = get_project_details_db(project_id)
    if not project_info:
        flash(f"پروژه با ID {project_id} یافت نشد.", "error")
        return redirect(url_for("index"))
    # <-- کلید session منحصر به فرد برای هر پروژه
    pending_doors = session.get(f"pending_doors_{project_id}", [])
    pending_count = len(pending_doors)
    print(
        f"DEBUG: نمایش فرم افزودن درب برای پروژه {project_id}. تعداد منتظر: {pending_count}"
    )
    return render_template(
        "add_door.html", project_info=project_info, pending_count=pending_count
    )


@app.route("/project/<int:project_id>/add_door", methods=["POST"])
def add_door_buffer(project_id):
    print(
        f"DEBUG: روت /project/{project_id}/add_door (POST - add_door_buffer) فراخوانی شد."
    )
    location = request.form.get("location")
    width_str = request.form.get("width")
    height_str = request.form.get("height")
    quantity_str = request.form.get("quantity")
    direction = request.form.get("direction")

    # فیلدهای سفارشی جدید
    rang = request.form.get("rang", "")
    noe_profile = request.form.get("noe_profile", "")
    vaziat = request.form.get("vaziat", "")
    lola = request.form.get("lola", "")
    ghofl = request.form.get("ghofl", "")
    accessory = request.form.get("accessory", "")
    # Existing and newly added doors are three-sided unless explicitly changed
    # to two-sided from door management / batch edit.
    kolaft = request.form.get("kolaft", "سه طرفه") or "سه طرفه"
    dastgire = request.form.get("dastgire", "")
    tozihat = request.form.get("tozihat", "")

    project_info = get_project_details_db(project_id)
    if not project_info:
        flash(f"پروژه با ID {project_id} یافت نشد.", "error")
        return redirect(url_for("index"))

    width = None
    height = None
    quantity = None
    errors = False
    try:
        # The project's stored unit is authoritative; a door form cannot
        # silently override it.
        measurement_unit = normalize_measurement_unit(
            project_info.get("measurement_unit", "cm")
        )
        if width_str:
            width = dimension_to_centimeters(width_str, measurement_unit)
        if height_str:
            height = dimension_to_centimeters(height_str, measurement_unit)
        if quantity_str:
            quantity = int(quantity_str)
        if (
            (width is not None and width <= 0)
            or (height is not None and height <= 0)
            or (quantity is not None and quantity <= 0)
        ):
            flash("مقادیر عرض، ارتفاع و تعداد باید مثبت باشند.", "error")
            errors = True
    except ValueError:
        flash("مقادیر عرض، ارتفاع و تعداد باید به صورت عددی وارد شوند.", "error")
        errors = True

    if errors:
        pending_doors = session.get(f"pending_doors_{project_id}", [])
        pending_count = len(pending_doors)
        print(f"DEBUG: خطای اعتبارسنجی در افزودن درب. بازگشت به فرم با داده‌های قبلی.")
        return render_template(
            "add_door.html",
            project_info=project_info,
            pending_count=pending_count,
            form_data=request.form,
        )

    # <-- کلید session منحصر به فرد
    pending_doors = session.get(f"pending_doors_{project_id}", [])
    new_door_data = {
        "location": location,
        "width": width,
        "height": height,
        "quantity": quantity,
        "direction": direction,
        "rang": rang,
        "noe_profile": noe_profile,
        "vaziat": vaziat,
        "lola": lola,
        "ghofl": ghofl,
        "accessory": accessory,
        "kolaft": kolaft,
        "dastgire": dastgire,
        "tozihat": tozihat,
    }
    pending_doors.append(new_door_data)
    # <-- کلید session منحصر به فرد
    session[f"pending_doors_{project_id}"] = pending_doors
    print(
        f"DEBUG: درب به لیست موقت پروژه {project_id} اضافه شد. تعداد منتظر: {len(pending_doors)}"
    )
    flash(
        "درب به لیست موقت اضافه شد. برای ذخیره نهایی از دکمه 'اتمام' استفاده کنید.",
        "success",
    )
    return redirect(url_for("add_door_form", project_id=project_id))


@app.route("/project/<int:project_id>/finish_doors", methods=["GET"])
def finish_adding_doors(project_id):
    print(
        f"DEBUG: روت /project/{project_id}/finish_doors (GET - finish_adding_doors) فراخوانی شد."
    )
    # <-- کلید session منحصر به فرد
    pending_doors = session.get(f"pending_doors_{project_id}", [])
    saved_count = 0
    error_count = 0

    if not pending_doors:
        flash("هیچ دربی در لیست موقت برای ذخیره وجود ندارد.", "warning")
        print(f"DEBUG: لیست موقت خالی بود، ریدایرکت به view_project {project_id}")
        return redirect(url_for("view_project", project_id=project_id))

    project_info = get_project_details_db(project_id)
    if not project_info:
        flash(f"پروژه با ID {project_id} یافت نشد.", "error")
        session.pop(f"pending_doors_{project_id}", None)
        return redirect(url_for("index"))

    print(
        f"DEBUG: شروع ذخیره {len(pending_doors)} درب از لیست موقت برای پروژه {project_id}..."
    )
    
    saved_count, error_count = save_doors_batch_db(project_id, pending_doors)

    # <-- کلید session منحصر به فرد
    session.pop(f"pending_doors_{project_id}", None)
    print(f"DEBUG: لیست موقت پروژه {project_id} از session پاک شد.")

    if error_count == 0:
        flash(f"{saved_count} درب با موفقیت در دیتابیس ذخیره شد.", "success")
    else:
        flash(
            f"{saved_count} درب ذخیره شد، اما در ذخیره {error_count} درب خطا رخ داد.",
            "error",
        )

    target_url = url_for("view_project", project_id=project_id)
    print(f"DEBUG: ذخیره نهایی انجام شد. ریدایرکت به: {target_url}")
    return redirect(target_url)


def initialize_visible_columns(project_id):
    """Load durable per-project field choices into the legacy session cache."""
    session_key = f"visible_columns_{project_id}"
    base_column_keys = ["location", "width", "height", "quantity", "direction"]
    selected_columns = get_project_visible_custom_columns(project_id)
    session[session_key] = base_column_keys + [col["key"] for col in selected_columns]
    session.modified = True
    return session[session_key]


def refresh_project_visible_columns(project_id):
    """Refresh the cache without undoing a user's per-project hide choice."""
    return initialize_visible_columns(project_id)


@app.route("/project/<int:project_id>/treeview")
def project_treeview(project_id):
    """نمایش درب‌های پروژه در قالب TreeView پیشرفته"""
    print(f"DEBUG: ++++ شروع روت project_treeview برای پروژه {project_id}")
    
    # برای اطمینان از عدم کش‌شدن، یک پارامتر زمانی اضافه کنیم
    refresh_param = int(time.time())
    print(f"DEBUG: پارامتر زمانی برای جلوگیری از کش: {refresh_param}")
    
    project_info = get_project_details_db(project_id)
    if not project_info:
        flash(f"پروژه با ID {project_id} یافت نشد.", "error")
        return redirect(url_for("index"))

    session_key = f"visible_columns_{project_id}"
    # پارامتر force_refresh از URL می‌آید و نشان می‌دهد که آیا باید session به‌روز شود یا خیر
    force_refresh_session = request.args.get("force_refresh", "0") == "1" 

    # The database is authoritative so the same project choices are visible to
    # every assigned worker and survive logouts or browser changes.
    refresh_project_visible_columns(project_id)
    
    visible_columns = session.get(session_key, [])
    
    # حذف بخش مربوط به if refresh_columns: چون دیگر نیازی به آن نیست.
    # تابع refresh_project_visible_columns مسئول به‌روزرسانی لیست بر اساس داده‌های واقعی است.

    print(f"DEBUG: ستون‌های نمایشی نهایی برای رندر در project_treeview: {visible_columns}")
    
    # درب‌ها را از دیتابیس دریافت می‌کنیم 
    doors = get_doors_for_project_db(project_id)
    print(f"DEBUG: دریافت {len(doors)} درب برای پروژه {project_id}")

    # اندازه‌ها همیشه به سانتی‌متر ذخیره می‌شوند، اما در این صفحه باید با
    # واحدی نمایش داده شوند که هنگام ساخت پروژه انتخاب شده است.
    project_measurement_unit = normalize_measurement_unit(
        project_info.get("measurement_unit", "cm")
    )
    unit_labels = measurement_unit_labels(project_measurement_unit)
    for door in doors:
        door["display_width"] = (
            format_measurement_value(
                centimeters_to_measurement_unit(door["width"], project_measurement_unit)
            )
            if door.get("width") is not None
            else ""
        )

        door["display_height"] = (
            format_measurement_value(
                centimeters_to_measurement_unit(door["height"], project_measurement_unit)
            )
            if door.get("height") is not None
            else ""
        )
        door["hardware_summary"] = hardware_summary(door)

    total_door_quantity = sum(int(door.get("quantity") or 0) for door in doors)
    
    # دریافت ستون‌های سفارشی فعال
    active_custom_columns = get_project_visible_custom_columns(project_id)
    
    # بررسی سریع مقادیر سفارشی
    for door in doors[:5]:  # فقط 5 درب اول را برای دیباگ بررسی می‌کنیم
        print(f"DEBUG: درب {door['id']} - مقادیر سفارشی: {door}")
    
    print("-" * 50)
    print(f"DEBUG (treeview): Preparing to render for project_id: {project_id}")
    print(f"DEBUG (treeview): Visible columns from session: {visible_columns}")
    print(f"DEBUG (treeview): Doors list from DB: {doors}")
    print(f"DEBUG (treeview): Active custom columns list: {active_custom_columns}")
    print("-" * 50)
    
    return render_template(
        "project_treeview.html", 
        project=project_info, 
        doors=doors, 
        refresh_param=refresh_param,
        visible_columns=visible_columns,
        active_custom_columns=active_custom_columns,
        measurement_unit_label=unit_labels["fa"],
        total_door_quantity=total_door_quantity,
        # پارامترهای force_refresh و refresh_columns دیگر به تمپلیت پاس داده نمی‌شوند
    )


@app.route("/project/<int:project_id>/delete_door/<int:door_id>", methods=["POST"])
@csrf_protected
@staff_or_admin_required
def delete_door(project_id, door_id):
    """حذف یک درب از پروژه"""
    print(f"DEBUG: درخواست برای حذف درب با ID {door_id} از پروژه {project_id}")
    
    # اتصال به دیتابیس
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ابتدا بررسی می‌کنیم که آیا درب متعلق به این پروژه است
        cursor.execute(
            "SELECT id FROM doors WHERE id = ? AND project_id = ?",
            (door_id, project_id),
        )
        door = cursor.fetchone()
        
        if not door:
            print(f"ERROR: درب با ID {door_id} در پروژه {project_id} یافت نشد")
            return jsonify({"success": False, "error": "درب مورد نظر یافت نشد"}), 404
        
        # حذف مقادیر ستون‌های سفارشی مربوط به این درب
        cursor.execute("DELETE FROM door_custom_values WHERE door_id = ?", (door_id,))
        
        # حذف درب از جدول اصلی
        cursor.execute("DELETE FROM doors WHERE id = ?", (door_id,))
        
        conn.commit()
        print(f"DEBUG: درب با ID {door_id} با موفقیت حذف شد")
        return jsonify({"success": True})
    except sqlite3.Error as e:
        print(f"ERROR: خطا در حذف درب: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@app.route("/hardware/settings", methods=["GET"])
@staff_or_admin_required
def hardware_catalog_settings():
    """Manage the global dropdown choices used by structured hardware forms."""
    return render_template(
        "hardware_catalog_settings.html",
        categories=HARDWARE_CATALOG_CATEGORIES,
        hardware_options=get_hardware_catalog_options(),
        profile_brackets=get_profile_bracket_settings(),
    )


@app.route("/hardware/settings/profile-brackets/<int:profile_id>", methods=["POST"])
@csrf_protected
@staff_or_admin_required
def hardware_profile_bracket_update(profile_id):
    bracket_name = request.form.get("bracket_name", "")
    if update_profile_bracket_setting(profile_id, bracket_name):
        flash("عنوان براکت این پروفیل ذخیره شد.", "success")
    else:
        flash("عنوان براکت معتبر نیست یا پروفیل پیدا نشد.", "error")
    return redirect(url_for("hardware_catalog_settings", section="profile-brackets"))


@app.route("/hardware/settings/options/add", methods=["POST"])
@csrf_protected
@staff_or_admin_required
def hardware_catalog_add():
    category = str(request.form.get("category", "")).strip()
    value = request.form.get("value", "")
    success, message = add_hardware_catalog_option(category, value)
    flash(message, "success" if success else "error")
    return redirect(url_for("hardware_catalog_settings", category=category))


@app.route("/hardware/settings/options/<int:option_id>/archive", methods=["POST"])
@csrf_protected
@staff_or_admin_required
def hardware_catalog_archive(option_id):
    category = str(request.form.get("category", "")).strip()
    if archive_hardware_catalog_option(option_id):
        flash("گزینه از فهرست‌های جدید برداشته شد؛ سفارش‌های قبلی محفوظ‌اند.", "success")
    else:
        flash("برداشتن گزینه انجام نشد.", "error")
    return redirect(url_for("hardware_catalog_settings", category=category))


@app.route("/hardware/settings/options/<int:option_id>/move", methods=["POST"])
@csrf_protected
@staff_or_admin_required
def hardware_catalog_move(option_id):
    category = str(request.form.get("category", "")).strip()
    direction = str(request.form.get("direction", ""))
    if not move_hardware_catalog_option(option_id, direction):
        flash("جابه‌جایی گزینه انجام نشد.", "warning")
    return redirect(url_for("hardware_catalog_settings", category=category))


def _build_factory_report(project_info, doors):
    profile_labels = {
        item["name"]: item["bracket_name"]
        for item in get_profile_bracket_settings(include_inactive=True)
    }
    report = calculate_factory_requirements(doors, profile_labels)
    measurement_unit = normalize_measurement_unit(
        project_info.get("measurement_unit", "cm")
    )
    unit_labels = measurement_unit_labels(measurement_unit)
    for row in report["details"]:
        row["display_width"] = format_measurement_value(
            centimeters_to_measurement_unit(row["width"], measurement_unit)
        )
        row["display_height"] = format_measurement_value(
            centimeters_to_measurement_unit(row["height"], measurement_unit)
        )
    return report, unit_labels


@app.route("/project/<int:project_id>/factory-requirements", methods=["GET"])
def factory_requirements_report(project_id):
    """Show the factory-only rubber and installation-bracket list."""
    project_info = get_project_details_db(project_id)
    if not project_info:
        flash("پروژه مورد نظر یافت نشد.", "error")
        return redirect(url_for("index"))
    doors = get_doors_for_project_db(project_id)
    if not doors:
        flash("هیچ دربی برای این پروژه ثبت نشده است.", "warning")
        return redirect(url_for("project_treeview", project_id=project_id))
    report, unit_labels = _build_factory_report(project_info, doors)
    return render_template(
        "factory_requirements.html",
        project=project_info,
        report=report,
        measurement_unit_label=unit_labels["short"],
    )


@app.route("/project/<int:project_id>/factory-requirements/excel", methods=["GET"])
def export_factory_requirements_excel(project_id):
    """Download the factory rubber and separated bracket requirements."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    project_info = get_project_details_db(project_id)
    if not project_info:
        flash("پروژه مورد نظر یافت نشد.", "error")
        return redirect(url_for("index"))
    doors = get_doors_for_project_db(project_id)
    if not doors:
        flash("هیچ دربی برای این پروژه ثبت نشده است.", "warning")
        return redirect(url_for("project_treeview", project_id=project_id))
    report, unit_labels = _build_factory_report(project_info, doors)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "خلاصه کارخانه"
    details = workbook.create_sheet("جزئیات درب‌ها")
    warnings = workbook.create_sheet("موارد ناقص")
    header_fill = PatternFill("solid", fgColor="168F79")
    title_fill = PatternFill("solid", fgColor="DDF2ED")
    warning_fill = PatternFill("solid", fgColor="FFF0C2")
    white_font = Font(color="FFFFFF", bold=True)

    def prepare(sheet, widths):
        sheet.sheet_view.rightToLeft = True
        sheet.freeze_panes = "A5"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    title = project_info.get("customer_name") or f"پروژه {project_id}"
    summary.merge_cells("A1:C1")
    summary["A1"] = f"لیست کارخانه — {title}"
    summary["A1"].font = Font(bold=True, size=15)
    summary["A1"].fill = title_fill
    summary.append(["متراژ کل لاستیک", report["total_rubber_meters"], "متر"])
    summary.append(["تعداد کل براکت", report["total_bracket_count"], "عدد"])
    summary.append(["نوع براکت", "تعداد کل", "واحد"])
    for cell in summary[4]:
        cell.fill = header_fill
        cell.font = white_font
    for item in report["bracket_summary"]:
        summary.append([item["label"], item["quantity"], "عدد"])
    prepare(summary, [42, 18, 14])

    detail_headers = [
        "کد درب", f"عرض ({unit_labels['short']})",
        f"ارتفاع ({unit_labels['short']})", "تعداد درب", "نوع چارچوب",
        "نوع پروفیل", "متراژ لاستیک (متر)", "نوع براکت", "تعداد براکت",
    ]
    details.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(detail_headers))
    details["A1"] = f"جزئیات کارخانه — {title}"
    details["A1"].font = Font(bold=True, size=15)
    details["A1"].fill = title_fill
    details.append([])
    details.append([])
    details.append(detail_headers)
    for cell in details[4]:
        cell.fill = header_fill
        cell.font = white_font
    for row in report["details"]:
        details.append(
            [
                row["door_code"], row["display_width"], row["display_height"],
                row["quantity"], row["frame_type"], row["profile_name"],
                row["rubber_meters"], row["bracket_label"], row["bracket_count"],
            ]
        )
        if row["has_warning"]:
            for cell in details[details.max_row]:
                cell.fill = warning_fill
    prepare(details, [24, 13, 13, 12, 16, 24, 20, 38, 16])

    warnings.merge_cells("A1:C1")
    warnings["A1"] = f"موارد نیازمند بررسی — {title}"
    warnings["A1"].font = Font(bold=True, size=15)
    warnings["A1"].fill = warning_fill
    warnings.append([])
    warnings.append([])
    warnings.append(["شناسه درب", "کد درب", "پیام"])
    for cell in warnings[4]:
        cell.fill = header_fill
        cell.font = white_font
    for warning in report["warnings"]:
        warnings.append(
            [warning["door_id"], warning["door_code"], warning["message"]]
        )
    if not report["warnings"]:
        warnings.append(["-", "-", "اطلاعات کارخانه تمام درب‌ها کامل است."])
    prepare(warnings, [14, 24, 64])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"factory_project_{project_id}_{get_shamsi_timestamp()}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _build_installer_rows(project_info, doors):
    measurement_unit = normalize_measurement_unit(
        project_info.get("measurement_unit", "cm")
    )
    unit_labels = measurement_unit_labels(measurement_unit)
    rows = []
    for index, door in enumerate(doors, start=1):
        code = door.get("door_code") or f"D-{index:02d}"
        width = format_measurement_value(
            centimeters_to_measurement_unit(door.get("width"), measurement_unit)
        )
        height = format_measurement_value(
            centimeters_to_measurement_unit(door.get("height"), measurement_unit)
        )
        locations = door.get("installation_locations") or [
            {"location": door.get("location") or "مکان ثبت‌نشده", "quantity": door.get("quantity") or 1}
        ]
        for location in locations:
            rows.append(
                {
                    "door_code": code,
                    "location": location["location"],
                    "quantity": int(location["quantity"]),
                    "display_width": width,
                    "display_height": height,
                    "direction": door.get("direction") or "—",
                }
            )
    return rows, unit_labels


@app.route("/project/<int:project_id>/installer", methods=["GET"])
def installer_report(project_id):
    """Show code-to-location instructions for the installer."""
    project_info = get_project_details_db(project_id)
    if not project_info:
        flash("پروژه مورد نظر یافت نشد.", "error")
        return redirect(url_for("index"))
    doors = get_doors_for_project_db(project_id)
    if not doors:
        flash("هیچ دربی برای این پروژه ثبت نشده است.", "warning")
        return redirect(url_for("view_project", project_id=project_id))
    rows, unit_labels = _build_installer_rows(project_info, doors)
    return render_template(
        "installer_report.html",
        project=project_info,
        doors=doors,
        rows=rows,
        total_quantity=sum(row["quantity"] for row in rows),
        measurement_unit_label=unit_labels["fa"],
    )


@app.route("/project/<int:project_id>/installer/excel", methods=["GET"])
def export_installer_report_excel(project_id):
    """Download the installer code and location map."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    project_info = get_project_details_db(project_id)
    if not project_info:
        flash("پروژه مورد نظر یافت نشد.", "error")
        return redirect(url_for("index"))
    doors = get_doors_for_project_db(project_id)
    if not doors:
        flash("هیچ دربی برای این پروژه ثبت نشده است.", "warning")
        return redirect(url_for("view_project", project_id=project_id))
    rows, unit_labels = _build_installer_rows(project_info, doors)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "لیست نصاب"
    sheet.sheet_view.rightToLeft = True
    headers = [
        "کد درب", "محل نصب", "تعداد", f"عرض ({unit_labels['short']})",
        f"ارتفاع ({unit_labels['short']})", "جهت",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="168F79")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in rows:
        sheet.append(
            [
                row["door_code"], row["location"], row["quantity"],
                row["display_width"], row["display_height"], row["direction"],
            ]
        )
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate([16, 34, 12, 16, 16, 12], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"installer_project_{project_id}_{get_shamsi_timestamp()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/project/<int:project_id>/hardware", methods=["GET"])
def hardware_report(project_id):
    """Preview the office/warehouse hardware list for one project."""
    project_info = get_project_details_db(project_id)
    if not project_info:
        flash("پروژه مورد نظر یافت نشد.", "error")
        return redirect(url_for("index"))

    doors = get_doors_for_project_db(project_id)
    if not doors:
        flash("هیچ دربی برای این پروژه ثبت نشده است.", "warning")
        return redirect(url_for("project_treeview", project_id=project_id))

    return render_template(
        "hardware_report.html",
        project=project_info,
        report=calculate_project_hardware(doors),
    )


@app.route("/project/<int:project_id>/hardware/excel", methods=["GET"])
def export_hardware_report_excel(project_id):
    """Download a two-sheet hardware report without writing a server-side file."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    project_info = get_project_details_db(project_id)
    if not project_info:
        flash("پروژه مورد نظر یافت نشد.", "error")
        return redirect(url_for("index"))

    doors = get_doors_for_project_db(project_id)
    if not doors:
        flash("هیچ دربی برای این پروژه ثبت نشده است.", "warning")
        return redirect(url_for("project_treeview", project_id=project_id))

    report = calculate_project_hardware(doors)
    project_measurement_unit = normalize_measurement_unit(
        project_info.get("measurement_unit", "cm")
    )
    unit_labels = measurement_unit_labels(project_measurement_unit)

    def export_length(value):
        return centimeters_to_measurement_unit(value, project_measurement_unit)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "خلاصه خرید"
    detail_sheet = workbook.create_sheet("جزئیات درب‌ها")
    warning_sheet = workbook.create_sheet("موارد ناقص")

    header_fill = PatternFill("solid", fgColor="168F79")
    title_fill = PatternFill("solid", fgColor="DDF2ED")
    warning_fill = PatternFill("solid", fgColor="FFF0C2")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="D5DEE2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_table(sheet, widths):
        sheet.sheet_view.rightToLeft = True
        sheet.freeze_panes = "A5"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    project_title = project_info.get("customer_name") or f"پروژه {project_id}"
    project_code = project_info.get("project_code") or "-"

    summary_sheet.merge_cells("A1:D1")
    summary_sheet["A1"] = f"لیست یراق پروژه — {project_title}"
    summary_sheet["A1"].font = Font(bold=True, size=15)
    summary_sheet["A1"].fill = title_fill
    summary_sheet["A2"] = "کد پروژه"
    summary_sheet["B2"] = project_code
    summary_sheet["C2"] = "تعداد درب محاسبه‌شده"
    summary_sheet["D2"] = report["included_door_count"]
    summary_sheet.append([])
    summary_sheet.append(["گروه", "مدل", "تعداد کل", "واحد"])
    for cell in summary_sheet[4]:
        cell.fill = header_fill
        cell.font = white_font
    for item in report["summary"]:
        summary_sheet.append([item["group"], item["model"], item["quantity"], item["unit"]])
    style_table(summary_sheet, [18, 32, 15, 12])

    detail_headers = [
        "کد درب",
        f"عرض ({unit_labels['short']})",
        f"ارتفاع ({unit_labels['short']})",
        "تعداد درب", "مدل لولا", "تعداد لولا",
        "مدل قفل", "تعداد قفل", "مدل دستگیره", "تعداد دستگیره", "تعداد سیلندر",
    ]
    detail_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(detail_headers))
    detail_sheet["A1"] = f"جزئیات یراق درب‌ها — {project_title}"
    detail_sheet["A1"].font = Font(bold=True, size=15)
    detail_sheet["A1"].fill = title_fill
    detail_sheet.append([])
    detail_sheet.append([])
    detail_sheet.append(detail_headers)
    for cell in detail_sheet[4]:
        cell.fill = header_fill
        cell.font = white_font
    for row in report["details"]:
        detail_sheet.append(
            [
                row["door_code"], export_length(row["width"]), export_length(row["height"]),
                row["quantity"], row["hinge_model"],
                row["hinge_count"], row["lock_model"], row["lock_count"], row["handle_model"],
                row["handle_count"], row["cylinder_count"],
            ]
        )
        if row["has_warning"]:
            for cell in detail_sheet[detail_sheet.max_row]:
                cell.fill = warning_fill
    style_table(detail_sheet, [24, 12, 12, 12, 22, 13, 22, 13, 24, 15, 15])

    warning_sheet.merge_cells("A1:D1")
    warning_sheet["A1"] = f"موارد نیازمند بررسی — {project_title}"
    warning_sheet["A1"].font = Font(bold=True, size=15)
    warning_sheet["A1"].fill = warning_fill
    warning_sheet.append([])
    warning_sheet.append([])
    warning_sheet.append(["شناسه درب", "کد درب", "فیلد", "پیام"])
    for cell in warning_sheet[4]:
        cell.fill = header_fill
        cell.font = white_font
    if report["warnings"]:
        for warning in report["warnings"]:
            warning_sheet.append(
                [warning["door_id"], warning["door_code"], warning["field"], warning["message"]]
            )
    else:
        warning_sheet.append(["-", "-", "-", "اطلاعات یراق تمام درب‌ها کامل است."])
    style_table(warning_sheet, [14, 24, 18, 62])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"hardware_project_{project_id}_{get_shamsi_timestamp()}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/project/<int:project_id>/export/excel", methods=["GET"])
def export_to_excel(project_id):
    """خروجی اکسل فرمت‌شده از داده‌های پروژه با استفاده از ستون‌های قابل نمایش"""
    try:
        import pandas as pd
        import os
        import jdatetime
        from datetime import datetime
        from flask import make_response
        import re
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        print(f"DEBUG: شروع صدور اکسل پیشرفته برای پروژه {project_id}")

        # دریافت اطلاعات پروژه
        project_info = get_project_details_db(project_id)
        if not project_info:
            print("DEBUG: پروژه یافت نشد")
            flash("پروژه مورد نظر یافت نشد.", "error")
            return redirect(url_for("index"))

        application_status = get_inventory_cutting_application_status(project_id)
        applied_plan_snapshot, export_warning = resolve_applied_cutting_plan(
            application_status
        )

        if export_warning and request.args.get("confirm_without_cutting") != "1":
            return render_template(
                "confirm_excel_without_cutting.html",
                project=project_info,
                warning_message=export_warning,
            )

        # دریافت داده‌های درب‌ها
        doors = get_doors_for_project_db(project_id)
        if not doors:
            print("DEBUG: هیچ دربی یافت نشد")
            flash("هیچ دربی برای این پروژه ثبت نشده است.", "warning")
            return redirect(url_for("project_treeview", project_id=project_id))
        
        print(f"DEBUG: {len(doors)} درب برای تبدیل به اکسل یافت شد")

        project_measurement_unit = normalize_measurement_unit(
            project_info.get("measurement_unit", "cm")
        )
        unit_labels = measurement_unit_labels(project_measurement_unit)

        def export_length(value):
            return centimeters_to_measurement_unit(value, project_measurement_unit)

        # دریافت ستون‌های قابل نمایش از session
        session_key = f"visible_columns_{project_id}"
        visible_columns = session.get(session_key, [])
        
        # اگر هیچ ستونی برای نمایش انتخاب نشده، ستون‌های پیش‌فرض را نمایش می‌دهیم
        if not visible_columns:
            # اجرای تابع مقداردهی اولیه
            initialize_visible_columns(project_id)
            # بازخوانی مجدد از سشن
            visible_columns = session.get(session_key, [])
        
        # اضافه کردن ستون‌های پایه که همیشه باید نمایش داده شوند
        basic_columns = ["location", "width", "height", "quantity", "direction"]
        for col in basic_columns:
            if col not in visible_columns:
                visible_columns.append(col)
        
        print(f"DEBUG: ستون‌های نمایشی برای اکسل: {visible_columns}")

        # ایجاد ترجمه فارسی برای نام ستون‌ها
        column_translations = {
            "id": "شماره ردیف",
            "location": "موقعیت",
            "width": f"عرض ({unit_labels['short']})",
            "height": f"ارتفاع ({unit_labels['short']})",
            "quantity": "تعداد درب",
            "direction": "جهت",
            "rang": "رنگ پروفیل آلومینیوم",
            "noe_profile": "نوع پروفیل",
            "vaziat": "وضعیت تولید درب",
            "lola": "نوع لولا",
            "ghofl": "نوع قفل",
            "accessory": "اکسسوری",
            "kolaft": "نوع چارچوب",
            "dastgire": "نوع دستگیره",
            "tozihat": "توضیحات"
        }
        
        # دریافت ستون‌های سفارشی فعال برای ترجمه نام‌ها
        custom_columns = get_all_custom_columns()
        for col in custom_columns:
            if col["key"] not in column_translations:
                column_translations[col["key"]] = col["display"]
        
        # ایجاد یک workbook جدید
        wb = Workbook()
        ws = wb.active
        ws.title = f"پروژه {project_id}"
        
        # استایل‌های مختلف برای سلول‌ها
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # آبی برای هدر
        alt_row_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")  # آبی کمرنگ برای ردیف‌های زوج
        
        # تنظیم استایل برای مرزها
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # نام مشتری و تاریخ در بالای اکسل - اصلاح بخش ادغام سلول‌ها
        # فرمت تاریخ شمسی به فارسی
        now_jalali = jdatetime.datetime.now()
        # نام ماه‌های فارسی
        persian_months = {
            1: 'فروردین', 2: 'اردیبهشت', 3: 'خرداد', 4: 'تیر', 5: 'مرداد', 6: 'شهریور',
            7: 'مهر', 8: 'آبان', 9: 'آذر', 10: 'دی', 11: 'بهمن', 12: 'اسفند'
        }
        # نام روزهای هفته فارسی
        persian_weekdays = {
            0: 'شنبه', 1: 'یکشنبه', 2: 'دوشنبه', 3: 'سه‌شنبه', 
            4: 'چهارشنبه', 5: 'پنج‌شنبه', 6: 'جمعه'
        }
        weekday_name = persian_weekdays.get(now_jalali.weekday(), '')
        month_name = persian_months.get(now_jalali.month, '')
        today_jalali = f"{weekday_name}، {now_jalali.day} {month_name} {now_jalali.year}"
        
        # دریافت اطلاعات پروژه
        customer_name = project_info.get("customer_name", "")
        order_ref = project_info.get("order_ref", "")
        
        # ردیف 1: تاریخ
        ws['A1'] = "تاریخ"
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws.merge_cells('A1:B1')
        
        ws['C1'] = today_jalali
        ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].font = Font(bold=True, size=11)
        ws['C1'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws.merge_cells('C1:E1')
        
        # ردیف 2: نام پروژه
        ws['A2'] = "نام پروژه"
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
        ws.merge_cells('A2:B2')
        
        ws['C2'] = customer_name if customer_name else "نامشخص"
        ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C2'].font = Font(bold=True, size=11)
        ws['C2'].fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
        ws.merge_cells('C2:E2')
        
        # ردیف 3: شماره سفارش و کد پروژه
        ws['A3'] = "شماره سفارش"
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws.merge_cells('A3:B3')
        
        ws['C3'] = order_ref if order_ref else "ندارد"
        ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C3'].font = Font(bold=True, size=11)
        ws['C3'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws.merge_cells('C3:E3')
        
        # فاصله بین جدول اصلی و هدر
        row_offset = 4
        
        # --- شروع ستون‌های نمایشی ---
        # ابتدا ستون شماره ردیف را به عنوان اولین ستون اضافه می‌کنیم
        visible_columns_with_translations = [{"key": "row_num", "display": "شماره ردیف"}]
        
        # اضافه کردن سایر ستون‌ها براساس لیست visible_columns
        for col_key in visible_columns:
            display_name = column_translations.get(col_key, col_key)
            visible_columns_with_translations.append({"key": col_key, "display": display_name})
        
        # درج هدر ستون‌ها
        for col_idx, col_info in enumerate(visible_columns_with_translations, 1):
            cell = ws.cell(row=row_offset+1, column=col_idx, value=col_info["display"])
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = header_fill
            cell.border = thin_border
            # تنظیم عرض ستون
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # درج داده‌های درب‌ها
        for row_idx, door in enumerate(doors, 1):
            # رنگ پس‌زمینه برای ردیف‌های زوج
            row_fill = alt_row_fill if row_idx % 2 == 0 else None
            
            # برای هر ستون قابل نمایش
            for col_idx, col_info in enumerate(visible_columns_with_translations, 1):
                col_key = col_info["key"]
                
                # مقدار ستون
                if col_key == "row_num":
                    value = row_idx
                else:
                    value = door.get(col_key, "")
                
                # تبدیل اعداد از string به عدد برای نمایش بهتر
                if col_key in ["width", "height"] and value:
                    try:
                        value = export_length(value)
                    except (ValueError, TypeError):
                        pass
                elif col_key == "quantity" and value:
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        pass
                
                # درج سلول
                cell = ws.cell(row=row_offset+1+row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                
                # اعمال رنگ پس‌زمینه برای ردیف‌های زوج
                if row_fill:
                    cell.fill = row_fill
                
                # قالب‌بندی خاص برای سلول‌های خاص
                if col_key == "vaziat" and value and "درآینده" in str(value):
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # قرمز کمرنگ
        
        # The cutting sheet must represent the exact plan that was committed to
        # inventory, never a fresh recalculation against already-mutated stock.
        if applied_plan_snapshot is not None:
            add_cutting_results_sheet(
                wb, applied_plan_snapshot, project_measurement_unit
            )

        # تنظیم مسیر فایل خروجی
        export_dir = "static/exports"
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
        
        timestamp = get_shamsi_timestamp()  # تاریخ شمسی
        customer_name = project_info.get("customer_name", "unknown")
        
        # ایجاد نام فایل با حروف انگلیسی برای ذخیره سازی
        safe_filename = f"project_{project_id}_{timestamp}.xlsx"
        excel_path = os.path.join(export_dir, safe_filename)
        
        # ذخیره به فایل اکسل
        wb.save(excel_path)
        print(f"DEBUG: فایل اکسل با فرمت جدید با موفقیت ذخیره شد در: {excel_path}")
        
        # اصلاح شده: استفاده از نام فایل انگلیسی برای هدر Content-Disposition
        # حذف کاراکترهای فارسی و غیرمجاز از نام فایل
        display_filename = f"project_{project_id}_{timestamp}.xlsx"
        order_ref = project_info.get("order_ref", "")
        
        # اگر شماره سفارش موجود است، از آن استفاده کنیم (اگر شامل حروف لاتین است)
        if order_ref and any(c.isalnum() and ord(c) < 128 for c in order_ref):
            # فقط کاراکترهای مجاز را نگه داریم
            safe_order_ref = ''.join(c for c in order_ref if c.isalnum() or c in '-_')
            if safe_order_ref:
                display_filename = f"project_{project_id}_{safe_order_ref}.xlsx"
        
        print(f"DEBUG: نام فایل دانلودی: {display_filename}")
        
        # ارسال فایل به کاربر با هدرهای مناسب برای نمایش دیالوگ "ذخیره به عنوان"
        response = make_response(send_file(excel_path, as_attachment=True))
        response.headers["Content-Disposition"] = f"attachment; filename={display_filename}"
        return response
        
    except Exception as e:
        print(f"ERROR در صدور اکسل: {e}")
        traceback.print_exc()
        flash(f"خطا در ایجاد فایل اکسل: {str(e)}", "error")
        return redirect(url_for("project_treeview", project_id=project_id))


@app.route("/project/<int:project_id>/calculate_cutting", methods=["GET"])
def calculate_cutting(project_id):
    """محاسبه طرح برش و گزارش وزن بر اساس مشخصات واقعی هر پروفیل."""
    stock_length = 600

    project_info = get_project_details_db(project_id)
    if not project_info:
        flash("پروژه مورد نظر یافت نشد.", "error")
        return redirect(url_for("index"))

    doors = get_doors_for_project_db(project_id)
    if not doors:
        flash("هیچ دربی برای این پروژه ثبت نشده است.", "warning")
        return redirect(url_for("view_project", project_id=project_id))

    settings = get_inventory_settings()
    use_inventory = settings.get("use_inventory_for_cutting", False)
    prefer_inventory_pieces = settings.get("prefer_inventory_pieces", False)
    optimization_strategy = settings.get("inventory_optimization_strategy", "minimize_waste")
    profiles = get_all_profile_types()

    available_pieces_by_profile = {}
    if use_inventory:
        profile_variants = {
            (str(door.get("noe_profile") or "").strip(), normalize_color_name(door.get("rang")))
            for door in doors
            if str(door.get("noe_profile") or "").strip()
        }
        available_pieces_by_profile = {
            make_inventory_variant_key(profile_name, color_name):
                get_available_inventory_pieces(profile_name, color_name)
            for profile_name, color_name in profile_variants
        }

    try:
        plan = build_cutting_plan(
            doors,
            profiles,
            available_pieces_by_profile=available_pieces_by_profile,
            use_inventory=use_inventory,
            prefer_inventory_pieces=prefer_inventory_pieces,
            optimization_strategy=optimization_strategy,
            stock_length=stock_length,
        )
    except CuttingPlanError as exc:
        flash(str(exc), "error")
        return redirect(url_for("view_project", project_id=project_id))

    if plan["invalid_rows"]:
        flash(
            f'{len(plan["invalid_rows"])} ردیف به دلیل عرض، ارتفاع، تعداد یا نوع پروفیل نامعتبر '
            "در محاسبه نادیده گرفته شد.",
            "warning",
        )

    session[f"cutting_result_{project_id}"] = {
        "profile_requirements": plan["inventory_application_data"],
        "stock_length": stock_length,
        "timestamp": get_shamsi_datetime_iso(),
        "used_inventory_pieces": plan["used_inventory_pieces"],
        "fingerprint": plan["fingerprint"],
    }

    stats = plan["stats"]
    return render_template(
        "cutting_result.html",
        project=project_info,
        bins=plan["processed_bins"],
        profile_summaries=plan["profile_summaries"],
        total_bins=plan["total_bins"],
        stock_length=stock_length,
        discarded_count=stats["discarded_count"],
        discarded_length=round(stats["discarded_length"], 1),
        discarded_weight=round(stats["discarded_weight"], 2),
        reusable_count=stats["reusable_count"],
        reusable_length=round(stats["reusable_length"], 1),
        reusable_weight=round(stats["reusable_weight"], 2),
        total_remaining_length=round(stats["total_remaining_length"], 1),
        total_remaining_weight=round(stats["total_remaining_weight"], 2),
        total_remaining_percentage=round(stats["total_remaining_percentage"], 1),
        blade_width_mm=round(plan["blade_width"] * 10, 1),
        total_kerf_length=round(stats["total_kerf_length"], 1),
        optimization_strategy=plan["optimization_strategy"],
    )

@app.route("/project/<int:project_id>/apply_cutting_plan", methods=["POST"])
def apply_cutting_plan(project_id):
    """Compatibility entrypoint: persist an order; never deduct stock directly."""
    if not user_can_edit_project(current_user.id, current_user.role, project_id):
        flash("فقط مسئول این پروژه یا مدیر می‌تواند سفارش برش بسازد.", "error")
        return redirect(url_for("view_project", project_id=project_id))
    try:
        order_id = create_cutting_order([project_id], current_user.id)
        session.pop(f"cutting_result_{project_id}", None)
        flash(
            "سفارش برش ثبت شد؛ هنوز هیچ موجودی رزرو یا کم نشده است.",
            "success",
        )
        return redirect(url_for("cutting_order_details", order_id=order_id))
    except CuttingOrderError as exc:
        flash(str(exc), "error")
        return redirect(url_for("calculate_cutting", project_id=project_id))

    # The legacy direct-deduction implementation intentionally remains below for
    # audit/reference compatibility, but is unreachable. All new physical
    # consumption must pass through reserve -> send -> per-bar confirmation.
    project_info = get_project_details_db(project_id)
    if not project_info:
        flash("پروژه مورد نظر یافت نشد.", "error")
        return redirect(url_for("index"))

    application_status = get_inventory_cutting_application_status(project_id)
    if application_status["status"] == "completed":
        application = application_status["application"]
        flash(
            "این پروژه قبلاً به‌طور کامل از انبار کسر شده است.\n"
            f"تاریخ ثبت: {application['applied_at']}\n"
            f"مجموع شاخه‌های کسرشده: {application['total_stock_deducted']}\n"
            f"قطعات مصرف‌شده: {application['pieces_consumed']}",
            "warning",
        )
        return redirect(url_for("view_project", project_id=project_id))

    if application_status["status"] == "legacy_unverified":
        deduction_details = "\n".join(
            f"• {item['profile_name']}: {item['quantity_deducted']} شاخه در تاریخ {item['deduction_date']}"
            for item in application_status["deductions"]
        )
        flash(
            "برای این پروژه سابقه کسر قدیمی یا نیمه‌کاره وجود دارد. برای جلوگیری از کسر دوباره، "
            "برنامه هیچ تغییری در انبار ایجاد نکرد.\n"
            "سوابق موجود:\n"
            f"{deduction_details}\n"
            "این سابقه باید جداگانه بازبینی و تعیین تکلیف شود.",
            "warning",
        )
        return redirect(url_for("view_project", project_id=project_id))

    if application_status["status"] == "error":
        flash("بررسی وضعیت قبلی کسر انبار انجام نشد. لطفاً دوباره تلاش کنید.", "error")
        return redirect(url_for("view_project", project_id=project_id))

    cutting_data = session.get(f"cutting_result_{project_id}")
    if not cutting_data:
        flash("ابتدا باید محاسبه برش را انجام دهید.", "warning")
        return redirect(url_for("calculate_cutting", project_id=project_id))

    profile_requirements = cutting_data.get("profile_requirements", {})
    used_inventory_pieces = cutting_data.get("used_inventory_pieces", {})
    if not profile_requirements:
        flash("اطلاعات پروفیل‌های مورد نیاز یافت نشد.", "error")
        return redirect(url_for("calculate_cutting", project_id=project_id))

    # Rebuild the plan immediately before application. If the order, profile
    # settings, cutting settings, or available offcuts changed after the report
    # was shown, applying that old report would make inventory unreliable.
    doors = get_doors_for_project_db(project_id)
    settings = get_inventory_settings()
    use_inventory = settings.get("use_inventory_for_cutting", False)
    prefer_inventory_pieces = settings.get("prefer_inventory_pieces", False)
    optimization_strategy = settings.get(
        "inventory_optimization_strategy", "minimize_waste"
    )
    profiles = get_all_profile_types()
    available_pieces_by_profile = {}
    if use_inventory:
        profile_variants = {
            (str(door.get("noe_profile") or "").strip(), normalize_color_name(door.get("rang")))
            for door in doors
            if str(door.get("noe_profile") or "").strip()
        }
        available_pieces_by_profile = {
            make_inventory_variant_key(profile_name, color_name):
                get_available_inventory_pieces(profile_name, color_name)
            for profile_name, color_name in profile_variants
        }

    try:
        current_plan = build_cutting_plan(
            doors,
            profiles,
            available_pieces_by_profile=available_pieces_by_profile,
            use_inventory=use_inventory,
            prefer_inventory_pieces=prefer_inventory_pieces,
            optimization_strategy=optimization_strategy,
            stock_length=600,
        )
    except CuttingPlanError as exc:
        session.pop(f"cutting_result_{project_id}", None)
        flash(
            f"طرح قبلی دیگر قابل اعمال نیست: {exc} لطفاً گزارش برش را دوباره محاسبه کنید.",
            "error",
        )
        return redirect(url_for("view_project", project_id=project_id))

    if (
        not cutting_data.get("fingerprint")
        or cutting_data["fingerprint"] != current_plan["fingerprint"]
    ):
        session.pop(f"cutting_result_{project_id}", None)
        flash(
            "پس از محاسبه، اطلاعات سفارش، پروفیل، تنظیمات برش یا موجودی تغییر کرده است. "
            "برای جلوگیری از کسر اشتباه، چیزی از انبار کم نشد؛ لطفاً گزارش برش را دوباره محاسبه کنید.",
            "warning",
        )
        return redirect(url_for("calculate_cutting", project_id=project_id))

    # Use only the freshly rebuilt server-side data for the inventory transaction.
    profile_requirements = current_plan["inventory_application_data"]
    used_inventory_pieces = current_plan["used_inventory_pieces"]

    result = apply_cutting_plan_inventory_transaction(
        project_id,
        project_info,
        profile_requirements,
        used_inventory_pieces,
        actor_user_id=current_user.id,
        plan_snapshot=create_cutting_plan_snapshot(current_plan),
    )

    if result["status"] == "success":
        messages = []
        for item in result["profile_results"]:
            details = []
            if item["stock_deducted"]:
                details.append(f"{item['stock_deducted']} شاخه کسر شد")
            if item["pieces_consumed"]:
                details.append(f"{item['pieces_consumed']} قطعه موجود مصرف شد")
            if item["pieces_returned"]:
                details.append(
                    f"{item['pieces_returned']} باقی‌مانده قابل‌استفاده به انبار برگشت"
                )
            if item["pieces_discarded"]:
                details.append(
                    f"{item['pieces_discarded']} ضایعات در انبار ضایعات ثبت شد"
                )
            messages.append(
                f"✓ {item['profile_name']} — {item.get('color_name', 'تعیین‌نشده')}: "
                f"{', '.join(details) or 'ثبت شد'}"
            )

        session.pop(f"cutting_result_{project_id}", None)
        flash("طرح برش به‌طور کامل در انبار اعمال شد.\n" + "\n".join(messages), "success")
        return redirect(url_for("view_project", project_id=project_id))

    if result["status"] == "validation_error":
        flash(
            "طرح برش اعمال نشد و موجودی هیچ‌کدام از پروفیل‌ها تغییر نکرد.\n"
            + "\n".join(f"• {error}" for error in result["errors"]),
            "error",
        )
        return redirect(url_for("calculate_cutting", project_id=project_id))

    if result["status"] == "already_applied":
        flash("این پروژه قبلاً به‌طور کامل از انبار کسر شده است.", "warning")
        return redirect(url_for("view_project", project_id=project_id))

    if result["status"] == "legacy_unverified":
        flash(
            "برای این پروژه سابقه کسر قدیمی یا نیمه‌کاره وجود دارد. "
            "برای جلوگیری از کسر دوباره، هیچ تغییری در انبار ایجاد نشد.",
            "warning",
        )
        return redirect(url_for("view_project", project_id=project_id))

    flash(
        "به دلیل خطای پایگاه داده، طرح برش اعمال نشد و تمام تغییرات احتمالی بازگردانده شد.",
        "error",
    )
    return redirect(url_for("calculate_cutting", project_id=project_id))





def _parse_batch_door_ids(raw_value):
    """Parse a stable, duplicate-free list of positive door ids."""
    parsed = []
    seen = set()
    for raw_id in str(raw_value or "").split(","):
        raw_id = raw_id.strip()
        if not raw_id:
            continue
        try:
            door_id = int(raw_id)
        except (TypeError, ValueError):
            return None
        if door_id <= 0:
            return None
        if door_id not in seen:
            seen.add(door_id)
            parsed.append(door_id)
    return parsed


def _load_project_batch_doors(project_id, raw_door_ids):
    """Return selected doors only when every id belongs to this project."""
    door_ids = _parse_batch_door_ids(raw_door_ids)
    if not door_ids:
        return None, None
    project_doors = get_doors_for_project_db(project_id)
    project_door_map = {int(door["id"]): door for door in project_doors}
    if any(door_id not in project_door_map for door_id in door_ids):
        return door_ids, None
    row_numbers = {
        int(door["id"]): row_number
        for row_number, door in enumerate(project_doors, start=1)
    }
    selected_doors = []
    for door_id in door_ids:
        door = dict(project_door_map[door_id])
        door["row_number"] = row_numbers[door_id]
        selected_doors.append(door)
    return door_ids, selected_doors


@app.route("/project/<int:project_id>/batch_edit", methods=["GET"])
def batch_edit_form(project_id):
    """Show a safe, reviewable batch-edit form for selected project doors."""
    raw_door_ids = request.args.get("door_ids")
    if not raw_door_ids:
        flash("هیچ دربی برای ویرایش انتخاب نشده است.", "warning")
        return redirect(url_for("project_treeview", project_id=project_id))

    project_info = get_project_details_db(project_id)
    if not project_info:
        flash("سفارش مورد نظر پیدا نشد.", "error")
        return redirect(url_for("index"))

    door_ids, selected_doors = _load_project_batch_doors(project_id, raw_door_ids)
    if door_ids is None:
        flash("فهرست درب‌های انتخاب‌شده معتبر نیست.", "error")
        return redirect(url_for("project_treeview", project_id=project_id))
    if selected_doors is None:
        flash("حداقل یکی از درب‌های انتخاب‌شده متعلق به این سفارش نیست.", "error")
        return redirect(url_for("project_treeview", project_id=project_id))

    measurement_unit = normalize_measurement_unit(
        project_info.get("measurement_unit", "cm")
    )
    unit_labels = measurement_unit_labels(measurement_unit)
    for door in selected_doors:
        door["display_width"] = format_measurement_value(
            centimeters_to_measurement_unit(door.get("width"), measurement_unit)
        ) if door.get("width") is not None else ""
        door["display_height"] = format_measurement_value(
            centimeters_to_measurement_unit(door.get("height"), measurement_unit)
        ) if door.get("height") is not None else ""
        door["hardware_summary"] = hardware_summary(door)

    hardware_counts = {}
    for door in selected_doors:
        summary = door["hardware_summary"]
        hardware_counts[summary] = hardware_counts.get(summary, 0) + 1
    hardware_distribution = [
        {"value": value, "count": count}
        for value, count in hardware_counts.items()
    ]
    hardware_mixed = len(hardware_distribution) > 1
    if hardware_mixed:
        hardware_current_summary = "مقادیر متفاوت — " + "، ".join(
            f"{item['count']} {item['value']}" for item in hardware_distribution
        )
    else:
        hardware_current_summary = f"همه: {hardware_distribution[0]['value']}"

    bracket_counts = {}
    for door in selected_doors:
        value = (
            "براکت گوشتی"
            if door.get("installation_bracket_mode") == "meaty"
            else "خودکار از نوع پروفیل"
        )
        bracket_counts[value] = bracket_counts.get(value, 0) + 1
    bracket_distribution = [
        {"value": value, "count": count}
        for value, count in bracket_counts.items()
    ]
    bracket_mixed = len(bracket_distribution) > 1
    if bracket_mixed:
        bracket_current_summary = "مقادیر متفاوت — " + "، ".join(
            f"{item['count']} {item['value']}" for item in bracket_distribution
        )
    else:
        bracket_current_summary = f"همه: {bracket_distribution[0]['value']}"

    column_options = []
    for column in get_project_visible_custom_columns(project_id):
        options = []
        if column.get("type") == "dropdown":
            options = [
                option["value"]
                for option in get_custom_column_options(column["id"])
            ]

        counts = {}
        for door in selected_doors:
            value = str(door.get(column["key"]) or "").strip() or "بدون مقدار"
            counts[value] = counts.get(value, 0) + 1
        distribution = [
            {"value": value, "count": count}
            for value, count in counts.items()
        ]
        is_mixed = len(distribution) > 1
        if is_mixed:
            current_summary = "مقادیر متفاوت — " + "، ".join(
                f"{item['count']} {item['value']}" for item in distribution
            )
        else:
            current_summary = f"همه: {distribution[0]['value']}"

        column_options.append(
            {
                "key": column["key"],
                "display": column["display"],
                "type": column.get("type", "text"),
                "options": options,
                "mixed": is_mixed,
                "distribution": distribution,
                "current_summary": current_summary,
            }
        )

    return render_template(
        "batch_edit.html",
        project=project_info,
        door_ids=door_ids,
        selected_doors=selected_doors,
        column_options=column_options,
        hardware_mixed=hardware_mixed,
        hardware_distribution=hardware_distribution,
        hardware_current_summary=hardware_current_summary,
        bracket_mixed=bracket_mixed,
        bracket_distribution=bracket_distribution,
        bracket_current_summary=bracket_current_summary,
        hardware_options=get_hardware_catalog_options(),
        measurement_unit_label=unit_labels["fa"],
    )


@app.route("/project/<int:project_id>/batch_edit", methods=["POST"])
@csrf_protected
@staff_or_admin_required
def apply_batch_edit(project_id):
    """Apply reviewed batch changes only to doors belonging to this project."""
    raw_door_ids = request.form.get("door_ids")
    if not raw_door_ids:
        flash("هیچ دربی برای ویرایش انتخاب نشده است.", "warning")
        return redirect(url_for("project_treeview", project_id=project_id))

    door_ids, selected_doors = _load_project_batch_doors(project_id, raw_door_ids)
    if door_ids is None or selected_doors is None:
        flash("درب‌های انتخاب‌شده معتبر نیستند یا به این سفارش تعلق ندارند.", "error")
        return redirect(url_for("project_treeview", project_id=project_id))

    allowed_columns = {
        column["key"]: column
        for column in get_project_visible_custom_columns(project_id)
    }
    columns_to_update = {}
    mixed_fields = []
    for field_key in allowed_columns:
        if request.form.get(f"update_{field_key}") != "on":
            continue
        mode = request.form.get(f"mode_{field_key}", "set")
        if mode == "clear":
            new_value = ""
        else:
            new_value = str(request.form.get(f"value_{field_key}", "")).strip()
            if not new_value:
                flash("برای فیلد فعال‌شده مقدار جدیدی انتخاب نشده است.", "error")
                return redirect(
                    url_for(
                        "batch_edit_form",
                        project_id=project_id,
                        door_ids=",".join(str(door_id) for door_id in door_ids),
                    )
                )
        columns_to_update[field_key] = new_value
        current_values = {
            str(door.get(field_key) or "").strip() for door in selected_doors
        }
        if len(current_values) > 1:
            mixed_fields.append(field_key)

    hardware_to_update = None
    if request.form.get("update_hardware") == "on":
        try:
            hardware_to_update = normalize_door_hardware(
                {
                    "hinge_brand": request.form.get("hardware_hinge_brand"),
                    "hinge_color": request.form.get("hardware_hinge_color"),
                    "hinge_count": request.form.get("hardware_hinge_count"),
                    "has_handle": request.form.get("hardware_has_handle"),
                    "handle_type": request.form.get("hardware_handle_type"),
                    "handle_brand": request.form.get("hardware_handle_brand"),
                    "handle_model": request.form.get("hardware_handle_model"),
                    "handle_color": request.form.get("hardware_handle_color"),
                    "lock_source": request.form.get("hardware_lock_source"),
                    "lock_brand": request.form.get("hardware_lock_brand"),
                    "lock_model": request.form.get("hardware_lock_model"),
                    "cylinder_brand": request.form.get("hardware_cylinder_brand"),
                    "cylinder_model": request.form.get("hardware_cylinder_model"),
                }
            )
        except HardwareValidationError as exc:
            flash(str(exc), "error")
            return redirect(
                url_for(
                    "batch_edit_form",
                    project_id=project_id,
                    door_ids=",".join(str(door_id) for door_id in door_ids),
                )
            )
        current_hardware = {hardware_summary(door) for door in selected_doors}
        if len(current_hardware) > 1:
            mixed_fields.append("__hardware__")

    bracket_mode_to_update = None
    if request.form.get("update_bracket_mode") == "on":
        try:
            bracket_mode_to_update = normalize_bracket_mode(
                request.form.get("bracket_mode")
            )
        except FactoryRequirementError as exc:
            flash(str(exc), "error")
            return redirect(
                url_for(
                    "batch_edit_form",
                    project_id=project_id,
                    door_ids=",".join(str(door_id) for door_id in door_ids),
                )
            )
        if len(
            {
                door.get("installation_bracket_mode") or "profile"
                for door in selected_doors
            }
        ) > 1:
            mixed_fields.append("__bracket_mode__")

    if (
        not columns_to_update
        and hardware_to_update is None
        and bracket_mode_to_update is None
    ):
        flash("هیچ فیلدی برای به‌روزرسانی انتخاب نشده است.", "warning")
        return redirect(
            url_for(
                "batch_edit_form",
                project_id=project_id,
                door_ids=",".join(str(door_id) for door_id in door_ids),
            )
        )

    if mixed_fields and request.form.get("acknowledge_mixed") != "1":
        flash("برای یکسان‌کردن مقادیر متفاوت، تأیید صریح شما لازم است.", "error")
        return redirect(
            url_for(
                "batch_edit_form",
                project_id=project_id,
                door_ids=",".join(str(door_id) for door_id in door_ids),
            )
        )

    # 🔄 بکاپ خودکار قبل از ویرایش گروهی
    print(f"ایجاد بکاپ خودکار قبل از ویرایش گروهی پروژه {project_id}...")
    backup_success, backup_result = backup_manager.create_backup(
        reason=f"before_batch_edit",
        user="system",
        metadata={"project_id": project_id, "action": "batch_edit", "door_count": len(door_ids)}
    )
    if backup_success:
        print(f"✓ بکاپ قبل از ویرایش گروهی ایجاد شد: {backup_result}")
    else:
        flash(f"ویرایش گروهی متوقف شد؛ ایجاد بکاپ ایمنی ناموفق بود: {backup_result}", "error")
        return redirect(url_for("project_treeview", project_id=project_id))

    # اعمال تغییرات روی درب‌های انتخاب شده
    successful_updates, failed_updates, success_messages, error_messages = batch_update_doors_db(
        door_ids, {}, columns_to_update, project_id=project_id,
        hardware_to_update=hardware_to_update,
        bracket_mode_to_update=bracket_mode_to_update,
    )
    
    # به‌روزرسانی ستون‌های قابل مشاهده بر اساس داده‌های جدید
    if successful_updates > 0:
        refresh_project_visible_columns(project_id)

    # نمایش پیام‌های مناسب
    if successful_updates > 0:
        success_summary = f"{successful_updates} درب با موفقیت به‌روزرسانی شد."
        flash(success_summary, "success")
    
    if failed_updates > 0:
        error_summary = f"{failed_updates} درب با خطا مواجه شد."
        flash(error_summary, "error")
    
    if successful_updates == 0 and failed_updates == 0:
        flash("هیچ به‌روزرسانی انجام نشد.", "warning")

    # به‌روزرسانی ستون‌های قابل مشاهده بر اساس داده‌های جدید
    # این فراخوانی باید انجام شود تا اگر ستونی خالی شده، از لیست نمایش حذف گردد.
    refresh_project_visible_columns(project_id)

    # افزودن پارامتر زمانی برای جلوگیری از کش شدن صفحه
    timestamp = int(time.time())
    return redirect(
        url_for(
            "project_treeview",
            project_id=project_id,
            t=timestamp,
            force_refresh=1,
            updated_doors=",".join(str(door_id) for door_id in door_ids),
        )
    )


@app.route("/project/<int:project_id>/toggle_column_display", methods=["POST"])
def toggle_column_display(project_id):
    """تغییر وضعیت نمایش یک ستون"""
    column_key = request.form.get("column_key")
    is_visible = request.form.get("is_visible", "0") == "1"  # تبدیل به بولین
    
    if not column_key:
        return jsonify({"success": False, "error": "کلید ستون ارسال نشده است"})
    
    try:
        session_key = f"visible_columns_{project_id}"
        visible_columns = session.get(session_key, [])
        
        # اگر ستون باید نمایش داده شود و در لیست نیست، اضافه می‌کنیم
        if is_visible and column_key not in visible_columns:
            visible_columns.append(column_key)
            session[session_key] = visible_columns
            print(f"DEBUG: ستون {column_key} به لیست ستون‌های نمایشی پروژه {project_id} اضافه شد")
            return jsonify({"success": True})
        
        # اگر ستون نباید نمایش داده شود و در لیست هست، حذف می‌کنیم
        elif not is_visible and column_key in visible_columns:
            # قبل از حذف بررسی کنیم که آیا ستون حاوی داده است یا خیر
            # اگر ستون دارای داده باشد، اجازه مخفی کردن نمی‌دهیم
            if column_key in ["width", "height", "quantity", "direction"]:
                return jsonify({
                    "success": False, 
                    "error": f"ستون '{column_key}' یک ستون پایه است و نمی‌تواند مخفی شود"
                })
                
            # بررسی تعداد داده‌های ستون با استفاده از اندپوینت check_column_can_hide
            column_check = check_column_can_hide_internal(project_id, column_key)
            if not column_check.get("can_hide", True):
                return jsonify({
                    "success": False, 
                    "error": column_check.get("reason", "این ستون دارای داده است و نمی‌تواند مخفی شود")
                })
            
            # اگر به اینجا رسیدیم، ستون می‌تواند مخفی شود
            visible_columns.remove(column_key)
            session[session_key] = visible_columns
            print(f"DEBUG: ستون {column_key} از لیست ستون‌های نمایشی پروژه {project_id} حذف شد")
            return jsonify({"success": True})
        
        # در غیر این صورت، نیازی به تغییر نیست
        return jsonify({"success": True, "info": "وضعیت ستون تغییری نکرد"})
        
    except Exception as e:
        print(f"ERROR: خطا در تغییر وضعیت نمایش ستون {column_key}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})





@app.route("/project/<int:project_id>/check_column_can_hide", methods=["POST"])
def check_column_can_hide(project_id):
    """بررسی می‌کند که آیا ستون مورد نظر می‌تواند مخفی شود یا خیر"""
    column_key = request.form.get("column_key")
    if not column_key:
        return jsonify({"can_hide": False, "reason": "کلید ستون مشخص نشده است."})
    
    return jsonify(check_column_can_hide_internal(project_id, column_key))


# --- مسیرهای مربوط به سیستم انبار ---

@app.route("/inventory")
def inventory_route():
    """صفحه اصلی مدیریت انبار"""
    try:
        # دریافت آمار واقعی از دیتابیس
        stats = get_inventory_stats()
        
        # دریافت لیست پروفیل‌ها برای نمایش در داشبورد
        profiles = get_all_profile_types()
        
        return render_template("inventory_dashboard.html", stats=stats, profiles=profiles)
    except Exception as e:
        print(f"!!!!!! Unexpected error in inventory_route: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش صفحه مدیریت انبار رخ داد.", "error")
        return redirect(url_for("index"))


@app.route("/inventory/profile_types")
def profile_types_route():
    """صفحه مدیریت انواع پروفیل"""
    try:
        # دریافت لیست انواع پروفیل از دیتابیس
        profile_types = get_all_profile_types(include_inactive=True)
        
        return render_template("profile_types.html", profile_types=profile_types)
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت profile_types_route: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش صفحه انواع پروفیل رخ داد.", "error")
        return redirect(url_for("inventory_route"))

@app.route("/inventory/profile_types/add", methods=["GET", "POST"])
def add_profile_type_route():
    """افزودن نوع پروفیل جدید"""
    try:
        if request.method == "POST":
            name = request.form.get("name")
            color = request.form.get("color_hex") or request.form.get("color")
            default_length = float(request.form.get("default_length") or 600)
            weight_per_meter = float(request.form.get("weight_per_meter") or 1.9)
            min_waste = float(request.form.get("min_waste") or 20)
            description = request.form.get("description")
            
            if not name:
                flash("نام پروفیل الزامی است.", "error")
                return render_template("add_profile_type.html")
            
            success, result = add_profile_type(name, description, default_length, weight_per_meter, color, min_waste)
            
            if success:
                flash("نوع پروفیل با موفقیت اضافه شد.", "success")
                return redirect(url_for("profile_types_route"))
            else:
                # result already contains a user-friendly Persian message
                flash(result, "error")
        
        return render_template("add_profile_type.html")
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت add_profile_type_route: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش صفحه افزودن پروفیل رخ داد.", "error")
        return redirect(url_for("profile_types_route"))

@app.route("/inventory/profile_types/edit/<int:profile_id>", methods=["GET", "POST"])
def edit_profile_type_route(profile_id):
    """ویرایش نوع پروفیل"""
    try:
        profile = get_profile_details(profile_id)
        if not profile:
            flash("پروفیل مورد نظر یافت نشد.", "error")
            return redirect(url_for("profile_types_route"))

        if request.method == "POST":
            name = request.form.get("name")
            color = request.form.get("color_hex") or request.form.get("color")
            default_length = float(request.form.get("default_length") or 600)
            weight_per_meter = float(request.form.get("weight_per_meter") or 1.9)
            min_waste = float(request.form.get("min_waste") or 20)
            description = request.form.get("description")
            
            if not name:
                flash("نام پروفیل الزامی است.", "error")
                return render_template("edit_profile_type.html", profile=profile)
            
            # در اینجا باید تابع ویرایش را فراخوانی کنیم
            success = update_profile_type(profile_id, name, description, default_length, weight_per_meter, color, min_waste)
            
            if success:
                flash("پروفیل با موفقیت ویرایش شد.", "success")
                return redirect(url_for("profile_types_route"))
            else:
                flash("خطا در ویرایش پروفیل.", "error")
        
        return render_template("edit_profile_type.html", profile=profile)
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت edit_profile_type_route: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش صفحه ویرایش پروفیل رخ داد.", "error")
        return redirect(url_for("profile_types_route"))

@app.route("/inventory/profile_types/delete/<int:profile_id>", methods=["POST"])
@manager_or_admin_required
def delete_profile_type_route(profile_id):
    """حذف نوع پروفیل"""
    try:
        result = delete_profile_type(
            profile_id,
            actor_user_id=current_user.id,
            reason=request.form.get("reason", ""),
        )
        flash(
            result.get("message", "عملیات انجام نشد."),
            "success" if result.get("status") in ("deleted", "archived") else "error",
        )
        return redirect(url_for("profile_types_route"))
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت delete_profile_type_route: {e}")
        traceback.print_exc()
        flash("خطایی در انجام عملیات حذف رخ داد.", "error")
        return redirect(url_for("profile_types_route"))


@app.route("/inventory/settings", methods=["GET", "POST"])
def inventory_settings_route():
    """صفحه تنظیمات انبار"""
    try:
        if request.method == "POST":
            # دریافت تنظیمات از فرم و ذخیره در دیتابیس
            new_settings = {
                "default_wastage": request.form.get("default_wastage", 20),
                "min_remaining_length": request.form.get("min_remaining_length", 20),
                "use_inventory_for_cutting": request.form.get("use_inventory_for_cutting") == "on",
                "prefer_inventory_pieces": request.form.get("prefer_inventory_pieces") == "on",
                "inventory_optimization_strategy": request.form.get("inventory_optimization_strategy", "minimize_waste"),
                "show_inventory_warnings": request.form.get("show_inventory_warnings") == "on",
                "low_inventory_threshold": request.form.get("low_inventory_threshold", 5)
            }
            
            if update_inventory_settings(new_settings):
                flash("تنظیمات انبار با موفقیت ذخیره شد.", "success")
            else:
                flash("خطا در ذخیره تنظیمات.", "error")
            
            return redirect(url_for("inventory_settings_route"))

        # دریافت تنظیمات فعلی از دیتابیس
        settings = get_inventory_settings()
        
        # اگر تنظیمات خالی بود (هنوز ست نشده)، مقادیر پیش‌فرض را نمایش بده
        if not settings:
            settings = {
                "waste_threshold": 70,
                "use_inventory": True,
                "prefer_pieces": True
            }
        
        return render_template("inventory_settings.html", settings=settings)
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت inventory_settings_route: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش صفحه تنظیمات انبار رخ داد.", "error")
        return redirect(url_for("inventory_route"))


@app.route("/inventory/logs")
@app.route("/inventory/logs/<int:profile_id>")
def inventory_logs_route(profile_id=None):
    """صفحه تاریخچه تغییرات انبار"""
    try:
        logs = get_inventory_logs(limit=100, profile_id=profile_id)
        profiles = get_all_profile_types()
        latest_operation = get_latest_reversible_inventory_operation()
        
        # افزودن ترجمه نوع تغییر
        change_type_map = {
            "add_stock": "افزایش موجودی",
            "remove_stock": "کاهش موجودی",
            "add_piece": "افزودن تکه",
            "remove_piece": "حذف تکه",
            "undo_stock": "بازگردانی موجودی",
            "undo_add_piece": "بازگردانی افزودن تکه",
            "undo_remove_piece": "بازگردانی حذف تکه",
        }
        
        logs_with_translation = []
        for log in logs:
            log_dict = dict(log)
            log_dict["change_type_fa"] = change_type_map.get(log_dict["change_type"], log_dict["change_type"])
            logs_with_translation.append(log_dict)
            
        return render_template(
            "inventory_logs.html",
            logs=logs_with_translation,
            profiles=profiles,
            profile_id=profile_id,
            latest_operation=latest_operation,
        )
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت inventory_logs_route: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش تاریخچه رخ داد.", "error")
        return redirect(url_for("inventory_route"))

@app.route("/inventory/details/<int:profile_id>")
def inventory_details_route(profile_id):
    """صفحه جزئیات موجودی یک پروفیل"""
    try:
        profile = get_profile_details(profile_id)
        if not profile:
            flash("پروفیل مورد نظر یافت نشد.", "error")
            return redirect(url_for("inventory_route"))
            
        details = get_profile_stock_details(profile_id)
        
        # سازماندهی داده‌ها برای قالب
        template_details = {
            "profile": profile,
            "full_items": details["complete_pieces"],
            "pieces": details["pieces"],
            "logs": details["logs"]
        }
        
        # افزودن ترجمه نوع تغییر به لاگ‌ها
        change_type_map = {
            "add_stock": "افزایش موجودی",
            "remove_stock": "کاهش موجودی",
            "add_piece": "افزودن تکه",
            "remove_piece": "حذف تکه",
            "undo_stock": "بازگردانی موجودی",
            "undo_add_piece": "بازگردانی افزودن تکه",
            "undo_remove_piece": "بازگردانی حذف تکه",
        }
        
        logs_with_translation = []
        for log in details["logs"]:
            log_dict = dict(log)
            log_dict["change_type_fa"] = change_type_map.get(log_dict["change_type"], log_dict["change_type"])
            logs_with_translation.append(log_dict)
            
        return render_template("profile_inventory_details.html", details=template_details, logs=logs_with_translation)
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت inventory_details_route: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش صفحه جزئیات انبار رخ داد.", "error")
        return redirect(url_for("profile_types_route"))

@app.route("/inventory/items/add/<int:profile_id>", methods=["POST"])
@staff_or_admin_required
def add_inventory_items_route(profile_id):
    """افزودن شاخه کامل به انبار"""
    try:
        quantity = int(request.form.get("quantity", 0))
        description = request.form.get("description", "")
        
        if quantity <= 0:
            flash("تعداد باید بزرگتر از صفر باشد.", "error")
        else:
            if add_inventory_stock(
                profile_id, quantity, description, actor_user_id=current_user.id
            ):
                flash("موجودی با موفقیت اضافه شد.", "success")
            else:
                flash("خطا در افزودن موجودی.", "error")
                
        return redirect(url_for("inventory_details_route", profile_id=profile_id))
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت add_inventory_items_route: {e}")
        traceback.print_exc()
        flash("خطایی در انجام عملیات رخ داد.", "error")
        return redirect(url_for("inventory_details_route", profile_id=profile_id))

@app.route("/inventory/items/remove/<int:profile_id>", methods=["POST"])
@staff_or_admin_required
def remove_inventory_items_route(profile_id):
    """کاهش شاخه کامل از انبار"""
    try:
        quantity = int(request.form.get("quantity", 0))
        description = request.form.get("description", "")
        
        if quantity <= 0:
            flash("تعداد باید بزرگتر از صفر باشد.", "error")
        else:
            # 🔄 بکاپ خودکار قبل از کسر موجودی انبار
            print(f"ایجاد بکاپ خودکار قبل از کسر موجودی انبار (profile_id={profile_id})...")
            backup_success, backup_result = backup_manager.create_backup(
                reason=f"before_inventory_deduction",
                user="system",
                metadata={"profile_id": profile_id, "action": "remove_stock", "quantity": quantity}
            )
            if backup_success:
                print(f"✓ بکاپ قبل از کسر موجودی ایجاد شد: {backup_result}")
            else:
                flash(f"کسر موجودی متوقف شد؛ ایجاد بکاپ ایمنی ناموفق بود: {backup_result}", "error")
                return redirect(url_for("inventory_details_route", profile_id=profile_id))
            
            success, msg = remove_inventory_stock(
                profile_id, quantity, description, actor_user_id=current_user.id
            )
            if success:
                flash("موجودی با موفقیت کسر شد.", "success")
            else:
                flash(f"خطا در کسر موجودی: {msg}", "error")
                
        return redirect(url_for("inventory_details_route", profile_id=profile_id))
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت remove_inventory_items_route: {e}")
        traceback.print_exc()
        flash("خطایی در انجام عملیات رخ داد.", "error")
        return redirect(url_for("inventory_details_route", profile_id=profile_id))

@app.route("/inventory/pieces/add/<int:profile_id>", methods=["POST"])
@staff_or_admin_required
def add_inventory_piece_route(profile_id):
    """افزودن تکه شاخه به انبار"""
    try:
        length = float(request.form.get("length", 0))
        description = request.form.get("description", "")
        
        if length <= 0:
            flash("طول باید بزرگتر از صفر باشد.", "error")
        else:
            if add_inventory_piece(
                profile_id, length, description, actor_user_id=current_user.id
            ):
                flash("تکه شاخه با موفقیت اضافه شد.", "success")
            else:
                flash("خطا در افزودن تکه شاخه.", "error")
                
        return redirect(url_for("inventory_details_route", profile_id=profile_id))
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت add_inventory_piece_route: {e}")
        traceback.print_exc()
        flash("خطایی در انجام عملیات رخ داد.", "error")
        return redirect(url_for("inventory_details_route", profile_id=profile_id))

@app.route("/inventory/pieces/remove/<int:piece_id>", methods=["POST"])
@staff_or_admin_required
def remove_inventory_piece_route(piece_id):
    """حذف تکه شاخه از انبار"""
    try:
        profile_id = request.form.get("profile_id")
        
        success, msg = remove_inventory_piece(
            piece_id,
            description="حذف دستی توسط کاربر",
            actor_user_id=current_user.id,
        )
        if success:
            flash("تکه شاخه با موفقیت حذف شد.", "success")
        else:
            flash(f"خطا در حذف تکه شاخه: {msg}", "error")
            
        if profile_id:
            return redirect(url_for("inventory_details_route", profile_id=profile_id))
        return redirect(url_for("inventory_route"))
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت remove_inventory_piece_route: {e}")
        traceback.print_exc()
        flash("خطایی در انجام عملیات رخ داد.", "error")
        return redirect(url_for("inventory_route"))


@app.route("/project/<int:project_id>/export_pdf", methods=["GET"])
def export_table_to_pdf_html(project_id):
    """صفحه خروجی PDF از جدول پروژه با استفاده از HTML"""
    try:
        # دریافت اطلاعات پروژه و درب‌ها
        project = get_project_details_db(project_id)
        if not project:
            flash(f"پروژه با شناسه {project_id} یافت نشد.", "error")
            return redirect(url_for("index"))
        
        doors = get_doors_for_project_db(project_id)
        
        # دریافت ستون‌های قابل نمایش از session
        session_key = f"visible_columns_{project_id}"
        visible_columns = session.get(session_key, [])
        
        # ایجاد یک نام فایل موقت برای خروجی HTML
        current_date = jdatetime.datetime.now().strftime("%Y%m%d")
        pdf_filename = f"project_{project_id}_{current_date}.pdf"
        
        # رندر قالب جدول برای PDF
        return render_template(
            "pdf_table_template.html",
            project=project,
            doors=doors,
            visible_columns=visible_columns,
            pdf_filename=pdf_filename
        )
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت export_table_to_pdf_html: {e}")
        traceback.print_exc()
        flash("خطایی در ایجاد خروجی PDF رخ داد.", "error")
        return redirect(url_for("project_treeview", project_id=project_id))


@app.route("/project/<int:project_id>/settings_columns", methods=["GET"])
def settings_columns(project_id):
    """Manage the reusable field library for one project's selected fields."""
    project = get_project_details_db(project_id)
    if not project:
        flash("سفارش مورد نظر پیدا نشد.", "error")
        return redirect(url_for("index"))

    columns = get_project_custom_columns(project_id)
    current_columns = [column for column in columns if column["is_selected"]]
    library_columns = [column for column in columns if not column["is_selected"]]
    total_door_quantity = sum(
        int(door.get("quantity") or 0) for door in get_doors_for_project_db(project_id)
    )
    initialize_visible_columns(project_id)
    return render_template(
        "project_column_settings.html",
        project=project,
        current_columns=current_columns,
        library_columns=library_columns,
        total_door_quantity=total_door_quantity,
    )


def _settings_column(project_id, column_id):
    return next(
        (
            column
            for column in get_project_custom_columns(project_id)
            if column["id"] == column_id
        ),
        None,
    )


@app.route(
    "/project/<int:project_id>/settings_columns/toggle", methods=["POST"]
)
@csrf_protected
@staff_or_admin_required
def project_column_toggle(project_id):
    data = request.get_json(silent=True) or request.form
    try:
        column_id = int(data.get("column_id", ""))
    except (TypeError, ValueError):
        return jsonify({"success": False, "error": "فیلد انتخاب‌شده معتبر نیست."}), 400

    raw_visibility = data.get("is_visible", False)
    is_visible = raw_visibility is True or str(raw_visibility).lower() in {"1", "true", "on"}
    column = _settings_column(project_id, column_id)
    if not column:
        return jsonify({"success": False, "error": "فیلد مورد نظر پیدا نشد."}), 404
    if not set_project_column_visibility(project_id, column_id, is_visible):
        return jsonify({"success": False, "error": "ذخیره تغییر انجام نشد."}), 500

    initialize_visible_columns(project_id)
    return jsonify(
        {
            "success": True,
            "message": (
                f"«{column['display']}» در این سفارش نمایش داده می‌شود."
                if is_visible
                else f"«{column['display']}» پنهان شد؛ اطلاعات قبلی آن محفوظ است."
            ),
        }
    )


@app.route(
    "/project/<int:project_id>/settings_columns/add-existing", methods=["POST"]
)
@csrf_protected
@staff_or_admin_required
def project_column_add_existing(project_id):
    data = request.get_json(silent=True) or request.form
    try:
        column_id = int(data.get("column_id", ""))
    except (TypeError, ValueError):
        return jsonify({"success": False, "error": "فیلد انتخاب‌شده معتبر نیست."}), 400
    column = _settings_column(project_id, column_id)
    if not column:
        return jsonify({"success": False, "error": "فیلد مورد نظر پیدا نشد."}), 404
    if not set_project_column_visibility(project_id, column_id, True):
        return jsonify({"success": False, "error": "افزودن فیلد انجام نشد."}), 500
    initialize_visible_columns(project_id)
    return jsonify(
        {"success": True, "message": f"«{column['display']}» به همین سفارش اضافه شد."}
    )


@app.route(
    "/project/<int:project_id>/settings_columns/remove", methods=["POST"]
)
@csrf_protected
@staff_or_admin_required
def project_column_remove(project_id):
    data = request.get_json(silent=True) or request.form
    try:
        column_id = int(data.get("column_id", ""))
    except (TypeError, ValueError):
        return jsonify({"success": False, "error": "فیلد انتخاب‌شده معتبر نیست."}), 400
    column = _settings_column(project_id, column_id)
    if not column:
        return jsonify({"success": False, "error": "فیلد مورد نظر پیدا نشد."}), 404
    if not remove_project_column(project_id, column_id):
        return jsonify({"success": False, "error": "برداشتن فیلد انجام نشد."}), 500
    initialize_visible_columns(project_id)
    return jsonify(
        {
            "success": True,
            "message": f"«{column['display']}» فقط از این سفارش برداشته شد و اطلاعات قبلی آن محفوظ ماند.",
        }
    )


@app.route(
    "/project/<int:project_id>/settings_columns/create", methods=["POST"]
)
@csrf_protected
@staff_or_admin_required
def project_column_create(project_id):
    data = request.get_json(silent=True) or request.form
    display_name = " ".join(str(data.get("display_name", "")).split())
    column_type = str(data.get("column_type", "text"))
    if not display_name:
        return jsonify({"success": False, "error": "نام فیلد را وارد کنید."}), 400
    if column_type not in {"text", "dropdown"}:
        return jsonify({"success": False, "error": "روش ورود اطلاعات معتبر نیست."}), 400

    existing = next(
        (
            column
            for column in get_project_custom_columns(project_id)
            if column["display"].strip().casefold() == display_name.casefold()
        ),
        None,
    )
    if existing:
        if existing["is_selected"]:
            return jsonify(
                {"success": False, "error": "فیلدی با این نام از قبل در سفارش وجود دارد."}
            ), 409
        set_project_column_visibility(project_id, existing["id"], True)
        initialize_visible_columns(project_id)
        return jsonify(
            {
                "success": True,
                "message": f"«{existing['display']}» از فهرست آماده به سفارش اضافه شد.",
            }
        )

    column_id = add_custom_column(display_name=display_name, column_type=column_type)
    if not column_id or not set_project_column_visibility(project_id, column_id, True):
        return jsonify({"success": False, "error": "ساخت فیلد انجام نشد."}), 500
    initialize_visible_columns(project_id)
    return jsonify(
        {"success": True, "message": f"«{display_name}» ساخته و به همین سفارش اضافه شد."}
    )


@app.route(
    "/project/<int:project_id>/settings_columns/<int:column_id>/options",
    methods=["GET"],
)
def project_column_options(project_id, column_id):
    column = _settings_column(project_id, column_id)
    if not column or column["type"] != "dropdown":
        return jsonify({"success": False, "error": "این فیلد فهرست انتخابی ندارد."}), 404
    return jsonify(
        {"success": True, "column": column, "options": get_custom_column_options(column_id)}
    )


@app.route(
    "/project/<int:project_id>/settings_columns/<int:column_id>/options/add",
    methods=["POST"],
)
@csrf_protected
@staff_or_admin_required
def project_column_option_add(project_id, column_id):
    column = _settings_column(project_id, column_id)
    if not column or column["type"] != "dropdown":
        return jsonify({"success": False, "error": "این فیلد فهرست انتخابی ندارد."}), 404
    if column["key"] == "noe_profile":
        return jsonify({"success": False, "error": "نوع پروفیل از بخش انبار مدیریت می‌شود."}), 400
    data = request.get_json(silent=True) or request.form
    value = " ".join(str(data.get("option_value", "")).split())
    if not value:
        return jsonify({"success": False, "error": "نام گزینه را وارد کنید."}), 400
    if not add_option_to_column(column_id, value):
        return jsonify({"success": False, "error": "افزودن گزینه انجام نشد."}), 500
    return jsonify({"success": True, "message": "گزینه اضافه شد."})


@app.route(
    "/project/<int:project_id>/settings_columns/<int:column_id>/options/<int:option_id>/edit",
    methods=["POST"],
)
@csrf_protected
@staff_or_admin_required
def project_column_option_edit(project_id, column_id, option_id):
    column = _settings_column(project_id, column_id)
    if not column or get_column_id_from_option_db(option_id) != column_id:
        return jsonify({"success": False, "error": "گزینه مورد نظر پیدا نشد."}), 404
    if column["key"] == "noe_profile":
        return jsonify({"success": False, "error": "نوع پروفیل از بخش انبار مدیریت می‌شود."}), 400
    data = request.get_json(silent=True) or request.form
    value = " ".join(str(data.get("new_value", "")).split())
    if not value:
        return jsonify({"success": False, "error": "نام گزینه را وارد کنید."}), 400
    if not update_custom_column_option(option_id, value):
        return jsonify({"success": False, "error": "ویرایش گزینه انجام نشد."}), 500
    return jsonify({"success": True, "message": "گزینه ویرایش شد."})


@app.route(
    "/project/<int:project_id>/settings_columns/<int:column_id>/options/<int:option_id>/delete",
    methods=["POST"],
)
@csrf_protected
@staff_or_admin_required
def project_column_option_delete(project_id, column_id, option_id):
    column = _settings_column(project_id, column_id)
    if not column or get_column_id_from_option_db(option_id) != column_id:
        return jsonify({"success": False, "error": "گزینه مورد نظر پیدا نشد."}), 404
    if column["key"] == "noe_profile":
        return jsonify({"success": False, "error": "نوع پروفیل از بخش انبار مدیریت می‌شود."}), 400
    if not delete_column_option(option_id):
        return jsonify({"success": False, "error": "حذف گزینه انجام نشد."}), 500
    return jsonify({"success": True, "message": "گزینه حذف شد."})

@app.route("/project/<int:project_id>/add_column", methods=["POST"])
def add_column_route(project_id):
    """افزودن ستون جدید سفارشی (برای سازگاری با قبل)"""
    # دریافت اطلاعات فرم
    display_name = request.form.get("display_name")
    column_key = request.form.get("column_key")
    column_type = request.form.get("column_type")
    
    # ذخیره اطلاعات در session برای انتقال به روت جدید
    session['temp_column_data'] = {
        'display_name': display_name,
        'column_key': column_key,
        'column_type': column_type,
        'action': 'add_column'
    }
    
    # ریدایرکت به روت جدید
    return redirect(url_for("manage_custom_columns", project_id=project_id))


@app.route("/project/<int:project_id>/update_column_display", methods=["POST"])
def update_column_display(project_id):
    """به‌روزرسانی تنظیمات نمایش ستون‌ها (برای سازگاری با قبل)"""
    # ریدایرکت به روت جدید
    return redirect(url_for("manage_custom_columns", project_id=project_id))


@app.route("/column/<int:column_id>/delete/<int:project_id>", methods=["GET"])
def delete_column_route(column_id, project_id):
    """حذف ستون سفارشی (برای سازگاری با قبل)"""
    # ذخیره اطلاعات در session برای انتقال به روت جدید
    session['temp_column_data'] = {
        'column_id': column_id,
        'action': 'delete_column'
    }
    
    # ریدایرکت به روت جدید
    return redirect(url_for("manage_custom_columns", project_id=project_id))


@app.route('/save_batch_edit_checkbox_state', methods=['POST'])
def save_batch_edit_checkbox_state():
    data = request.get_json()
    column = data.get('column')
    checked = data.get('checked')
    
    if not column:
        return jsonify({'success': False, 'error': 'Column name is required'})
    
    # Initialize the session key if it doesn't exist
    if 'batch_edit_checked_columns' not in session:
        session['batch_edit_checked_columns'] = {}
    
    # Update the session with the new checkbox state
    session['batch_edit_checked_columns'][column] = checked
    session.modified = True
    
    return jsonify({'success': True})


@app.route("/project/<int:project_id>/save_batch_edit_checkbox_state", methods=["POST"])
def save_batch_edit_checkbox_state_project(project_id):
    """ذخیره وضعیت چک‌باکس‌های ویرایش گروهی"""
    column_key = request.form.get("column_key")
    is_checked = request.form.get("is_checked", "0") == "1"
    door_id = request.form.get("door_id")  # اضافه کردن شناسه درب
    
    if not column_key:
        return jsonify({"success": False, "error": "کلید ستون ارسال نشده است"})
    
    try:
        # ۱. ذخیره وضعیت چک‌باکس‌های ویرایش گروهی برای هر درب
        batch_edit_checked_key = f"batch_edit_checked_{project_id}_{door_id}"
        checked_columns = session.get(batch_edit_checked_key, [])
        
        # به‌روزرسانی لیست ستون‌های تیک خورده برای این درب
        if is_checked and column_key not in checked_columns:
            checked_columns.append(column_key)
        elif not is_checked and column_key in checked_columns:
            checked_columns.remove(column_key)
        
        session[batch_edit_checked_key] = checked_columns
        session.modified = True
        
        # ۲. به‌روزرسانی وضعیت نمایش ستون‌ها
        session_key = f"visible_columns_{project_id}"
        visible_columns = session.get(session_key, [])
        
        # ستون‌های پایه که همیشه باید نمایش داده شوند
        basic_columns = ["location", "width", "height", "quantity", "direction"]
        
        # اگر ستون پایه نیست (جزو ستون‌های سفارشی است)
        if column_key not in basic_columns:
            # وقتی ستون تیک دارد، باید از لیست نمایش حذف شود
            if is_checked:
                if column_key in visible_columns:
                    visible_columns.remove(column_key)
                    print(f"DEBUG: ستون {column_key} تیک خورده و از لیست نمایشی حذف شد")
            # وقتی ستون تیک ندارد، باید به لیست نمایش اضافه شود
            else:
                if column_key not in visible_columns:
                    visible_columns.append(column_key)
                    print(f"DEBUG: ستون {column_key} تیک ندارد و به لیست نمایشی اضافه شد")
            
            session[session_key] = visible_columns
            session.modified = True
        
        print(f"DEBUG: درب {door_id} - ستون '{column_key}' به وضعیت تیک {is_checked} تغییر یافت.")
        print(f"DEBUG: ستون‌های تیک خورده برای درب {door_id}: {checked_columns}")
        print(f"DEBUG: ستون‌های نمایشی: {visible_columns}")
        
        return jsonify({
            "success": True,
            "checked_columns": checked_columns,
            "visible_columns": visible_columns
        })
        
    except Exception as e:
        print(f"ERROR: خطا در ذخیره وضعیت چک‌باکس {column_key} برای درب {door_id}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/project/<int:project_id>/get_visible_columns", methods=["GET"])
def get_visible_columns(project_id):
    """دریافت لیست ستون‌های نمایشی"""
    try:
        session_key = f"visible_columns_{project_id}"
        visible_columns = session.get(session_key, [])
        
        # اگر لیست ستون‌های نمایشی خالی است، آن را با مقادیر پیش‌فرض پر کنیم
        if not visible_columns:
            initialize_visible_columns(project_id)
            visible_columns = session.get(session_key, [])
        
        # اضافه کردن ستون‌های پایه که همیشه باید نمایش داده شوند
        basic_columns = ["location", "width", "height", "quantity", "direction"]
        for col in basic_columns:
            if col not in visible_columns:
                visible_columns.append(col)
        
        print(f"DEBUG: ستون‌های نمایشی پروژه {project_id}: {visible_columns}")
        
        return jsonify({"success": True, "visible_columns": visible_columns})
    except Exception as e:
        print(f"ERROR: خطا در دریافت لیست ستون‌های نمایشی: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@app.route("/project/<int:project_id>/batch_remove_column_value", methods=["POST"])
@csrf_protected
@staff_or_admin_required
def batch_remove_column_value_route(project_id):
    """
    این روت درخواست AJAX برای حذف مقادیر یک ستون خاص 
    برای درب‌های انتخاب شده را پردازش می‌کند.
    """
    print(f"DEBUG: ورود به روت batch_remove_column_value_route برای پروژه ID: {project_id}")

    if not request.is_json:
        print("ERROR: درخواست باید JSON باشد.")
        return jsonify({"success": False, "error": "درخواست باید با فرمت JSON باشد"}), 400

    data = request.get_json()
    door_ids_str_list = data.get('door_ids')  # انتظار داریم لیستی از IDها به صورت رشته باشد
    column_key_to_remove = data.get('column_key_to_remove')

    print(f"DEBUG: داده‌های دریافتی: door_ids={door_ids_str_list}, column_key={column_key_to_remove}")

    if not door_ids_str_list or not isinstance(door_ids_str_list, list) or not column_key_to_remove:
        error_msg = "ID درب‌ها (به صورت لیست) و کلید ستون مورد نیاز است."
        print(f"ERROR: {error_msg}")
        return jsonify({"success": False, "error": error_msg}), 400

    try:
        # تبدیل ID درب‌ها از رشته به عدد صحیح
        door_ids = [int(d_id) for d_id in door_ids_str_list]
    except ValueError:
        error_msg = "فرمت ID درب‌ها نامعتبر است. باید لیستی از اعداد صحیح باشد."
        print(f"ERROR: {error_msg}")
        return jsonify({"success": False, "error": error_msg}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        allowed_column_keys = {
            column["key"]
            for column in get_project_visible_custom_columns(project_id)
        }
        if column_key_to_remove not in allowed_column_keys:
            return jsonify({
                "success": False,
                "error": "این فیلد برای سفارش مورد نظر قابل ویرایش نیست.",
            }), 400

        placeholders = ",".join("?" for _ in door_ids)
        cursor.execute(
            f"SELECT id FROM doors WHERE project_id = ? AND id IN ({placeholders})",
            [project_id, *door_ids],
        )
        matching_door_ids = {int(row[0]) for row in cursor.fetchall()}
        if matching_door_ids != set(door_ids):
            return jsonify({
                "success": False,
                "error": "حداقل یکی از درب‌ها متعلق به این سفارش نیست.",
            }), 400

        # دریافت ID ستون از روی کلید (column_key)
        column_id = get_column_id_by_key(column_key_to_remove) 
        
        if not column_id:
            error_msg = f"ستون با کلید '{column_key_to_remove}' یافت نشد."
            print(f"ERROR: {error_msg}")
            return jsonify({"success": False, "error": error_msg}), 404

        print(f"DEBUG: ID ستون '{column_key_to_remove}' یافت شد: {column_id}")
        
        # تلاش برای دریافت نام نمایشی ستون
        display_name = None
        try:
            cursor.execute("SELECT display_name FROM custom_columns WHERE id = ?", (column_id,))
            result = cursor.fetchone()
            if result:
                display_name = result[0]  # نام نمایشی ستون
        except Exception as e:
            print(f"WARNING: خطا در دریافت نام نمایشی ستون: {e}")
        
        # اگر نام نمایشی دریافت نشد، از کلید ستون استفاده می‌کنیم
        column_identifier_for_message = display_name if display_name else column_key_to_remove

        deleted_count_total = 0
        for door_id in door_ids:
            # حذف مقدار از جدول door_custom_values
            print(f"DEBUG: تلاش برای حذف مقدار ستون {column_id} برای درب {door_id}")
            cursor.execute("""
                DELETE FROM door_custom_values 
                WHERE door_id = ? AND column_id = ?
            """, (door_id, column_id))
            
            if cursor.rowcount > 0:
                deleted_count_total += 1
                print(f"DEBUG: مقدار ستون {column_id} برای درب {door_id} حذف شد.")
            else:
                print(f"DEBUG: مقداری برای ستون {column_id} و درب {door_id} جهت حذف یافت نشد (ممکن است از قبل خالی بوده باشد).")
        
        conn.commit()
        print(f"DEBUG: عملیات حذف commit شد. تعداد کل رکوردهای حذف شده: {deleted_count_total}")
        
        # تهیه پیام مناسب بر اساس تعداد رکوردهای حذف شده
        if deleted_count_total == 0:
            message = f"مقداری برای ستون '{column_identifier_for_message}' جهت حذف یافت نشد (ممکن است از قبل خالی بوده یا تغییرات UI هنوز ذخیره نشده باشند)."
        else:
            message = f"{deleted_count_total} مقدار از ستون '{column_identifier_for_message}' با موفقیت از دیتابیس حذف شد."

        # اگر مقداری حذف شده یا حتی اگر حذف نشده (چون ممکن است آخرین مقدار بوده)، visible_columns را به‌روز کن
        # شرط if deleted_count_total > 0: می‌تواند برای بهینه‌سازی باشد، اما برای اطمینان بیشتر، همیشه فراخوانی می‌کنیم.
        refresh_project_visible_columns(project_id)

        return jsonify({"success": True, "message": message})

    except sqlite3.Error as e:
        if conn:
            conn.rollback()
        print(f"ERROR - sqlite3.Error در batch_remove_column_value_route: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": f"خطای دیتابیس: {str(e)}"}), 500
    except Exception as e:
        print(f"ERROR - Exception عمومی در batch_remove_column_value_route: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": f"خطای عمومی سرور: {str(e)}"}), 500
    finally:
        if conn:
            conn.close()

@app.route("/settings/custom_columns", methods=["GET", "POST"])
def manage_custom_columns():
    """صفحه مدیریت ستون‌های سفارشی"""
    try:
        # دریافت شناسه پروژه برای بازگشت (در صورت وجود)
        # ممکن است در query string یا form data باشد
        project_id = request.args.get("project_id") or request.form.get("project_id")
        
        # بررسی وجود اطلاعات در session (برای سازگاری با روت‌های قدیمی)
        temp_data = session.pop('temp_column_data', None)
        
        action = request.form.get("action") if request.method == "POST" else temp_data.get("action") if temp_data else None
        
        # افزودن ستون جدید
        if action == "add_column":
            if request.method == "POST":
                display_name = request.form.get("display_name")
                column_type = request.form.get("column_type")
            else:
                display_name = temp_data.get("display_name")
                column_type = temp_data.get("column_type")
            
            if not display_name or not column_type:
                flash("لطفاً نام نمایشی و نوع ستون را وارد کنید.", "error")
                return redirect(url_for("manage_custom_columns", project_id=project_id))
            
            if column_type not in ['text', 'dropdown']:
                flash("نوع ستون انتخاب شده نامعتبر است. لطفاً 'متنی' یا 'دراپ‌داون' را انتخاب کنید.", "error")
                return redirect(url_for("manage_custom_columns", project_id=project_id))
            
            # کلید داخلی پایدار به‌صورت خودکار در دیتابیس ساخته می‌شود.
            new_column_id = add_custom_column(
                display_name=display_name,
                column_type=column_type,
            )
            if new_column_id:
                flash(f"ستون '{display_name}' با موفقیت اضافه شد.", "success")
            else:
                flash("خطا در افزودن ستون جدید.", "error")
            return redirect(url_for("manage_custom_columns", project_id=project_id))
        
        # حذف ستون
        elif action == "delete_column":
            if request.method == "POST":
                column_id = request.form.get("column_id")
            else:
                column_id = temp_data.get("column_id")
            
            if column_id:
                column_id = int(column_id)
                conn = None
                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    
                    # حذف مقادیر مربوط به این ستون
                    cursor.execute("DELETE FROM door_custom_values WHERE column_id = ?", (column_id,))
                    
                    # حذف ستون
                    cursor.execute("DELETE FROM custom_columns WHERE id = ?", (column_id,))
                    
                    conn.commit()
                    flash("ستون با موفقیت حذف شد.", "success")
                except Exception as e:
                    print(f"خطا در حذف ستون: {e}")
                    flash("خطا در حذف ستون.", "error")
                finally:
                    if conn:
                        conn.close()
            return redirect(url_for("manage_custom_columns", project_id=project_id))
        
        # تغییر وضعیت فعال/غیرفعال ستون
        elif action == "toggle_status":
            column_id_str = request.form.get("column_id")
            if column_id_str:
                column_id = int(column_id_str)
                # اگر کلید 'is_active' در request.form وجود داشت و مقدارش '1' بود، یعنی چک‌باکس تیک خورده است.
                # در غیر این صورت (یعنی کلید 'is_active' اصلاً در فرم نبود چون تیک نخورده)، مقدار آن False خواهد بود.
                is_active_bool = request.form.get("is_active") == "1"
                
                success = update_custom_column_status(column_id, is_active_bool)
                if success:
                    status_text = "فعال" if is_active_bool else "غیرفعال"
                    flash(f"وضعیت ستون با موفقیت به {status_text} تغییر کرد.", "success")
                else:
                    flash(f"خطا در تغییر وضعیت ستون با شناسه {column_id}.", "error")
            else:
                flash("شناسه ستون برای تغییر وضعیت ارسال نشده است.", "error")
            return redirect(url_for("manage_custom_columns", project_id=project_id))
        
        # پردازش درخواست GET (نمایش صفحه)
        all_columns = get_all_custom_columns()
        
        column_type_display_map = {
            'text': 'متنی',
            'dropdown': 'دراپ‌داون'
        }
        processed_columns = []
        for col in all_columns:
            col_copy = col.copy() 
            col_copy['type_display'] = column_type_display_map.get(col_copy.get('type'), col_copy.get('type', 'نامشخص'))
            if col_copy.get('type') == 'dropdown':
                col_copy['options'] = get_custom_column_options(col_copy['id'])
            else:
                col_copy['options'] = []
            processed_columns.append(col_copy)
        
        return render_template(
            "column_settings.html",
            all_columns=processed_columns,
            project_id=project_id
        )
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت manage_custom_columns: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش صفحه تنظیمات ستون‌ها رخ داد.", "error")
        return redirect(url_for("index"))

@app.route("/project/<int:project_id>/settings_combos", methods=["GET"])
def settings_combos(project_id):
    """صفحه مدیریت گزینه‌های کمبوباکس (dropdown) برای پروژه"""
    try:
        # دریافت اطلاعات پروژه
        project_info = get_project_details_db(project_id)
        if not project_info:
            flash("پروژه مورد نظر یافت نشد.", "error")
            return redirect(url_for("index"))
        
        # دریافت تمام ستون‌های dropdown
        all_columns = get_all_custom_columns()
        dropdown_columns = [col for col in all_columns if col.get('type') == 'dropdown']
        
        return render_template(
            "settings_combos.html",
            project=project_info,
            columns=dropdown_columns
        )
    except Exception as e:
        print(f"!!!!!! خطای غیرمنتظره در روت settings_combos: {e}")
        traceback.print_exc()
        flash("خطایی در نمایش صفحه مدیریت گزینه‌ها رخ داد.", "error")
        return redirect(url_for("project_treeview", project_id=project_id))

@app.route("/api/custom_columns/<int:column_id>/options", methods=["GET"])
def get_column_options_api(column_id):
    try:
        # تابع get_custom_column_options(column_id) از قبل موجود است 
        # و لیستی از رشته‌ها (مقادیر گزینه‌ها) را برمی‌گرداند.
        options = get_custom_column_options(column_id) 
        return jsonify({"success": True, "options": options})
    except Exception as e:
        print(f"Error fetching options for column {column_id}: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": "خطا در دریافت گزینه‌ها"}), 500

@app.route("/api/custom_columns/<int:column_id>/options/add", methods=["POST"])
def add_column_option_api(column_id):
    """افزودن گزینه جدید به ستون دراپ‌داون"""
    try:
        # بررسی و دریافت داده‌های ارسالی
        data = request.get_json()
        if not data or 'option_value' not in data:
            return jsonify({"success": False, "error": "مقدار گزینه ارسال نشده است"}), 400
        
        option_value = data['option_value']
        if not option_value.strip():
            return jsonify({"success": False, "error": "مقدار گزینه نمی‌تواند خالی باشد"}), 400
        
        # بررسی نوع ستون قبل از افزودن گزینه
        column_type = get_column_type_db(column_id)
        
        if not column_type:
            return jsonify({"success": False, "error": "ستون مورد نظر یافت نشد"}), 404
        
        if column_type != 'dropdown':
            return jsonify({"success": False, "error": "فقط می‌توان به ستون‌های دراپ‌داون گزینه اضافه کرد"}), 400
            
        # افزودن گزینه با استفاده از تابع موجود
        success = add_option_to_column(column_id, option_value)
        
        if success:
            return jsonify({"success": True, "message": "گزینه با موفقیت اضافه شد"})
        else:
            return jsonify({"success": False, "error": "خطا در افزودن گزینه"}), 500
            
    except Exception as e:
        print(f"Error adding option to column {column_id}: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": f"خطای سرور: {str(e)}"}), 500







@app.route("/api/custom_columns/options/<int:option_id>/delete", methods=["POST"])
def delete_column_option_api(option_id):
    """حذف یک گزینه از ستون دراپ‌داون براساس شناسه گزینه"""
    try:
        # بررسی وجود گزینه قبل از حذف
        column_id = get_column_id_from_option_db(option_id)
        
        if not column_id:
            return jsonify({"success": False, "error": "گزینه مورد نظر یافت نشد"}), 404
            
        # حذف گزینه با استفاده از تابع موجود
        success = delete_column_option(option_id)
        
        if success:
            return jsonify({"success": True, "message": "گزینه با موفقیت حذف شد", "column_id": column_id})
        else:
            return jsonify({"success": False, "error": "خطا در حذف گزینه"}), 500
            
    except Exception as e:
        print(f"Error deleting option {option_id}: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": f"خطای سرور: {str(e)}"}), 500



@app.route("/api/custom_columns/options/<int:option_id>/edit", methods=["POST"])
def edit_column_option_api(option_id):
    """ویرایش متن یک گزینه از ستون دراپ‌داون براساس شناسه گزینه"""
    try:
        # بررسی و دریافت داده‌های ارسالی
        data = request.get_json()
        if not data or 'new_value' not in data:
            return jsonify({"success": False, "error": "مقدار جدید گزینه ارسال نشده است"}), 400
        
        new_value = data['new_value']
        if not new_value.strip():
            return jsonify({"success": False, "error": "مقدار گزینه نمی‌تواند خالی باشد"}), 400
        
        # بررسی وجود گزینه قبل از ویرایش
        column_id = get_column_id_from_option_db(option_id)
        
        if not column_id:
            return jsonify({"success": False, "error": "گزینه مورد نظر یافت نشد"}), 404
        
        # ویرایش گزینه با استفاده از تابع موجود
        success = update_custom_column_option(option_id, new_value)
        
        if success:
            return jsonify({
                "success": True, 
                "message": "گزینه با موفقیت ویرایش شد", 
                "updated_option": {"id": option_id, "value": new_value},
                "column_id": column_id
            })
        else:
            return jsonify({"success": False, "error": "خطا در ویرایش گزینه"}), 500
            
    except Exception as e:
        print(f"Error editing option {option_id}: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": f"خطای سرور: {str(e)}"}), 500


# --- Price Calculator Constants ---
قیمت_انواع_پروفیل = {
    "فریم لس قدیمی": 1.7,
    "فریم لس قالب جدید": 1.9,
    "توچوب دار": 1.5,
    "دور آلومینیوم": 1.5,
}

قیمت_ملزومات_نصب = {
    "لاستیک": 98000,
    "بست نصب": 600000,
}

قیمت_اجرت_ماشین_کاری = {
    "چهارچوب فریم لس": 20000000,
    "داخل چوب": 40000000,
    "دور آلومینیوم": 50000000,
}

قیمت_رنگ_آلومینیوم_جدول = {
    "خام": 3450000,
    "آنادایز": 3950000,
    "رنگی": 3750000,
}

قیمت_جنس_درب = {
    "ام دی اف": 0,
    "پلای وود": 19000000,
}

قیمت_پایه_درب_خام_بر_اساس_ارتفاع = {
    "تا 260 سانتی متر": 121000000,
    "261 تا 320 سانتی متر": 133100000,
    "321 تا 360 سانتی متر": 145200000,
    "بیش از 360 سانتی متر": 145200000,
}

قیمت_خدمات_رنگ = {
    ("رنگ نهایی", "خارجی"): 27000000,
    ("رنگ نهایی", "ایرانی"): 20000000,
    ("زیر سازی", "خارجی"): 22000000,
    ("زیر سازی", "ایرانی"): 15000000,
    ("کد رنگ", "خارجی"): 33000000,
    ("کد رنگ", "ایرانی"): 25000000,
}

قیمت_یراق_آلات = {
    "لولا": 18000000,
    "قفل": 14000000,
    "سیلندر": 6800000,
}

def get_قیمت_پایه_درب_خام(height_cm):
    if height_cm <= 260:
        return قیمت_پایه_درب_خام_بر_اساس_ارتفاع["تا 260 سانتی متر"]
    elif height_cm <= 320:
        return قیمت_پایه_درب_خام_بر_اساس_ارتفاع["261 تا 320 سانتی متر"]
    elif height_cm <= 360:
        return قیمت_پایه_درب_خام_بر_اساس_ارتفاع["321 تا 360 سانتی متر"]
    else:
        return قیمت_پایه_درب_خام_بر_اساس_ارتفاع["بیش از 360 سانتی متر"]

def format_price(price):
    """Format price with thousand separators"""
    return "{:,}".format(int(price))

@app.route("/price_calculator", methods=["GET", "POST"])
def price_calculator():
    """صفحه محاسبه قیمت درب"""
    try:
        # دریافت مقادیر از دیتابیس (مربوط به تنظیمات قیمت، نه ت فرم کاربر)
        # دریافت مقادیر از دیتابیس (مربوط به تنظیمات قیمت، نه ت فرم کاربر)
        db_prices = get_price_settings_db()

        # مقادیر پیش‌فرض قیمت‌ها از تنظیمات
        prices = {
            "فریم_لس_قدیمی": db_prices.get("فریم_لس_قدیمی", 0),
            "فریم_لس_قالب_جدید": db_prices.get("فریم_لس_قالب_جدید", 0),
            "توچوب_دار": db_prices.get("توچوب_دار", 0),
            "دور_آلومینیوم": db_prices.get("دور_آلومینیوم", 0),
            "لاستیک": db_prices.get("لاستیک", 0),
            "بست_نصب": db_prices.get("بست_نصب", 0),
            "چهارچوب_فریم_لس": db_prices.get("چهارچوب_فریم_لس", 0),
            "داخل_چوب": db_prices.get("داخل_چوب", 0),
            "دور_آلومینیوم_ماشین": db_prices.get("دور_آلومینیوم_ماشین", 0),
            "خام": db_prices.get("خام", 0), # قیمت رنگ آلومینیوم
            "آنادایز": db_prices.get("آنادایز", 0), # قیمت رنگ آلومینیوم
            "رنگی": db_prices.get("رنگی", 0), # قیمت رنگ آلومینیوم (جدیداً سفید شده)
            "سفید": db_prices.get("سفید", db_prices.get("رنگی",0)), # برای سازگاری اگر "رنگی" هنوز در دیتابیس باشد
            "پلای_وود": db_prices.get("پلای_وود", 0),
            "تا_260": db_prices.get("تا_260", 0),
            "261_تا_320": db_prices.get("261_تا_320", 0),
            "321_تا_360": db_prices.get("321_تا_360", 0),
            "بیش_از_360": db_prices.get("بیش_از_360", 0),
            "رنگ_نهایی_خارجی": db_prices.get("رنگ_نهایی_خارجی", 0),
            "رنگ_نهایی_ایرانی": db_prices.get("رنگ_نهایی_ایرانی", 0),
            "زیر_سازی_خارجی": db_prices.get("زیر_سازی_خارجی", 0),
            "زیر_سازی_ایرانی": db_prices.get("زیر_سازی_ایرانی", 0),
            "کد_رنگ_خارجی": db_prices.get("کد_رنگ_خارجی", 0),
            "کد_رنگ_ایرانی": db_prices.get("کد_رنگ_ایرانی", 0),
            "لولا": db_prices.get("لولا", 0),
            "قفل": db_prices.get("قفل", 0),
            "سیلندر": db_prices.get("سیلندر", 0)
        }
        # print("DEBUG: Initialized 'prices':", prices) # Removed

        today_shamsi = jdatetime.date.today().strftime("%Y/%m/%d")

        # مقادیر پیش‌فرض برای فیلدهای اصلی فرم (برای GET request)
        initial_form_values = {
            "عرض_درب": "110",
            "ارتفاع_درب": "280",
            "measurement_unit": "cm",
            "نوع_پروفیل_فریم_لس": "فریم لس قالب جدید",
            "رنگ_آلومینیوم": "سفید", # قبلا "رنگی" بود، طبق درخواست کاربر به سفید و آنادایز محدود شد
            "جنس_درب": "ام دی اف",
            "شرایط_رنگ": "بدون رنگ",
            "رند_رنگ": "بدون رنگ",
            "نام_مشتری": "", # اطمینان از وجود کلید برای نام مشتری
            "موبایل_مشتری": "", # اطمینان از وجود کلید برای موبایل مشتری
            "تاریخ_سفارش": today_shamsi # اضافه کردن تاریخ شمسی فعلی
        }

        # گزینه‌های دراپ‌داون
        dropdown_options = {
            "نوع_پروفیل_فریم_لس": ["فریم لس قدیمی", "فریم لس قالب جدید", "توچوب دار", "دور آلومینیوم"],
            "رنگ_آلومینیوم": ["آنادایز", "سفید"], # تغییر یافته طبق درخواست کاربر
            "جنس_درب": ["ام دی اف", "پلای وود"],
            "شرایط_رنگ": ["بدون رنگ", "رنگ نهایی", "زیر سازی", "کد رنگ"],
            "رند_رنگ": ["بدون رنگ", "خارجی", "ایرانی"]
        }

        # ساختار و مقادیر پیش‌فرض اولیه برای بخش انتخاب مولفه‌ها (برای GET request)
        initial_selections_config = {
            "درب_خام": (False, 0),
            "درب_با_رنگ_کامل": (True, 30),
            "فریم": (True, 30),
            "یراق_کامل": (True, 10),
            "رنگ_کاری": (False, 0)
        }

        if request.method == "POST":
            results = None
            current_selections_for_template = {}
            component_markup_rules = {} # برای محاسبات داخلی (درصد به صورت اعشاری)

            try:
                # دریافت مقادیر اصلی فرم
                width_str = request.form.get("عرض_درب", initial_form_values["عرض_درب"])
                height_str = request.form.get("ارتفاع_درب", initial_form_values["ارتفاع_درب"])
                measurement_unit = normalize_measurement_unit(
                    request.form.get("measurement_unit", "cm")
                )
                width = dimension_to_centimeters(width_str, measurement_unit)
                height = dimension_to_centimeters(height_str, measurement_unit)
                profile_type_from_form = request.form.get("نوع_پروفیل_فریم_لس", initial_form_values["نوع_پروفیل_فریم_لس"])
                # Normalize profile_type_from_form to match the keys in the prices dictionary
                profile_type = profile_type_from_form.strip().replace(" ", "_")
                # print(f"DEBUG: profile_type_from_form: '{profile_type_from_form}'") # Removed
                # print(f"DEBUG: Normalized profile_type for prices lookup: '{profile_type}'") # Removed

                aluminum_color_from_form = request.form.get("رنگ_آلومینیوم", initial_form_values["رنگ_آلومینیوم"])
                # Normalize aluminum_color_from_form if necessary, though it seems to be working.
                # For consistency, let's normalize it as well if it might contain spaces.
                aluminum_color = aluminum_color_from_form.strip() # Assuming keys in prices don't have underscores for colors
                # print(f"DEBUG: aluminum_color_from_form: '{aluminum_color_from_form}'") # Removed
                # print(f"DEBUG: Normalized aluminum_color for prices lookup: '{aluminum_color}'") # Removed

                door_type = request.form.get("جنس_درب", initial_form_values["جنس_درب"])
                paint_type = request.form.get("شرایط_رنگ", initial_form_values["شرایط_رنگ"])
                paint_origin = request.form.get("رند_رنگ", initial_form_values["رند_رنگ"])

                # دریافت اطلاعات مشتری از فرم
                customer_name = request.form.get("نام_مشتری", "")
                customer_mobile = request.form.get("موبایل_مشتری", "")
                shamsi_order_date_from_form = request.form.get("تاریخ_سفارش", today_shamsi) # خواندن تاریخ از فرم

                # پردازش انتخاب‌ها و درصدها از فرم برای نمایش و محاسبه
                for key, (default_is_selected, default_percentage_value) in initial_selections_config.items():
                    is_selected_from_form = request.form.get(f"checkbox_{key}") == "on"
                    percentage_str_from_form = request.form.get(f"percentage_{key}")

                    template_percentage_to_use = 0.0 # برای نمایش در فرم (0-100)
                    calc_contrib_decimal_to_use = 0.0 # برای محاسبات (0.0-1.0)

                    if is_selected_from_form:
                        # اگر کاربر درصدی وارد کرده، آن را استفاده کن، در غیر این صورت از پیش‌فرض اولیه برای حالت انتخاب شده استفاده کن
                        fallback_percentage = float(default_percentage_value) # پیش‌فرض اولیه برای این مولفه
                        if percentage_str_from_form:
                            try:
                                parsed_percentage = float(percentage_str_from_form)
                                if 0 <= parsed_percentage <= 100:
                                    template_percentage_to_use = parsed_percentage
                                    calc_contrib_decimal_to_use = parsed_percentage / 100.0
                                else:
                                    flash(f"درصد برای '{key}' ({parsed_percentage}) خارج از محدوده بود. مقدار پیش‌فرض ({fallback_percentage}%) استفاده شد.", "warning")
                                    template_percentage_to_use = fallback_percentage
                                    calc_contrib_decimal_to_use = fallback_percentage / 100.0
                            except ValueError:
                                flash(f"مقدار درصد نامعتبر ('{percentage_str_from_form}') برای '{key}'. مقدار پیش‌فرض ({fallback_percentage}%) استفاده شد.", "warning")
                                template_percentage_to_use = fallback_percentage
                                calc_contrib_decimal_to_use = fallback_percentage / 100.0
                        else: # تیک خورده ولی فیلد درصد خالی است
                            flash(f"درصدی برای '{key}' وارد نشده. مقدار پیش‌فرض ({fallback_percentage}%) استفاده شد.", "warning")
                            template_percentage_to_use = fallback_percentage
                            calc_contrib_decimal_to_use = fallback_percentage / 100.0
                    # else: # اگر تیک نخورده، درصد نمایشی و محاسباتی صفر است
                        # template_percentage_to_use و calc_contrib_decimal_to_use به طور پیش‌فرض صفر هستند
                    
                    current_selections_for_template[key] = (is_selected_from_form, template_percentage_to_use)
                    component_markup_rules[key] = (is_selected_from_form, calc_contrib_decimal_to_use)

                # --- شروع محاسبات قیمت (منطق قبلی با استفاده از component_markup_rules) ---
                base_price = 0
                if height <= 260: base_price = prices["تا_260"]
                elif height <= 320: base_price = prices["261_تا_320"]
                elif height <= 360: base_price = prices["321_تا_360"]
                else: base_price = prices["بیش_از_360"]
                
                profile_weight_price = prices.get(profile_type, 0) # قیمت بر اساس وزن پروفیل انتخابی
                
                # هزینه فریم بر اساس طول و قیمت وزنی پروفیل
                # فرض می‌کنیم `prices[profile_type]` قیمت هر واحد وزن (مثلا کیلوگرم) است
                # و باید منطق محاسبه وزن کل پروفیل را داشته باشیم یا قیمت‌ها بر اساس متر باشند.
                # در اینجا فرض بر این است که قیمت‌های وارد شده در تنظیمات، قیمت نهایی به ازای هر متر یا واحد مناسب است.
                # اگر prices[profile_type] قیمت بر کیلوگرم است، باید وزن کل را محاسبه کنیم.
                # با توجه به نام فیلدها در price_calculator_settings.html مثل "فریم لس قدیمی:" (بدون واحد وزن)
                # به نظر میرسد قیمت‌های پروفیل در settings قیمت بر متر یا یک واحد دیگر است.
                # در اینجا، قیمت پروفیل را مستقیماً از prices[profile_type] میخوانیم (که در settings با وزن مشخص شده)
                # این بخش نیاز به شفاف‌سازی دارد که آیا قیمت‌های پروفیل در settings قیمت واحد وزن است یا قیمت واحد طول.
                # فعلا فرض می‌کنیم prices.get(profile_type,0) قیمت نهایی به ازای متر است.
                # اگر این قیمت وزنی است، باید وزن متر پروفیل را هم داشته باشیم.
                # با توجه به اینکه در settings کاربر وزن وارد می‌کند، prices[profile_type] باید وزن باشد.
                # پس باید قیمت واحد آلومینیوم را هم داشته باشیم.
                # این قسمت از محاسبات ممکن است نیاز به بازنگری بر اساس معنای دقیق مقادیر settings داشته باشد.
                # فعلا فرض می‌کنیم `profile_weight_price` وزن بر متر است و باید در قیمت واحد آلومینیوم ضرب شود.
                # اما در کد قبلی مستقیم ضرب میشد. پس قیمت واحد آلومینیوم در این وزن‌ها لحاظ شده.
                
                # قیمت واحد آلومینیوم بر اساس رنگ انتخابی
                aluminum_unit_price = prices.get(aluminum_color, prices.get("سفید", 0)) # اگر "آنادایز" یا "سفید" نبود، پیش‌فرض "سفید"

                total_profile_length_meters_raw = (width + (2 * height)) / 100.0 # برای اطمینان از تقسیم اعشاری
                total_profile_length_meters = ceil(total_profile_length_meters_raw) # گرد کردن به بالا
                
                # print(f"DEBUG: total_profile_length_meters_raw: {total_profile_length_meters_raw}") # Removed
                # print(f"DEBUG: total_profile_length_meters (ceil-ed): {total_profile_length_meters}") # Removed
                # print(f"DEBUG: profile_type (used for lookup): {profile_type}") # Removed
                # print(f"DEBUG: prices.get(profile_type, 0) (profile weight from prices): {prices.get(profile_type, 0)}") # Removed
                # print(f"DEBUG: aluminum_color (used for lookup): {aluminum_color}") # Removed
                # print(f"DEBUG: aluminum_unit_price (from prices): {aluminum_unit_price}") # Removed
                
                # هزینه فریم = طول کل پروفیل * وزن بر متر پروفیل انتخابی * قیمت واحد آلومینیوم بر اساس رنگ
                frame_cost = total_profile_length_meters * prices.get(profile_type, 0) * aluminum_unit_price

                rubber_cost = total_profile_length_meters * prices["لاستیک"]
                # installation_cost = prices["بست_نصب"] # Original line to be replaced
                half_bracket_unit_price = prices["بست_نصب"]          # قیمت یک «بست نصف»
                half_bracket_per_side   = ceil(height / 60)          # هر ۶۰ cm یک بست نصف
                total_half_bracket      = half_bracket_per_side * 2  # چون دو طرف در نصب می‌شود
                installation_cost       = total_half_bracket * half_bracket_unit_price
                
                machining_cost_key_map = {
                    "فریم_لس_قدیمی": "چهارچوب_فریم_لس", # کلید مقصد در prices
                    "فریم_لس_قالب_جدید": "چهارچوب_فریم_لس",
                    "توچوب_دار": "داخل_چوب",
                    "دور_آلومینیوم": "دور_آلومینیوم_ماشین" 
                }
                # profile_type از قبل نرمال شده است (مثلا "فریم_لس_قالب_جدید")
                machining_key_to_lookup_in_prices = machining_cost_key_map.get(profile_type, "چهارچوب_فریم_لس")
                machining_cost = prices.get(machining_key_to_lookup_in_prices, 0)

                paint_service_cost = 0
                if paint_type != "بدون رنگ" and paint_origin != "بدون رنگ":
                    paint_key = f"{paint_type}_{paint_origin}"
                    unit_paint_service_cost_per_sqm = prices.get(paint_key.replace(" ", "_"), 0) # e.g. رنگ_نهایی_خارجی
                    
                    # Calculate paint area (paint_area_sqm) based on door dimensions
                    # width و height اینجا باید مقادیر عددی سانتی متر باشند
                    if width > 10 and height > 6:  # برای جلوگیری از مساحت منفی
                        paint_area_sqm = ((width - 10.0) * (height - 6.0) * 2.0) / 10000.0
                    else:
                        paint_area_sqm = 0.0  # یا مقدار پیش فرض دیگر یا ایجاد خطا
                    
                    # Calculate total paint service cost
                    paint_service_cost = paint_area_sqm * unit_paint_service_cost_per_sqm
                
                # محاسبه تعداد لولا بر اساس ارتفاع درب
                height_meters = height / 100.0  # تبدیل ارتفاع به متر
                
                if height_meters <= 0:  # مدیریت ارتفاع نامعتبر
                    num_hinges = 2  # یا مقدار پیش فرض دیگر یا ایجاد خطا
                    # flash("ارتفاع درب نامعتبر است، تعداد لولا پیش‌فرض در نظر گرفته شد.", "warning")
                elif height_meters <= 1.8:
                    num_hinges = 2
                elif height_meters <= 2.1:
                    num_hinges = 3
                elif height_meters <= 2.4:
                    num_hinges = 3
                elif height_meters <= 2.7:
                    num_hinges = 4
                elif height_meters <= 3.2:
                    num_hinges = 5
                elif height_meters <= 3.6:
                    num_hinges = 6
                else:
                    # برای ارتفاع‌های بیشتر از ۳.۶ متر، یا یک مقدار ثابت در نظر بگیرید،
                    # یا بر اساس یک الگو ادامه دهید، یا خطا ایجاد کنید.
                    # فعلا فرض می کنیم برای ارتفاع بیشتر هم ۶ لولا کافی است یا باید بررسی شود.
                    num_hinges = 6  # یا مثلاً: num_hinges = 6 + math.ceil((height_meters - 3.6) / 0.5) اگر یک الگوی افزایشی دارید
                    # flash(f"ارتفاع درب ({height_meters} متر) بسیار زیاد است، تعداد لولا ({num_hinges}) بر اساس حداکثر پیش‌بینی شده در نظر گرفته شد.", "warning")
                
                # محاسبه هزینه کل یراق آلات
                # num_hinges از مرحله بالا محاسبه شده است
                price_per_hinge = prices.get("لولا", 0.0)  # قیمت هر عدد لولا از تنظیمات
                price_per_lock = prices.get("قفل", 0.0)   # قیمت پایه قفل از تنظیمات
                price_per_cylinder = prices.get("سیلندر", 0.0)  # قیمت پایه سیلندر از تنظیمات
                
                total_hinge_cost = num_hinges * price_per_hinge
                hardware_cost = total_hinge_cost + price_per_lock + price_per_cylinder
                door_material_cost = prices["پلای_وود"] if door_type == "پلای وود" else 0
                
                results = {}
                # سهم درب خام (شامل هزینه جنس درب)
                هزینه_پایه_درب_خام = base_price + door_material_cost
                is_selected_درب_خام, contrib_decimal_درب_خام = component_markup_rules["درب_خام"]
                if is_selected_درب_خام:
                    results["D14_هزینه_درب_خام_یک_درب"] = هزینه_پایه_درب_خام * (1 + contrib_decimal_درب_خام)
                else:
                    results["D14_هزینه_درب_خام_یک_درب"] = 0
                
                # سهم درب با رنگ کامل (شامل هزینه جنس درب و رنگ کاری)
                هزینه_پایه_درب_با_رنگ_کامل = base_price + door_material_cost + paint_service_cost
                is_selected_درب_با_رنگ, contrib_decimal_درب_با_رنگ = component_markup_rules["درب_با_رنگ_کامل"]
                if is_selected_درب_با_رنگ:
                    results["C11_درب_با_رنگ_کامل"] = هزینه_پایه_درب_با_رنگ_کامل * (1 + contrib_decimal_درب_با_رنگ)
                else:
                    results["C11_درب_با_رنگ_کامل"] = 0
                
                # سهم فریم
                # print(f"DEBUG: frame_cost: {frame_cost}") # Removed
                # print(f"DEBUG: rubber_cost: {rubber_cost}") # Removed
                # print(f"DEBUG: installation_cost: {installation_cost}") # Removed
                # print(f"DEBUG: machining_cost: {machining_cost}") # Removed
                هزینه_پایه_فریم = frame_cost + rubber_cost + installation_cost + machining_cost
                is_selected_فریم, contrib_decimal_فریم = component_markup_rules["فریم"]
                if is_selected_فریم:
                    results["D11_فریم"] = هزینه_پایه_فریم * (1 + contrib_decimal_فریم)
                else:
                    results["D11_فریم"] = 0
                
                # گرد کردن سهم نهایی فریم
                if results.get("D11_فریم") is not None and results["D11_فریم"] > 0: # فقط اگر مقدار مثبت و معناداری دارد
                    results["D11_فریم"] = ceil(results["D11_فریم"] / 1000000.0) * 1000000
                elif results.get("D11_فریم") is None: # اگر کلید اصلا وجود نداشت یا None بود
                     results["D11_فریم"] = 0 # یا مقدار مناسب دیگر
                # اگر صفر بود، صفر باقی می ماند

                # سهم یراق کامل
                هزینه_پایه_یراق_کامل = hardware_cost
                is_selected_یراق, contrib_decimal_یراق = component_markup_rules["یراق_کامل"]
                if is_selected_یراق:
                    results["E11_یراق_کامل"] = هزینه_پایه_یراق_کامل * (1 + contrib_decimal_یراق)
                else:
                    results["E11_یراق_کامل"] = 0
                
                # سهم رنگ کاری (فقط هزینه خدمات رنگ)
                هزینه_پایه_رنگ_کاری = paint_service_cost
                is_selected_رنگ_کاری, contrib_decimal_رنگ_کاری = component_markup_rules["رنگ_کاری"]
                if is_selected_رنگ_کاری:
                    results["رنگ_کاری_contrib"] = هزینه_پایه_رنگ_کاری * (1 + contrib_decimal_رنگ_کاری)
                else:
                    results["رنگ_کاری_contrib"] = 0
                
                # مقادیر نمایشی که حذف شده بودند، برای سازگاری و جلوگیری از خطا None یا 0 میگذاریم
                results["G14_هزینه_فریم_کل"] = frame_cost + rubber_cost + installation_cost + machining_cost # این در محاسبات اصلی استفاده نمیشود، فقط برای نمایش اگر لازم شد
                results["N14_هزینه_کل_رنگ_کاری_یک_درب"] = paint_service_cost # اینم

                results["total_cost"] = sum(filter(None, [
                    results.get("D14_هزینه_درب_خام_یک_درب"),
                    results.get("C11_درب_با_رنگ_کامل"),
                    results.get("D11_فریم"),
                    results.get("E11_یراق_کامل"),
                    results.get("رنگ_کاری_contrib")
                ]))
                # گرد کردن قیمت نهایی کل
                if results.get("total_cost") is not None and results["total_cost"] > 0: # فقط اگر مقدار مثبت و معناداری دارد
                    results["total_cost"] = ceil(results["total_cost"] / 1000000.0) * 1000000
                elif results.get("total_cost") is None:
                    results["total_cost"] = 0
                # اگر صفر بود، صفر باقی می ماند
                # --- پایان محاسبات ---

                # Check if this is an AJAX request
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    # Collect flash messages for AJAX response
                    flashed_messages = []
                    for category, message in get_flashed_messages(with_categories=True):
                        flashed_messages.append({"category": category, "message": message})
                    
                    return jsonify(success=True, results=results, flash_messages=flashed_messages)

                return render_template(
                    "price_calculator.html",
                    results=results,
                    default_values=request.form, # برای حفظ مقادیر فرم اصلی
                    dropdown_options=dropdown_options,
                    selections=current_selections_for_template # برای حفظ وضعیت چک‌باکس‌ها و درصدها
                )
                
            except ValueError as ve: # خطای تبدیل نوع مثل float()
                flash(f"خطا در مقادیر ورودی: {str(ve)}. لطفاً مقادیر عددی صحیح وارد کنید.", "error")
                traceback.print_exc()
                # در صورت خطا، فرم را با مقادیر وارد شده توسط کاربر نمایش بده
                preserved_selections_on_error = {}
                for key, (default_sel, default_perc) in initial_selections_config.items():
                    is_selected = request.form.get(f"checkbox_{key}") == "on"
                    percentage_str = request.form.get(f"percentage_{key}")
                    perc_to_display = 0.0
                    if is_selected:
                        fallback_percentage = float(default_perc)
                        if percentage_str:
                            try: perc_to_display = float(percentage_str)
                            except: perc_to_display = fallback_percentage
                            if not (0 <= perc_to_display <= 100): perc_to_display = fallback_percentage
                        else: perc_to_display = fallback_percentage
                    preserved_selections_on_error[key] = (is_selected, perc_to_display)
                
                # Check if this is an AJAX request
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    flashed_messages = []
                    for category, message in get_flashed_messages(with_categories=True):
                        flashed_messages.append({"category": category, "message": message})
                    return jsonify(success=False, error=str(ve), default_values=request.form.to_dict(), 
                                 selections=preserved_selections_on_error, flash_messages=flashed_messages), 400

                return render_template(
                    "price_calculator.html", results=None, default_values=request.form,
                    dropdown_options=dropdown_options, selections=preserved_selections_on_error
                )
            except Exception as e:
                flash(f"خطا در محاسبه قیمت: {str(e)}", "error")
                traceback.print_exc()
                # تلاش برای حفظ حالت فرم در صورت خطای عمومی
                preserved_selections_on_error = {}
                for key, (default_sel, default_perc) in initial_selections_config.items():
                    is_selected = request.form.get(f"checkbox_{key}") == "on"
                    percentage_str = request.form.get(f"percentage_{key}")
                    perc_to_display = 0.0
                    if is_selected:
                        fallback_percentage = float(default_perc)
                        if percentage_str:
                            try: perc_to_display = float(percentage_str)
                            except: perc_to_display = fallback_percentage
                            if not (0 <= perc_to_display <= 100): perc_to_display = fallback_percentage
                        else: perc_to_display = fallback_percentage
                    preserved_selections_on_error[key] = (is_selected, perc_to_display)

                # Check if this is an AJAX request
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    flashed_messages = []
                    for category, message in get_flashed_messages(with_categories=True):
                        flashed_messages.append({"category": category, "message": message})
                    return jsonify(success=False, error=str(e), flash_messages=flashed_messages), 500

                return render_template(
                    "price_calculator.html", results=None, default_values=request.form,
                    dropdown_options=dropdown_options, selections=preserved_selections_on_error
                )
        
        # GET request
        # مقادیر پیش‌فرض اولیه برای نمایش در اولین بارگذاری صفحه
        # اطمینان از اینکه selections به درستی برای قالب آماده شده (شامل مقادیر پیش‌فرض درصد)
        prepared_initial_selections = {
            key: (val[0], val[1] if val[1] is not None else 0) 
            for key, val in initial_selections_config.items()
        }

        # مقادیر پیش‌فرض فرم را ابتدا با مقادیر اولیه پر کن
        current_default_values = initial_form_values.copy() 

        # بررسی اطلاعات مشتری فلش شده از روت save_quote
        preserved_customer_data = session.pop('preserved_customer_info_data', None) # NEW way: Get from session

        if preserved_customer_data: # NEW way: Check if data exists
            if isinstance(preserved_customer_data, dict): # اطمینان از اینکه دیکشنری است
                current_default_values['نام_مشتری'] = preserved_customer_data.get('customer_name', initial_form_values.get('نام_مشتری', ''))
                current_default_values['موبایل_مشتری'] = preserved_customer_data.get('customer_mobile', initial_form_values.get('موبایل_مشتری', ''))
                # سایر فیلدهای ورودی (عرض، ارتفاع و ...) باید از initial_form_values باشند تا برای یک محاسبه جدید ریست شوند
                # فقط نام و موبایل حفظ می شوند.

        return render_template(
            "price_calculator.html",
            results=None, # برای سفارش جدید یا پس از ذخیره، نتایج باید خالی باشند
            default_values=current_default_values, # استفاده از مقادیری که ممکن است اطلاعات مشتری را حفظ کرده باشند
            dropdown_options=dropdown_options,
            selections=prepared_initial_selections # انتخاب‌های پیش‌فرض اولیه با مقادیر درصد صحیح
        )
        
    except Exception as e:
        print(f"خطای کلی در روت price_calculator: {e}")
        traceback.print_exc()
        flash("خطایی در بارگذاری صفحه محاسبه قیمت رخ داد.", "error")
        return redirect(url_for("index"))

@app.route("/price_calculator_settings", methods=["GET", "POST"])
@manager_or_admin_required
def price_calculator_settings():
    """صفحه تنظیمات قیمت پایه"""
    # print("\\n--- Initiating price_calculator_settings ---") # Removed
    try:
        if request.method == "POST":
            # print("--- Method: POST ---") # Removed
            
            # # چاپ مقادیر خام از فرم # Removed
            # print("DEBUG: Raw form data:") # Removed
            # for key in request.form: # Removed
            #     print(f"  {key}: {request.form.getlist(key)}") # استفاده از getlist برای دیدن همه مقادیر در صورت وجود کلید تکراری # Removed
            
            value_for_sefid_price = request.form.get("رنگی") 
            # print(f"DEBUG: Raw 'رنگی' value from form: {value_for_sefid_price}") # Removed

            prices_to_save = {
                "فریم_لس_قدیمی": float(request.form.get("فریم_لس_قدیمی")),
                "فریم_لس_قالب_جدید": float(request.form.get("فریم_لس_قالب_جدید")),
                "توچوب_دار": float(request.form.get("توچوب_دار")),
                "دور_آلومینیوم": float(request.form.get("دور_آلومینیوم")),
                "لاستیک": float(request.form.get("لاستیک")),
                "بست_نصب": float(request.form.get("بست_نصب")),
                "چهارچوب_فریم_لس": float(request.form.get("چهارچوب_فریم_لس")),
                "داخل_چوب": float(request.form.get("داخل_چوب")),
                "دور_آلومینیوم_ماشین": float(request.form.get("دور_آلومینیوم_ماشین")),
                "خام": float(request.form.get("خام")),
                "آنادایز": float(request.form.get("آنادایز")),
                "سفید": float(value_for_sefid_price),
                "پلای_وود": float(request.form.get("پلای_وود")),
                "تا_260": float(request.form.get("تا_260")),
                "261_تا_320": float(request.form.get("261_تا_320")),
                "321_تا_360": float(request.form.get("321_تا_360")),
                "بیش_از_360": float(request.form.get("بیش_از_360")),
                "رنگ_نهایی_خارجی": float(request.form.get("رنگ_نهایی_خارجی")),
                "رنگ_نهایی_ایرانی": float(request.form.get("رنگ_نهایی_ایرانی")),
                "زیر_سازی_خارجی": float(request.form.get("زیر_سازی_خارجی")),
                "زیر_سازی_ایرانی": float(request.form.get("زیر_سازی_ایرانی")),
                "کد_رنگ_خارجی": float(request.form.get("کد_رنگ_خارجی")),
                "کد_رنگ_ایرانی": float(request.form.get("کد_رنگ_ایرانی")),
                "لولا": float(request.form.get("لولا")),
                "قفل": float(request.form.get("قفل")),
                "سیلندر": float(request.form.get("سیلندر"))
            }
            # print(f"DEBUG: Prices to save (after _to_float): {prices_to_save}") # Removed
            
            conn = None # Initialize conn to None
            try:
                conn = get_db_connection()
                # print("DEBUG: Database connection obtained for POST.") # Removed
                cursor = conn.cursor()
                # price_settings table is created by migration 003_create_price_settings
                # print("DEBUG: 'price_settings' table ensured.") # Removed
                
                # print("DEBUG: Attempting to save to DB:") # Removed
                for key, value in prices_to_save.items():
                    # print(f"  Saving: {key} = {value} (Type: {type(value)})") # Removed
                    cursor.execute("INSERT OR REPLACE INTO price_settings (key, value) VALUES (?, ?)", (key, value))
                
                conn.commit()
                # print("DEBUG: conn.commit() executed successfully.") # Removed
            except sqlite3.Error as db_err:
                print(f"!!!!!! DATABASE ERROR during POST: {db_err}")
                traceback.print_exc()
                flash(f"خطای دیتابیس هنگام ذخیره: {db_err}", "error")
            finally:
                if conn:
                    conn.close()
                    # print("DEBUG: Database connection closed for POST.") # Removed
            
            if not flash_messages_exist(category_filter="error"): # Only flash success if no DB error occurred
                 flash("تنظیمات با موفقیت ذخیره شد.", "success")
            return redirect(url_for("price_calculator_settings"))
        
        # GET request - نمایش فرم
        # print("--- Method: GET ---") # Removed
        conn = None # Initialize conn to None
        current_prices = {}
        try:
            conn = get_db_connection()
            # print("DEBUG: Database connection obtained for GET.") # Removed
            cursor = conn.cursor()
            # price_settings table is created by migration 003_create_price_settings
            # print("DEBUG: 'price_settings' table ensured for GET.") # Removed
            cursor.execute("SELECT key, value FROM price_settings")
            rows = cursor.fetchall()
            # print(f"DEBUG: Rows fetched from DB: {len(rows)} rows") # Removed
            current_prices = {row[0]: row[1] for row in rows}
            # print(f"DEBUG: Current prices from DB: {current_prices}") # Removed
        except sqlite3.Error as db_err:
            print(f"!!!!!! DATABASE ERROR during GET: {db_err}")
            traceback.print_exc()
            flash(f"خطای دیتابیس هنگام خواندن تنظیمات: {db_err}", "error")
            # اگر در خواندن از دیتابیس خطا رخ دهد، current_prices خالی می‌ماند و مقادیر پیش‌فرض استفاده می‌شوند
        finally:
            if conn:
                conn.close()
                # print("DEBUG: Database connection closed for GET.") # Removed
        
        display_prices = {
            "فریم_لس_قدیمی": current_prices.get("فریم_لس_قدیمی", 0.0),
            "فریم_لس_قالب_جدید": current_prices.get("فریم_لس_قالب_جدید", 0.0),
            "توچوب_دار": current_prices.get("توچوب_دار", 0.0),
            "دور_آلومینیوم": current_prices.get("دور_آلومینیوم", 0.0),
            "لاستیک": current_prices.get("لاستیک", 0.0),
            "بست_نصب": current_prices.get("بست_نصب", 0.0),
            "چهارچوب_فریم_لس": current_prices.get("چهارچوب_فریم_لس", 0.0),
            "داخل_چوب": current_prices.get("داخل_چوب", 0.0),
            "دور_آلومینیوم_ماشین": current_prices.get("دور_آلومینیوم_ماشین", 0.0),
            "خام": current_prices.get("خام", 0.0),
            "آنادایز": current_prices.get("آنادایز", 0.0),
            "رنگی": current_prices.get("سفید", 0.0), 
            "پلای_وود": current_prices.get("پلای_وود", 0.0),
            "تا_260": current_prices.get("تا_260", 0.0),
            "261_تا_320": current_prices.get("261_تا_320", 0.0),
            "321_تا_360": current_prices.get("321_تا_360", 0.0),
            "بیش_از_360": current_prices.get("بیش_از_360", 0.0),
            "رنگ_نهایی_خارجی": current_prices.get("رنگ_نهایی_خارجی", 0.0),
            "رنگ_نهایی_ایرانی": current_prices.get("رنگ_نهایی_ایرانی", 0.0),
            "زیر_سازی_خارجی": current_prices.get("زیر_سازی_خارجی", 0.0),
            "زیر_سازی_ایرانی": current_prices.get("زیر_سازی_ایرانی", 0.0),
            "کد_رنگ_خارجی": current_prices.get("کد_رنگ_خارجی", 0.0),
            "کد_رنگ_ایرانی": current_prices.get("کد_رنگ_ایرانی", 0.0),
            "لولا": current_prices.get("لولا", 0.0),
            "قفل": current_prices.get("قفل", 0.0),
            "سیلندر": current_prices.get("سیلندر", 0.0),
        }
        # print(f"DEBUG: Display prices sent to template: {display_prices}") # Removed
        
        return render_template("price_calculator_settings.html", prices=display_prices)
        
    except Exception as e:
        print(f"!!!!!! خطای کلی در روت price_calculator_settings: {e}")
        traceback.print_exc()
        flash("خطایی در تنظیمات قیمت پایه رخ داد.", "error")
        return redirect(url_for("index")) # تغییر به ایندکس برای جلوگیری از حلقه احتمالی

# Helper function to check if flash messages of a certain category exist
def flash_messages_exist(category_filter=None):
    if '_flashes' in session:
        for category, message in session['_flashes']:
            if category_filter is None or category == category_filter:
                return True
    return False

@app.route("/save_quote", methods=["POST"])
def save_quote():
    if request.method == "POST":
        conn = None  # Initialize conn to None
        try:
            # Ensure the request is JSON
            if not request.is_json:
                flash("درخواست باید با فرمت JSON باشد.", "danger")
                # Return JSON error for AJAX, redirect otherwise (though AJAX is expected)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify(success=False, error="درخواست باید با فرمت JSON باشد"), 400
                return redirect(url_for('price_calculator'))

            data = request.get_json()
            if not data:
                flash("اطلاعات ارسال نشده یا فرمت نامعتبر است.", "danger")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify(success=False, error="اطلاعات ارسال نشده یا فرمت نامعتبر است."), 400
                return redirect(url_for('price_calculator'))

            # Log دریافت داده‌ها
            print(f"DEBUG: Data received in /save_quote: {data}")

            customer_name = data.get("customer_name")
            customer_mobile = data.get("customer_mobile")
            input_width = data.get("input_width")
            input_height = data.get("input_height")
            profile_type = data.get("profile_type")
            aluminum_color = data.get("aluminum_color")
            door_material = data.get("door_material")
            paint_condition = data.get("paint_condition")
            paint_brand = data.get("paint_brand")
            selections_details = data.get("selections_details") # Already a JSON string from JS
            final_price = data.get("final_price")
            shamsi_order_date = data.get("shamsi_date", "") # دریافت تاریخ شمسی از payload

            # اعتبارسنجی اولیه
            if not all([customer_name, input_width, input_height, profile_type, selections_details, final_price]):
                error_message = "اطلاعات ضروری برای ذخیره قیمت ناقص است."
                flash(error_message, "danger")
                # حفظ اطلاعات مشتری حتی در صورت خطا در سایر فیلدها
                if customer_name or customer_mobile:
                     # flash({'customer_name': customer_name, 'customer_mobile': customer_mobile}, 'preserved_customer_info') # OLD way
                     session['preserved_customer_info_data'] = {'customer_name': customer_name, 'customer_mobile': customer_mobile} # NEW way
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify(success=False, error=error_message, preserved_info={'customer_name': customer_name, 'customer_mobile': customer_mobile}), 400
                return redirect(url_for('price_calculator'))


            # فراخوانی تابع دیتابیس برای ذخیره
            data_to_save = {
                'customer_name': customer_name,
                'customer_mobile': customer_mobile,
                'input_width': input_width,
                'input_height': input_height,
                'profile_type': profile_type,
                'aluminum_color': aluminum_color,
                'door_material': door_material,
                'paint_condition': paint_condition,
                'paint_brand': paint_brand,
                'selections_details': selections_details,
                'final_price': final_price,
                'shamsi_order_date': shamsi_order_date
            }
            
            if save_quote_db(data_to_save):
                success_message = "قیمت‌دهی با موفقیت ذخیره شد."
                flash(success_message, "success")
                session['preserved_customer_info_data'] = {'customer_name': customer_name, 'customer_mobile': customer_mobile}
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify(success=True, message=success_message, preserved_info={'customer_name': customer_name, 'customer_mobile': customer_mobile})
                return redirect(url_for('price_calculator'))
            else:
                raise Exception("خطا در عملیات ذخیره در پایگاه داده")

        except Exception as e:
            print(f"Error in /save_quote: {e}")
            traceback.print_exc()
            error_message = f"خطا در ذخیره اطلاعات: {str(e)}"
            flash(error_message, "danger")

            preserved_customer_name = ""
            preserved_customer_mobile = ""
            if request.is_json:
                data_for_flash = request.get_json() or {}
                preserved_customer_name = data_for_flash.get("customer_name", "")
                preserved_customer_mobile = data_for_flash.get("customer_mobile", "")
            
            if preserved_customer_name or preserved_customer_mobile:
                session['preserved_customer_info_data'] = {'customer_name': preserved_customer_name, 'customer_mobile': preserved_customer_mobile}

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify(success=False, error=error_message, preserved_info={'customer_name': preserved_customer_name, 'customer_mobile': preserved_customer_mobile}), 500
            return redirect(url_for('price_calculator'))
    
    # اگر متد POST نبود یا خطای دیگری قبل از try رخ داد
    flash("درخواست نامعتبر برای ذخیره قیمت.", "warning")
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify(success=False, error="درخواست نامعتبر برای ذخیره قیمت."), 405 # Method Not Allowed
    return redirect(url_for('price_calculator'))

@app.route("/saved_quotes")
def saved_quotes():
    """نمایش قیمت‌دهی‌های ذخیره شده با قابلیت گروه‌بندی و باز/بسته شدن"""
    try:
        quotes = get_all_saved_quotes_db()
        
        grouped_quotes = defaultdict(list)
        for quote_data in quotes:
            quote_dict = {
                'id': quote_data['id'],
                'customer_name': quote_data['customer_name'] if quote_data['customer_name'] else "بدون نام مشتری",
                'customer_mobile': quote_data['customer_mobile'],
                'input_width': quote_data['input_width'],
                'input_height': quote_data['input_height'],
                'profile_type': quote_data['profile_type'],
                'aluminum_color': quote_data['aluminum_color'],
                'door_material': quote_data['door_material'],
                'paint_condition': quote_data['paint_condition'],
                'paint_brand': quote_data['paint_brand'],
                'final_calculated_price': quote_data['final_calculated_price'],
                'shamsi_order_date': quote_data['shamsi_order_date'] if quote_data['shamsi_order_date'] else "تاریخ نامشخص"
            }

            timestamp_val = quote_data['timestamp']
            if isinstance(timestamp_val, str):
                try:
                    timestamp_val = datetime.strptime(timestamp_val.split('.')[0], '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        timestamp_val = datetime.strptime(timestamp_val, '%Y-%m-%d %H:%M')
                    except ValueError:
                        print(f"WARNING: Could not parse timestamp string: {quote_data['timestamp']} for quote id {quote_dict['id']}. Setting to None.")
                        timestamp_val = None
            elif timestamp_val is None:
                # print(f"WARNING: Timestamp is None for quote id {quote_dict['id']}. Setting to None.")
                timestamp_val = None
            quote_dict['timestamp'] = timestamp_val
            
            try:
                if quote_data['selections_details']:
                    # Handle both JSON string and already parsed JSON (if DB driver did it, though sqlite returns str)
                    if isinstance(quote_data['selections_details'], str):
                         quote_dict['selections_details'] = json.loads(quote_data['selections_details'])
                    else:
                         quote_dict['selections_details'] = quote_data['selections_details']
                else:
                    quote_dict['selections_details'] = {}
            except json.JSONDecodeError as json_err:
                print(f"ERROR: JSONDecodeError for quote id {quote_dict['id']}: {json_err}")
                quote_dict['selections_details'] = {}
            except Exception as e_json:
                print(f"ERROR: Unknown error parsing selections_details for quote id {quote_dict['id']}: {e_json}")
                quote_dict['selections_details'] = {}
                
            customer_key = quote_dict['customer_name']
            grouped_quotes[customer_key].append(quote_dict)
        
        all_quotes_for_js = [quote for customer_quotes in grouped_quotes.values() for quote in customer_quotes]
        quotes_json_list = []
        for quote_data_dict in all_quotes_for_js:
            temp_quote = quote_data_dict.copy()
            if temp_quote.get('timestamp') and not isinstance(temp_quote['timestamp'], str):
                try:
                    temp_quote['timestamp'] = temp_quote['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
                except AttributeError: # Should not happen if previous logic is correct, but as a safeguard
                    temp_quote['timestamp'] = str(temp_quote['timestamp'])
            quotes_json_list.append(temp_quote)
        all_quotes_json = json.dumps(quotes_json_list)

        return render_template("saved_quotes.html", grouped_quotes=grouped_quotes, all_quotes_json=all_quotes_json)
        
    except Exception as e:
        print(f"!!!!!! ERROR in saved_quotes route: {e}") 
        traceback.print_exc()
        flash("خطایی در بارگذاری قیمت‌دهی‌های ذخیره شده رخ داد.", "error")
        return redirect(url_for("index"))

@app.route("/delete_quote/<int:quote_id>", methods=["POST"])
def delete_quote(quote_id):
    """پاک کردن یک قیمت‌دهی ذخیره شده"""
    try:
        # استفاده از تابع جدید
        if delete_quote_db(quote_id):
            flash("قیمت‌دهی با موفقیت پاک شد.", "success")
        else:
            flash("قیمت‌دهی مورد نظر یافت نشد یا خطا در حذف.", "error")
        return redirect(url_for("saved_quotes"))
        
    except Exception as e:
        print(f"خطا در پاک کردن قیمت‌دهی: {e}")
        traceback.print_exc()
        flash("خطایی در پاک کردن قیمت‌دهی رخ داد.", "error")
        return redirect(url_for("saved_quotes"))

@app.route("/delete_multiple_quotes", methods=["POST"])
def delete_multiple_quotes():
    """پاک کردن چندین قیمت‌دهی انتخاب شده"""
    try:
        # دریافت لیست شناسه‌های انتخاب شده
        selected_ids = request.form.getlist('selected_quotes')
        
        if not selected_ids:
            flash("هیچ قیمت‌دهی‌ای انتخاب نشده است.", "warning")
            return redirect(url_for("saved_quotes"))
        
        conn = None # To avoid UnboundLocalError in finally if used
        
        # تبدیل شناسه‌ها به لیست
        if not selected_ids:
             flash("هیچ موردی انتخاب نشده", "warning")
             return redirect(url_for("saved_quotes"))

        deleted_count = delete_multiple_quotes_db(selected_ids)
        
        if deleted_count > 0:
            flash(f"{deleted_count} قیمت‌دهی با موفقیت پاک شدند.", "success")
        else:
            flash("هیچ قیمت‌دهی‌ای پاک نشد.", "warning")
            
        return redirect(url_for("saved_quotes"))
        
    except Exception as e:
        print(f"خطا در پاک کردن قیمت‌دهی‌های انتخاب شده: {e}")
        traceback.print_exc()
        flash("خطایی در پاک کردن قیمت‌دهی‌ها رخ داد.", "error")
        return redirect(url_for("saved_quotes"))

# ============================================================================
# مدیریت بکاپ (Backup Management Routes)
# ============================================================================

def _backup_csrf_token():
    token = session.get("backup_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["backup_csrf_token"] = token
    return token


def _valid_backup_csrf():
    expected = session.get("backup_csrf_token", "")
    provided = request.form.get("csrf_token", "")
    return bool(expected and provided and hmac.compare_digest(expected, provided))


@app.route("/backup")
@admin_required
def backup_management():
    """صفحه مدیریت بکاپ"""
    try:
        backups = backup_manager.list_backups()
        stats = backup_manager.get_backup_stats()
        
        return render_template(
            "backup_manager.html",
            backups=backups,
            stats=stats,
            csrf_token=_backup_csrf_token(),
            message=session.pop('backup_message', None),
            message_type=session.pop('backup_message_type', None)
        )
    except Exception as e:
        print(f"خطا در صفحه مدیریت بکاپ: {e}")
        traceback.print_exc()
        flash("خطا در بارگذاری صفحه مدیریت بکاپ", "error")
        return redirect(url_for("index"))

@app.route("/backup/create", methods=["POST"])
@admin_required
def backup_create():
    """ایجاد بکاپ دستی"""
    try:
        if not _valid_backup_csrf():
            return "درخواست نامعتبر است.", 400
        success, result = backup_manager.create_backup(
            reason="manual_backup",
            user="admin",
            metadata={"source": "web_interface"}
        )
        
        if success:
            session['backup_message'] = f"بکاپ با موفقیت ایجاد شد."
            session['backup_message_type'] = "success"
        else:
            session['backup_message'] = f"خطا در ایجاد بکاپ: {result}"
            session['backup_message_type'] = "error"
            
    except Exception as e:
        print(f"خطا در ایجاد بکاپ: {e}")
        traceback.print_exc()
        session['backup_message'] = f"خطا در ایجاد بکاپ: {str(e)}"
        session['backup_message_type'] = "error"
    
    return redirect(url_for("backup_management"))

@app.route("/backup/restore/<filename>", methods=["POST"])
@admin_required
def backup_restore(filename):
    """بازگردانی از بکاپ"""
    try:
        if not _valid_backup_csrf():
            return "درخواست نامعتبر است.", 400
        enable_maintenance("web_restore", {"backup": filename})
        try:
            success, message = backup_manager.restore_backup(filename, create_pre_restore_backup=True)
            if success:
                init_db(allow_migrations=True)
        finally:
            disable_maintenance()
        
        if success:
            session['backup_message'] = "دیتابیس با موفقیت بازگردانی، مهاجرت و بررسی شد؛ رمز کاربران تغییر نکرد."
            session['backup_message_type'] = "success"
        else:
            session['backup_message'] = f"خطا در بازگردانی: {message}"
            session['backup_message_type'] = "error"
            
    except Exception as e:
        print(f"خطا در بازگردانی بکاپ: {e}")
        traceback.print_exc()
        session['backup_message'] = f"خطا در بازگردانی: {str(e)}"
        session['backup_message_type'] = "error"
    
    return redirect(url_for("backup_management"))

@app.route("/backup/delete/<filename>", methods=["POST"])
@admin_required
def backup_delete(filename):
    """حذف بکاپ"""
    try:
        if not _valid_backup_csrf():
            return "درخواست نامعتبر است.", 400
        success, message = backup_manager.delete_backup(filename)
        
        if success:
            session['backup_message'] = "بکاپ با موفقیت حذف شد."
            session['backup_message_type'] = "success"
        else:
            session['backup_message'] = f"خطا در حذف بکاپ: {message}"
            session['backup_message_type'] = "error"
            
    except Exception as e:
        print(f"خطا در حذف بکاپ: {e}")
        traceback.print_exc()
        session['backup_message'] = f"خطا در حذف: {str(e)}"
        session['backup_message_type'] = "error"
    
    return redirect(url_for("backup_management"))

@app.route("/backup/download/<filename>")
@admin_required
def backup_download(filename):
    """دانلود فایل بکاپ"""
    try:
        success, file_path = backup_manager.download_backup(filename)
        
        if success:
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            session['backup_message'] = file_path  # error message
            session['backup_message_type'] = "error"
            return redirect(url_for("backup_management"))
            
    except Exception as e:
        print(f"خطا در دانلود بکاپ: {e}")
        traceback.print_exc()
        session['backup_message'] = f"خطا در دانلود: {str(e)}"
        session['backup_message_type'] = "error"
        return redirect(url_for("backup_management"))

@app.route("/backup/cleanup", methods=["POST"])
@admin_required
def backup_cleanup():
    """پاکسازی بکاپ‌های قدیمی (بیشتر از 7 روز)"""
    try:
        if not _valid_backup_csrf():
            return "درخواست نامعتبر است.", 400
        deleted_count = backup_manager.cleanup_old_backups()
        
        if deleted_count > 0:
            session['backup_message'] = f"{deleted_count} بکاپ قدیمی حذف شد."
            session['backup_message_type'] = "success"
        else:
            session['backup_message'] = "هیچ بکاپ قدیمی برای حذف وجود ندارد."
            session['backup_message_type'] = "warning"
            
    except Exception as e:
        print(f"خطا در پاکسازی بکاپ‌ها: {e}")
        traceback.print_exc()
        session['backup_message'] = f"خطا در پاکسازی: {str(e)}"
        session['backup_message_type'] = "error"
    
    return redirect(url_for("backup_management"))

# افزودن کد راه‌اندازی Flask در انتهای فایل
if __name__ == "__main__":
    # Set UTF-8 encoding for Windows
    import sys
    import io
    if sys.platform == 'win32':
        # Try to set UTF-8 for stdout/stderr
        try:
            if hasattr(sys.stdout, 'buffer'):
                sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
            if hasattr(sys.stderr, 'buffer'):
                sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
        except:
            pass
    
    # Set environment variable for Werkzeug
    import os
    os.environ['PYTHONIOENCODING'] = 'utf-8'
    
    # تابع ensure_default_custom_columns() حذف شد چون مایگریشن 002 این کار را انجام می‌دهد
    # اگر این تابع قبل از مایگریشن اجرا شود، ستون‌ها را بدون column_type اضافه می‌کند
    # همه ستون‌های پیش‌فرض اکنون به درستی به عنوان dropdown تنظیم شده‌اند
    
    port = int(os.getenv("PORT", "5000"))
    app.run(debug=False, host='0.0.0.0', port=port, use_reloader=False)
